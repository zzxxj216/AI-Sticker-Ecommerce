#!/usr/bin/env python3
"""
根据 Shopify 导出的 products CSV，为每个产品调用 OpenAI 生成英文评论，
并汇总为一个 Excel（含 product_url 列，便于按链接区分产品块）。

评分、国家、日期由脚本按规则分配（近一年、以 US 为主并少量欧洲国家、全局少量 3 星）；
评论正文与 reviewer 姓名由模型生成。

用法（在项目根目录）:
  python scripts/generate_product_reviews_xlsx.py \\
    --input output/shopify_products_export_products.csv \\
    --output output/product_reviews.xlsx

环境变量:
  OPENAI_API_KEY / OPENAI_BASE_URL
  OPENAI_REVIEW_MODEL  默认 gpt-5.4（也可用 OPENAI_MODEL）
"""

from __future__ import annotations

import argparse
import json
import os
import random
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import csv

from openpyxl import Workbook
from openpyxl.styles import Font

from src.core.config import Config
from src.core.logger import get_logger
from src.services.ai.openai_service import OpenAIService

logger = get_logger("generate_product_reviews_xlsx")

# 以 US 为主；少量西欧/北欧等（不含 CA/AU）
EU_COUNTRY_POOL = (
    "UK",
    "DE",
    "FR",
    "NL",
    "IE",
    "ES",
    "IT",
    "SE",
    "AT",
    "BE",
    "PL",
    "PT",
    "DK",
    "FI",
    "NO",
    "CH",
    "GR",
    "CZ",
)

# 每条评论选国家：约 82% US，其余在欧洲池中均匀
US_WEIGHT = 0.82

# 单次 JSON 条数上限，避免超长输出截断
REVIEWS_CHUNK_SIZE = 14


def pick_country_code(rng: random.Random) -> str:
    if rng.random() < US_WEIGHT:
        return "US"
    return rng.choice(EU_COUNTRY_POOL)


def strip_html(html: str) -> str:
    text = re.sub(r"<[^>]+>", " ", html or "")
    return re.sub(r"\s+", " ", text).strip()


def build_rating_sequence(n: int, num_threes: int = 2) -> list[int]:
    """多数 4–5 星；全局约 ``num_threes`` 条 3 星（不超过 ``n``）。"""
    if n <= 0:
        return []
    threes = min(max(0, num_threes), n)
    rest = n - threes
    fives = int(rest * 0.68) if rest else 0
    fours = rest - fives
    seq = [3] * threes + [4] * fours + [5] * fives
    while len(seq) < n:
        seq.append(5)
    seq = seq[:n]
    random.shuffle(seq)
    return seq


def random_dates_in_range(n: int, start: date, end: date, rng: random.Random) -> list[str]:
    if start > end:
        start, end = end, start
    span = (end - start).days + 1
    out: list[str] = []
    for _ in range(n):
        d = start + timedelta(days=rng.randrange(span))
        out.append(d.isoformat())
    return out


def read_products_csv(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not row.get("id"):
                continue
            rows.append({k: (v or "") for k, v in row.items()})
    return rows


def llm_reviews_for_product(
    openai: OpenAIService,
    title: str,
    description: str,
    k: int,
    *,
    batch_index: int = 0,
) -> list[dict[str, str]]:
    system = (
        "You write authentic customer reviews for an e-commerce sticker shop. "
        "Lengths and styles must vary like real marketplace feedback (not uniform). "
        "Output strict JSON only."
    )
    batch_note = ""
    if batch_index > 0:
        batch_note = (
            f"\n(This is continuation batch {batch_index + 1} for the same product. "
            "Use entirely new reviewer names and different angles; do not repeat "
            "earlier phrasing. Still mix very short, medium, and longer reviews in this batch.)"
        )
    prompt = f"""Product title: {title}
Product description (plain text): {description}
{batch_note}

Generate exactly {k} different customer reviews. Each review:
- reviewer_name: realistic Western-style first + last name (two words), varied, not repetitive
- content: natural English, specific to this product when possible. No hashtags.

Length rules (important — mimic messy real reviews, do NOT make them similar length):
- Spread lengths across the batch: include some very brief ones (roughly 3–15 words, e.g. "Great quality, fast ship."), some medium (1–2 sentences), and a few longer (3–5 sentences or a short paragraph with a bit of story: use case, gift, small business, packaging).
- Avoid a "template" feel: do not start many reviews with the same opening words; vary tone (casual / enthusiastic / matter-of-fact).
- Word counts should look random: it is fine if one review is 6 words and another is 120+ words in the same batch.

Return JSON object: {{"reviews": [{{"reviewer_name": "...", "content": "..."}}, ...]}}
Must contain exactly {k} items in the array."""
    max_tokens = min(28000, max(2048, k * 220 + 900))
    data = openai.generate_json(
        prompt=prompt,
        system=system,
        temperature=0.92,
        max_tokens=max_tokens,
    )
    raw = data.get("reviews") if isinstance(data, dict) else None
    if not isinstance(raw, list) or len(raw) != k:
        raise ValueError(f"Expected {k} reviews, got {raw!r}")
    cleaned: list[dict[str, str]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        name = str(item.get("reviewer_name", "")).strip()
        content = str(item.get("content", "")).strip()
        if not name or not content:
            raise ValueError(f"Invalid review item: {item!r}")
        cleaned.append({"reviewer_name": name, "content": content})
    if len(cleaned) != k:
        raise ValueError(f"Parsed {len(cleaned)} valid reviews, need {k}")
    return cleaned


def llm_reviews_for_product_batched(
    openai: OpenAIService,
    title: str,
    description: str,
    total: int,
    chunk_size: int = REVIEWS_CHUNK_SIZE,
) -> list[dict[str, str]]:
    """分多批请求，凑满 ``total`` 条（50–70 等）。"""
    all_pairs: list[dict[str, str]] = []
    batch_index = 0
    while len(all_pairs) < total:
        n = min(chunk_size, total - len(all_pairs))
        batch = llm_reviews_for_product(
            openai, title, description, n, batch_index=batch_index
        )
        all_pairs.extend(batch)
        batch_index += 1
    return all_pairs[:total]


def process_one_product(
    args: tuple[Any, ...],
) -> tuple[str, list[dict[str, Any]] | None, str | None]:
    (
        product,
        ratings,
        dates,
        countries,
        reviews_per_product,
        model,
        base_url,
        api_key,
    ) = args
    pid = product.get("id", "")
    url = product.get("storefront_url", "")
    title = product.get("title", "")
    body = strip_html(product.get("body_html", ""))
    desc = body[:1200] if body else ""

    try:
        openai = OpenAIService(api_key=api_key, base_url=base_url or None, model=model)
        pairs = llm_reviews_for_product_batched(
            openai, title, desc, reviews_per_product
        )
    except Exception as e:
        logger.exception("Product %s failed: %s", pid, e)
        return pid, None, str(e)

    rows_out: list[dict[str, Any]] = []
    for i in range(reviews_per_product):
        rows_out.append(
            {
                "product_url": url,
                "content": pairs[i]["content"],
                "rating": ratings[i],
                "name": pairs[i]["reviewer_name"],
                "country_code": countries[i],
                "created_at": dates[i],
            }
        )
    return pid, rows_out, None


def write_xlsx(
    path: Path,
    headers: list[str],
    blocks: list[tuple[str, list[dict[str, Any]]]],
) -> None:
    """每个产品一块：先一行仅 product_url（分隔），再若干数据行。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "reviews"
    bold = Font(bold=True)
    sep_font = Font(italic=True, color="0066CC")
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = bold

    for product_url, reviews in blocks:
        ws.append([product_url] + [""] * (len(headers) - 1))
        ws.cell(row=ws.max_row, column=1).font = sep_font
        for r in reviews:
            ws.append(
                [
                    r.get("product_url", ""),
                    r.get("content", ""),
                    int(r.get("rating", 5)),
                    r.get("name", ""),
                    r.get("country_code", ""),
                    r.get("created_at", ""),
                ]
            )

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate synthetic reviews XLSX from Shopify CSV")
    parser.add_argument(
        "--input",
        "-i",
        type=Path,
        default=ROOT / "output" / "shopify_products_export_products.csv",
        help="shopify_products_export_products.csv 路径",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=ROOT / "output" / "product_reviews_generated.xlsx",
        help="输出 .xlsx 路径",
    )
    parser.add_argument("--workers", type=int, default=6, help="并发线程数")
    parser.add_argument(
        "--reviews-min",
        type=int,
        default=50,
        help="每个产品评论条数下限（含）",
    )
    parser.add_argument(
        "--reviews-max",
        type=int,
        default=70,
        help="每个产品评论条数上限（含）",
    )
    parser.add_argument(
        "--max-products",
        type=int,
        default=0,
        help="仅处理前 N 个产品，0 表示全部",
    )
    parser.add_argument("--seed", type=int, default=42, help="随机种子（评分/日期/国家）")
    args = parser.parse_args()

    Config()
    rng = random.Random(args.seed)

    inp = args.input.resolve()
    if not inp.is_file():
        print(f"Input not found: {inp}", file=sys.stderr)
        return 1

    model = os.getenv("OPENAI_REVIEW_MODEL") or os.getenv("OPENAI_MODEL") or "gpt-5.4"
    base_url = os.getenv("OPENAI_BASE_URL", "") or None
    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        print("OPENAI_API_KEY is not set.", file=sys.stderr)
        return 1

    products = read_products_csv(inp)
    if not products:
        print("No products in CSV.", file=sys.stderr)
        return 1
    if args.max_products and args.max_products > 0:
        products = products[: args.max_products]

    rmin, rmax = args.reviews_min, args.reviews_max
    if rmin < 1:
        rmin = 1
    if rmax < rmin:
        rmax = rmin

    counts_per_product = [rng.randint(rmin, rmax) for _ in products]
    total = sum(counts_per_product)
    num_threes = max(2, min(50, int(total * 0.0025 + 2))) if total >= 8 else (1 if total >= 4 else 0)
    rating_flat = build_rating_sequence(total, num_threes=num_threes)
    end_d = date.today()
    start_d = end_d - timedelta(days=365)
    date_flat = random_dates_in_range(total, start_d, end_d, rng)
    country_flat = [pick_country_code(rng) for _ in range(total)]

    work: list[tuple[Any, ...]] = []
    idx = 0
    for p, rpp in zip(products, counts_per_product, strict=True):
        ratings = rating_flat[idx : idx + rpp]
        dates = date_flat[idx : idx + rpp]
        countries = country_flat[idx : idx + rpp]
        idx += rpp
        work.append((p, ratings, dates, countries, rpp, model, base_url, api_key))

    results: dict[str, tuple[list[dict[str, Any]] | None, str | None]] = {}
    max_workers = max(1, min(args.workers, len(work)))

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = {ex.submit(process_one_product, w): w[0].get("id") for w in work}
        for fut in as_completed(futs):
            pid, rows_out, err = fut.result()
            results[pid] = (rows_out, err)

    header = ["product_url", "content", "rating", "name", "country_code", "created_at"]
    blocks: list[tuple[str, list[dict[str, Any]]]] = []
    errors: list[str] = []
    row_count = 0
    for p in products:
        pid = p.get("id", "")
        url = p.get("storefront_url", "")
        tup = results.get(pid, (None, "missing"))
        rows_out, err = tup
        if err or rows_out is None:
            errors.append(f"{pid} ({p.get('title','')[:40]}): {err or 'unknown'}")
            continue
        blocks.append((url, rows_out))
        row_count += len(rows_out)

    if errors:
        for e in errors[:20]:
            logger.warning("%s", e)
        if len(errors) > 20:
            logger.warning("... and %s more errors", len(errors) - 20)

    out_path = args.output.resolve()
    write_xlsx(out_path, header, blocks)
    print(
        json.dumps(
            {
                "ok": len(errors) == 0,
                "review_rows": row_count,
                "product_blocks": len(blocks),
                "output": str(out_path),
                "error_count": len(errors),
            },
            ensure_ascii=False,
        )
    )
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
