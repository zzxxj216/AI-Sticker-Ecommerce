"""TKShop product service — C.1 / C.2 / C.3 / C.4 of the V2 pipeline.

C.1: create product row from a pack, AI-generate detail two-step
C.2: register product images (rows + files via PackStore)
C.3: publish via the multi-channel-api FastAPI service at
     ``POST {TKSHOP_SERVER_URL}/api/v1/tiktok/products/sticker_publish``
C.4: status sync via
     ``GET  {TKSHOP_SERVER_URL}/api/v1/tiktok/products/{tiktok_product_id}``

The wrapper at multi-channel-api owns price / warehouse / weight defaults
(see its sticker_field_config). We only send the listing copy + image URLs
+ category + seller_sku + quantity. multi-channel-api fills the rest.
"""

from __future__ import annotations

import json
import os
import re
import sqlite3
import time
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

import requests
from PIL import Image

from src.core.exceptions import APIError
from src.core.logger import get_logger
from src.services.ai.router import AIRouter, get_router
from src.services.storage.pack_store import PackStore, get_pack_store
from src.services.tkshop.prompts import (
    APPEND_SECONDARY_EXTRACT_SCHEMA,
    APPEND_SECONDARY_INSTRUCTIONS,
    DETAIL_EXTRACT_INSTRUCTIONS,
    DETAIL_EXTRACT_SCHEMA,
    DETAIL_MAIN_SYSTEM_PROMPT,
    IMAGE_DESIGN_EXTRACT_SCHEMA,
    IMAGE_DESIGN_INSTRUCTIONS,
    IMAGE_DESIGN_SYSTEM_PROMPT,
    MAIN_BUNDLE_PROMPT,
    SECONDARY_HAND_HOLD_BUNDLE_BELOW_PROMPT,
    SELF_HEAL_EXTRACT_SCHEMA,
    SELF_HEAL_INSTRUCTIONS,
    SELF_HEAL_SYSTEM_PROMPT,
    build_append_secondary_prompt,
    fifth_image_style_is_hand_hold,
    hand_hold_secondary_spec,
    build_detail_main_prompt,
    build_image_design_prompt,
    build_self_heal_prompt,
)

logger = get_logger("service.tkshop")

DEFAULT_DB_PATH = Path("data/ops_workbench.db")

# Local publish_status vocabulary → Chinese label + pill class.
# Single source of truth for both UI and API responses.
# pill class refers to .v2-status-pill modifier in v2_base.html (ok/error/pending/running).
STATUS_LABELS_ZH: dict[str, tuple[str, str]] = {
    "draft":              ("本地草稿（未上传）",     "pending"),
    "publishing":         ("上传中…",                "running"),
    "failed":             ("上传失败",               "error"),
    "draft_on_platform":  ("平台草稿（买家不可见）", "pending"),
    "pending":            ("平台审核中",             "running"),
    "published":          ("已上架销售中",           "ok"),
    # Aliases for raw TikTok status passthroughs that older rows may still
    # carry. New writes go through _local_status_from_tiktok() which folds
    # them into 'published'.
    "activate":           ("已上架销售中",           "ok"),
    "active":             ("已上架销售中",           "ok"),
    "live":               ("已上架销售中",           "ok"),
    "seller_deactivated": ("已下架（可恢复）",       "pending"),
    # Legacy rows that came from before sync-status normalization. New code
    # writes 'seller_deactivated' instead — relabel for consistency.
    "suspended":          ("已下架（可恢复）",       "pending"),
    "deleted":            ("平台已删除",             "error"),
    "unknown":            ("未知状态",               "error"),
}


def status_label_zh(status: str) -> str:
    return STATUS_LABELS_ZH.get(status or "", (status or "—", "pending"))[0]


def status_pill_cls(status: str) -> str:
    return STATUS_LABELS_ZH.get(status or "", ("", "pending"))[1]


def _format_search_keywords(keywords: list[str]) -> str:
    """TikTok Seller Center「后台关键词」— comma-separated, no hashtags."""
    parts = [k.strip().lstrip("#") for k in keywords if (k or "").strip()]
    return ", ".join(parts)[:500]


def _sku_meta_from_platform_data(data: dict) -> dict[str, Any]:
    """Extract price/stock from multi-channel-api product GET payload."""
    skus = data.get("skus") or []
    sku = skus[0] if skus else {}
    price_obj = sku.get("price") or {}
    sale = str(
        price_obj.get("sale_price") or price_obj.get("amount") or ""
    ).strip()
    currency = str(price_obj.get("currency") or "USD").strip()
    inv = sku.get("inventory") or []
    stock = inv[0].get("quantity") if inv else None
    # Promotional / strikethrough price when TikTok exposes it.
    discount = str(
        price_obj.get("original_price")
        or price_obj.get("list_price")
        or ""
    ).strip()
    return {
        "sale_price": sale,
        "discount_price": discount,
        "currency": currency,
        "stock": stock,
    }


def _price_from_default_template(raw: str) -> tuple[str, str]:
    """Best-effort read sale/discount price from default_template_json."""
    try:
        tpl = json.loads(raw or "{}")
    except Exception:
        return "", ""
    if not isinstance(tpl, dict):
        return "", ""
    skus = tpl.get("skus") or []
    sku = skus[0] if skus else tpl
    price = sku.get("price") if isinstance(sku, dict) else {}
    if not isinstance(price, dict):
        price = tpl.get("price") if isinstance(tpl.get("price"), dict) else {}
    sale = str(price.get("amount") or price.get("sale_price") or "").strip()
    discount = str(
        price.get("original_price") or price.get("list_price") or ""
    ).strip()
    return sale, discount

# multi-channel-api FastAPI service. Defaults to localhost:8000 because
# both processes run on the same box during development.
TKSHOP_SERVER_URL = os.getenv("TKSHOP_SERVER_URL", "http://localhost:8000")
# 300s default: sticker_publish does N image fetches + N TikTok uploads +
# create + activate, easily exceeds 60s for 4-6 images. Override via env.
TKSHOP_SERVER_TIMEOUT = int(os.getenv("TKSHOP_SERVER_TIMEOUT", "300"))

# Public URL the multi-channel-api server uses to fetch our local product
# images. It must be reachable from that server's perspective. Both run
# on localhost so the existing /v2-outputs static mount works.
TKSHOP_IMAGE_PUBLIC_BASE = os.getenv(
    "TKSHOP_IMAGE_PUBLIC_BASE", "http://localhost:5000"
).rstrip("/")

# Multi-shop support (matches multi-channel-api's TIKTOK_SHOPS_JSON registry).
# Comma-separated canonical shop names; the first one is the default for new
# rows and UI dropdown selection.
TKSHOP_SHOPS = [
    s.strip()
    for s in (
        os.getenv("TKSHOP_SHOPS") or "inkelligentsticker,inkelligentstudio"
    ).split(",")
    if s.strip()
]
# Default commerce fields for export / operator spreadsheets (TikTok Seller
# Center bulk edit). Live listings override sale_price/stock via platform GET.
TKSHOP_DEFAULT_SALE_PRICE = os.getenv("TKSHOP_DEFAULT_SALE_PRICE", "13.98")
TKSHOP_DEFAULT_DISCOUNT_PRICE = (os.getenv("TKSHOP_DEFAULT_DISCOUNT_PRICE") or "6.99").strip()
TKSHOP_DEFAULT_QUANTITY = os.getenv("TKSHOP_DEFAULT_QUANTITY", "16")

# Auto discount: after a successful publish, ask the middle layer to put the
# product on a N% discount. 0 (or negative) disables it. Default 50% off
# $13.98 list → $6.99 buyer price. The middle layer turns this into a TikTok
# Promotion activity (see docs/tiktok_product_shipping_discount_contract.md).
try:
    TKSHOP_AUTO_DISCOUNT_PERCENT = float(os.getenv("TKSHOP_AUTO_DISCOUNT_PERCENT", "50") or 0)
except (TypeError, ValueError):
    TKSHOP_AUTO_DISCOUNT_PERCENT = 50.0

# Discount activity window (days). TikTok promotions need begin/end times and
# CAP the period at 90 days (error 17029007), so default to 89. The middle
# layer also clamps to 90 defensively. Activity type is passed through.
try:
    TKSHOP_DISCOUNT_WINDOW_DAYS = int(os.getenv("TKSHOP_DISCOUNT_WINDOW_DAYS", "89"))
except (TypeError, ValueError):
    TKSHOP_DISCOUNT_WINDOW_DAYS = 89
TKSHOP_DISCOUNT_ACTIVITY_TYPE = os.getenv("TKSHOP_DISCOUNT_ACTIVITY_TYPE", "DIRECT_DISCOUNT")

# Free shipping is handled at the SHOP level by a default free-shipping freight
# template (set once in Seller Center → applies to every product automatically),
# so per-product automation is NOT needed and is OFF by default. The
# SHIPPING_DISCOUNT promotion path (apply_free_shipping) stays available for
# targeted free-shipping promos — opt in with TKSHOP_DEFAULT_FREE_SHIPPING=1.
_fs = (os.getenv("TKSHOP_DEFAULT_FREE_SHIPPING") or "0").strip().lower()
TKSHOP_DEFAULT_FREE_SHIPPING = _fs in ("1", "true", "yes", "on")

# Per-shop seller_sku prefix. Format: ``shopA:INK1,shopB:INK2``. When a shop
# isn't listed, prefix derives from the shop's index in TKSHOP_SHOPS:
# first shop → INK1, second → INK2, etc. This keeps SKUs on different shops
# from colliding when the same pack gets cloned across shops.
def _parse_shop_sku_prefixes(raw: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for chunk in (raw or "").split(","):
        chunk = chunk.strip()
        if not chunk or ":" not in chunk:
            continue
        shop, pfx = chunk.split(":", 1)
        shop = shop.strip()
        pfx = re.sub(r"[^A-Z0-9]", "", (pfx or "").strip().upper())
        if shop and pfx:
            out[shop] = pfx
    return out


_TKSHOP_SKU_PREFIX_OVERRIDES = _parse_shop_sku_prefixes(
    os.getenv("TKSHOP_SKU_PREFIXES", "")
)


def get_shop_sku_prefix(shop: str) -> str:
    """Return the SKU prefix for ``shop`` (e.g. ``INK1`` / ``INK2``).

    Resolution order: explicit override in ``TKSHOP_SKU_PREFIXES`` env →
    index in ``TKSHOP_SHOPS`` (1-based, so the first shop → ``INK1``) →
    fallback ``INK`` for unknown shops.
    """
    s = canonical_shop_key(shop or TKSHOP_DEFAULT_SHOP)
    if s in _TKSHOP_SKU_PREFIX_OVERRIDES:
        return _TKSHOP_SKU_PREFIX_OVERRIDES[s]
    legacy = (shop or "").strip()
    if legacy in _TKSHOP_SKU_PREFIX_OVERRIDES:
        return _TKSHOP_SKU_PREFIX_OVERRIDES[legacy]
    try:
        idx = TKSHOP_SHOPS.index(s) + 1
        return f"INK{idx}"
    except ValueError:
        return "INK"


# Optional per-key label overrides via ``TKSHOP_SHOP_LABELS=main:2店·…`` env.
# Default labels come from ``_SHOP_UI_LABELS`` (see get_shop_label).
def _parse_shop_labels(raw: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for chunk in (raw or "").split(","):
        chunk = chunk.strip()
        if not chunk or ":" not in chunk:
            continue
        shop, label = chunk.split(":", 1)
        shop = shop.strip()
        label = label.strip()
        if shop and label:
            out[shop] = label
    return out


_TKSHOP_SHOP_LABEL_OVERRIDES = _parse_shop_labels(
    os.getenv("TKSHOP_SHOP_LABELS", "")
)


# ---------------------------------------------------------------------------
# Shop registry — canonical keys vs legacy aliases vs UI labels
# ---------------------------------------------------------------------------
# Canonical DB / env keys:
#   inkelligentsticker  → 2店 · inkelligentsticker (SKU prefix INK / INK1)
#   inkelligentstudio   → 3店 · inkelligentstudio  (SKU prefix INK2)
# Legacy keys (read-only alias, never write new rows):
#   main   → inkelligentsticker
#   second → inkelligentstudio
# ---------------------------------------------------------------------------

_LEGACY_SHOP_ALIASES: dict[str, str] = {
    "main": "inkelligentsticker",
    "second": "inkelligentstudio",
}


def canonical_shop_key(shop: str) -> str:
    """Normalize legacy registry keys to the canonical shop name."""
    s = (shop or "").strip()
    if not s:
        return "inkelligentsticker"
    return _LEGACY_SHOP_ALIASES.get(s, s)


def normalize_tkshop_shop_list(shops: list[str]) -> list[str]:
    """Dedupe after canonicalizing legacy keys."""
    out: list[str] = []
    for raw in shops:
        key = canonical_shop_key(raw)
        if key and key not in out:
            out.append(key)
    return out


TKSHOP_SHOPS = normalize_tkshop_shop_list(TKSHOP_SHOPS)
TKSHOP_DEFAULT_SHOP = TKSHOP_SHOPS[0] if TKSHOP_SHOPS else "inkelligentsticker"

_SHOP_UI_LABELS: dict[str, tuple[str, str]] = {
    # (店号, 店铺名)
    "inkelligentsticker": ("2店", "inkelligentsticker"),
    "inkelligentstudio": ("3店", "inkelligentstudio"),
    # Legacy read-only aliases — same stores, never shown in UI as raw keys.
    "main": ("2店", "inkelligentsticker"),
    "second": ("3店", "inkelligentstudio"),
}


def get_shop_label(shop: str) -> str:
    """Human-facing shop name for templates. Never returns ``main`` / ``second``."""
    s = (shop or "").strip()
    if not s:
        return ""
    if s in _TKSHOP_SHOP_LABEL_OVERRIDES:
        return _TKSHOP_SHOP_LABEL_OVERRIDES[s]
    if s in _SHOP_UI_LABELS:
        num, name = _SHOP_UI_LABELS[s]
        return f"{num} · {name}"
    return s


def get_shop_store_name(shop: str) -> str:
    """Store name only (no 店号), e.g. ``inkelligentsticker``."""
    s = (shop or "").strip()
    if s in _SHOP_UI_LABELS:
        return _SHOP_UI_LABELS[s][1]
    if s in _TKSHOP_SHOP_LABEL_OVERRIDES:
        return _TKSHOP_SHOP_LABEL_OVERRIDES[s]
    if s == "main":
        return "inkelligentsticker"
    return s


def get_shop_number(shop: str) -> str:
    """Return ``2店`` / ``3店`` or empty for unknown keys."""
    s = (shop or "").strip()
    if s in _SHOP_UI_LABELS:
        return _SHOP_UI_LABELS[s][0]
    return ""


def is_legacy_shop_key(shop: str) -> bool:
    """True for registry keys that should not appear as new listing targets."""
    return (shop or "").strip() in _LEGACY_SHOP_ALIASES


def get_shop_labels_map() -> dict[str, str]:
    """All configured shops with their display labels."""
    return {s: get_shop_label(s) for s in TKSHOP_SHOPS}


# Size whitelist accepted by JieKou's gpt-image-2 text-to-image / edit endpoints
# (from its 400 VALIDATION_ERROR schema). Anything outside this set is rejected,
# so callers must snap to a member before sending.
_JIEKOU_IMAGE_SIZES = {
    "1024x1024", "1024x1536", "1536x1024",
    "688x2048", "880x2048", "1024x2048", "1152x2048", "1360x2048", "1536x2048", "2048x2048",
    "2048x688", "2048x880", "2048x1024", "2048x1152", "2048x1360", "2048x1536",
}


def _get_product_shop(db_path: Path, product_id: int) -> str:
    """Return the shop name a product is bound to. Falls back to the default
    shop if the row has an empty shop value (legacy rows pre-migration get
    'main' from the column DEFAULT)."""
    with _open_db(db_path) as conn:
        row = conn.execute(
            "SELECT shop FROM tkshop_products WHERE id = ?",
            (product_id,),
        ).fetchone()
    if not row:
        return TKSHOP_DEFAULT_SHOP
    return canonical_shop_key(row["shop"] or TKSHOP_DEFAULT_SHOP)

# Self-heal toggles
TKSHOP_PUBLISH_AUTO_FIX_DEFAULT = (
    os.getenv("TKSHOP_PUBLISH_AUTO_FIX", "true").strip().lower()
    in ("1", "true", "yes", "on")
)
TKSHOP_PUBLISH_MAX_RETRIES_DEFAULT = int(
    os.getenv("TKSHOP_PUBLISH_MAX_RETRIES", "3")
)


def _open_db(db_path: Path = DEFAULT_DB_PATH) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path), timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_SKU_VALID_RE = re.compile(r"^[A-Z0-9-]{1,25}$")

# TikTok Shop hard limit on product title.
TKSHOP_TITLE_MAX = 255


def _clamp_title(title: str) -> str:
    """Trim title to TKSHOP_TITLE_MAX. Strip trailing separators so we
    don't end on a dangling ``|`` or comma.
    """
    t = (title or "").strip()
    if len(t) <= TKSHOP_TITLE_MAX:
        return t
    return t[:TKSHOP_TITLE_MAX].rstrip(" |,-—·•")


def _friendly_platform_error(action: str, raw: str) -> str:
    """Map opaque TikTok error text into a Chinese hint for the operator."""
    raw_low = (raw or "").lower()
    expected_state = {
        "activate":   "FAILED / INACTIVE / DRAFT (审核通过后)",
        "deactivate": "LIVE",
        "recover":    "DELETED",
    }.get(action, "—")
    # 12052901 = product status invalid for this action
    if "12052901" in raw or "product status invalid" in raw_low or "status invalid" in raw_low:
        return (
            f"商品当前状态不允许「{action}」操作（TikTok 要求状态: {expected_state}）。"
            f"请先点🔄 同步平台状态查看真实状态 — 例如已是 DRAFT/PENDING 不能下架，已是 SUSPENDED 不能再下架。"
            f"\n原始错误: {raw[:200]}"
        )
    if "11017002" in raw or "category" in raw_low and "audit" in raw_low:
        return f"分类资质审核未通过 — TikTok 卖家中心查看商品审核状态。原始: {raw[:200]}"
    if "12017004" in raw or "permission" in raw_low:
        return f"无权限执行此操作 — token scope 可能缺失。原始: {raw[:200]}"
    return f"TikTok 拒绝: {raw[:300]}"


def _slugify_for_sku(pack_uid: str) -> str:
    """Uppercased alphanumeric slug from the NAME part of a pack_uid.

    ``pack_uid`` is ``{YYYYMMDD}_{name}_{4hex}``. We strip the leading date
    block and the trailing hex token first, so the slug reflects the pack
    *name* rather than the (shared) date prefix. Otherwise every pack created
    on the same day collapses to the same date-dominated first 10 chars and
    their default SKUs collide.
    """
    s = re.sub(r"^\d{8}_", "", pack_uid or "")        # drop leading YYYYMMDD_
    s = re.sub(r"_[0-9a-fA-F]{4}$", "", s)            # drop trailing _xxxx hex
    return re.sub(r"[^A-Za-z0-9]", "", s).upper()[:12]


def compute_default_seller_sku(
    *, pack_uid: str, total_stickers: int = 0, pack_id: int,
    shop: Optional[str] = None,
) -> str:
    """Default seller_sku = ``INK-{name_slug}-{pack_id}`` — cross-shop unified.

    Historically this prepended a per-shop prefix (``INK1``/``INK2``/…) so the
    same pack on two stores got distinct SKUs. The operator now wants the
    SAME SKU on every shop for the same pack, mirroring the master catalog's
    SKU. Pre-existing rows keep their old prefixed SKUs unchanged; only new
    rows get the unified form. Cross-shop uniqueness is enforced per-shop by
    ``_ensure_unique_seller_sku(conn, sku, shop=...)``, so two shops can
    legitimately share ``INK-COASTAL-7`` without colliding.

    ``shop`` and ``total_stickers`` are accepted for backward compat with
    existing call sites but no longer encoded.
    """
    slug = _slugify_for_sku(pack_uid) or f"P{pack_id}"
    sku = f"INK-{slug}-{pack_id}"
    return sku[:25]


def _ensure_unique_seller_sku(
    conn: sqlite3.Connection, sku: str, *,
    exclude_product_id: Optional[int] = None,
    shop: Optional[str] = None,
) -> str:
    """Return a SKU guaranteed not to collide with any existing row.

    Checks ``tkshop_products.seller_sku`` (case-insensitive) and, on
    collision, appends ``-2``, ``-3``, … until free. Truncates to 25 chars
    while leaving room for the disambiguator. ``exclude_product_id`` lets
    an UPDATE keep its own existing SKU.

    ``shop`` scopes the uniqueness check to that shop only — needed since the
    operator now wants the same pack to share one SKU across shops, so two
    shops legitimately holding ``INK-COASTAL-7`` is no longer a collision.
    When ``shop`` is None the check is global (master-level callers, legacy).
    """
    base = (sku or "").strip().upper()[:25]
    if not base:
        return base
    candidate = base
    n = 1
    while True:
        sql = "SELECT id FROM tkshop_products WHERE UPPER(seller_sku) = ?"
        params: list = [candidate]
        if exclude_product_id is not None:
            sql += " AND id != ?"
            params.append(exclude_product_id)
        if shop is not None:
            sql += " AND shop = ?"
            params.append(shop)
        if conn.execute(sql, tuple(params)).fetchone() is None:
            return candidate
        n += 1
        suffix = f"-{n}"
        head_len = 25 - len(suffix)
        candidate = (base[:head_len] + suffix)


def _is_valid_seller_sku(sku: str) -> bool:
    return bool(sku) and bool(_SKU_VALID_RE.match(sku))


_MOJIBAKE_MARKERS = ("�", "��", "锛", "涓", "鍙", "鈥", "馃", "浜", "鎻", "è", "ɳ")


def _mojibake_score(s: str) -> int:
    if not s:
        return 0
    return sum(s.count(m) for m in _MOJIBAKE_MARKERS)


def _ascii_alpha_count(s: str) -> int:
    return sum(1 for c in (s or "") if c.isascii() and c.isalpha())


def _clean_listing_context_text(s: str, *, max_len: int = 300) -> str:
    """Keep model context buyer-useful when upstream bilingual fields contain
    mojibake. Prefer the English side of slash-delimited names and drop noisy
    replacement/control characters without trying to recover corrupted text.
    """
    s = (s or "").replace("\r", " ").replace("\n", " ").strip()
    s = s.translate(str.maketrans({
        "’": "'", "‘": "'", "“": '"', "”": '"',
        "—": "-", "–": "-", "·": " ", "•": " ",
    }))
    if "/" in s:
        parts = [p.strip(" -|") for p in s.split("/") if p.strip(" -|")]
        if parts:
            s = max(parts, key=lambda p: (_ascii_alpha_count(p), -_mojibake_score(p)))
    s = re.sub(r"[^\x09\x0A\x0D\x20-\x7E]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" -|")
    for src, dst in {
        "Father s": "Father's",
        "Mother s": "Mother's",
        "Valentine s": "Valentine's",
        "World s": "World's",
        "New Year s": "New Year's",
        "Dad s": "Dad's",
        "Mom s": "Mom's",
    }.items():
        s = s.replace(src, dst)
    return s[:max_len]


def _clean_brief_text(s: str, *, max_len: int = 120) -> str:
    s = _clean_listing_context_text(s, max_len=max_len)
    if len(s) < 3 or _mojibake_score(s) > 0:
        return ""
    return s


def _normalize_disk_path(path: str) -> str:
    return (path or "").replace("\\", "/").strip()


def _push_sync_baseline(*, last_success_push_at: int, published_at: int) -> int:
    """Last moment TikTok is known to reflect local master content."""
    return max(int(last_success_push_at or 0), int(published_at or 0))


def _listing_needs_platform_push(
    *,
    master_updated_at: int,
    tiktok_product_id: str,
    last_success_push_at: int,
    published_at: int,
) -> bool:
    if not (tiktok_product_id or "").strip():
        return False
    return int(master_updated_at or 0) > _push_sync_baseline(
        last_success_push_at=last_success_push_at,
        published_at=published_at,
    )


def _fetch_last_success_push_by_product(
    conn: sqlite3.Connection, product_ids: list[int],
) -> dict[int, int]:
    if not product_ids:
        return {}
    ph = ",".join(["?"] * len(product_ids))
    rows = conn.execute(
        f"""
        SELECT product_id, MAX(created_at) AS last_push
          FROM tkshop_publish_logs
         WHERE success = 1 AND product_id IN ({ph})
         GROUP BY product_id
        """,
        tuple(product_ids),
    ).fetchall()
    return {int(r["product_id"]): int(r["last_push"] or 0) for r in rows}


def _resolve_disk_path(path: str, *, root: Optional[Path] = None) -> Path:
    """Resolve a project-relative asset path; tolerate Windows-style separators."""
    norm = _normalize_disk_path(path)
    if not norm:
        return Path()
    p = Path(norm)
    if p.is_file():
        return p
    if root is None:
        root = DEFAULT_DB_PATH.resolve().parent.parent
    candidate = root / norm
    return candidate if candidate.is_file() else Path()


def _local_path_to_public_url(local_path: str) -> str:
    """Turn ``output/packs/<uid>/...`` into the public ``/v2-outputs/...`` URL.

    The multi-channel-api server is running on the same host (localhost),
    so it hits our FastAPI app's static mount at ``/v2-outputs/`` to fetch
    the image. If the path doesn't contain the canonical ``output/packs/``
    marker we return an empty string (caller should skip).
    """
    if not local_path:
        return ""
    p = local_path.replace("\\", "/")
    marker = "output/packs/"
    idx = p.find(marker)
    if idx < 0:
        return ""
    rel = p[idx + len(marker):]
    return f"{TKSHOP_IMAGE_PUBLIC_BASE}/v2-outputs/{rel}"


def _stage_path_for_mca(local_path: str) -> str:
    """Copy to a no-space temp path so mca can read via image_paths (not URL)."""
    import hashlib
    import shutil
    import tempfile

    abs_path = local_path if os.path.isabs(local_path) else os.path.abspath(local_path)
    if not os.path.isfile(abs_path) or not abs_path.isascii():
        return ""
    if " " not in abs_path:
        return abs_path
    digest = hashlib.md5(abs_path.encode()).hexdigest()[:12]
    ext = os.path.splitext(abs_path)[1] or ".png"
    stage_root = os.path.join(tempfile.gettempdir(), "tkshop_push_staging")
    os.makedirs(stage_root, exist_ok=True)
    staged = os.path.join(stage_root, f"{digest}{ext}")
    if not os.path.isfile(staged) or os.path.getmtime(staged) < os.path.getmtime(abs_path):
        shutil.copy2(abs_path, staged)
    return staged


class TKShopService:
    def __init__(
        self,
        router: Optional[AIRouter] = None,
        store: Optional[PackStore] = None,
        db_path: Path = DEFAULT_DB_PATH,
    ) -> None:
        self.router = router or get_router()
        self.store = store or get_pack_store()
        self.db_path = db_path

    # ------------------------------------------------------------------
    # C.1 create + AI generate
    # ------------------------------------------------------------------

    def create_product_from_pack(self, pack_id: int,
                                  shop: Optional[str] = None) -> int:
        """Mint a draft tkshop_products row for this pack. Idempotent for
        the (pack_id, shop) pair — if a product already exists for this pack
        on the chosen shop, return its id (so re-clicking "create" doesn't
        spawn duplicates). Different shops still get their own rows.

        Also ensures a ``local_products`` master row exists for the pack and
        links the new listing to it. The master is the single source of
        truth for title/description/images; listings carry shop-specific
        platform IDs and a snapshot of what was last pushed.
        """
        # Always store the canonical shop key so a legacy alias (main/second)
        # and its canonical name (inkelligentsticker/inkelligentstudio) are
        # never treated as two different shops — that mismatch is what used to
        # spawn a duplicate listing for the same physical store.
        target_shop = canonical_shop_key(shop or TKSHOP_DEFAULT_SHOP)
        # Master is mandatory — created first so the listing FK can point at
        # it from row insert (avoids a second UPDATE).
        local_product_id = self.get_or_create_local_product(pack_id)
        with _open_db(self.db_path) as conn:
            # Idempotency is canonical-shop-aware: an existing row stored under
            # a legacy alias counts as the same shop, so we reuse it instead of
            # minting a duplicate.
            existing = conn.execute(
                "SELECT id, shop FROM tkshop_products WHERE pack_id = ? "
                "ORDER BY id DESC",
                (pack_id,),
            ).fetchall()
            for row in existing:
                if canonical_shop_key(row["shop"]) == target_shop:
                    return row["id"]
            now = int(time.time())
            # Snapshot master fields into the listing — if master is still
            # empty (operator hasn't run AI gen yet) the listing inherits the
            # empties, which is the same as the old behavior.
            master = conn.execute(
                "SELECT title, description_html, selling_points, keywords, "
                "       detail_main_raw_text, category_id, default_template_json, "
                "       seller_sku "
                "FROM local_products WHERE id = ?",
                (local_product_id,),
            ).fetchone()
            cur = conn.execute(
                """
                INSERT INTO tkshop_products
                    (pack_id, shop, tiktok_product_id, detail_main_raw_text,
                     title, description_html, selling_points, keywords,
                     seller_sku, category_id, default_template_json,
                     publish_status, created_at, published_at, local_product_id)
                VALUES (?, ?, '', ?, ?, ?, ?, ?, ?, ?, ?, 'draft', ?, NULL, ?)
                """,
                (
                    pack_id, target_shop,
                    master["detail_main_raw_text"] or "",
                    master["title"] or "",
                    master["description_html"] or "",
                    master["selling_points"] or "[]",
                    master["keywords"] or "[]",
                    # Inherit the master SKU so every shop's listing shares one
                    # SKU. Empty master SKU falls back at publish time.
                    master["seller_sku"] or "",
                    master["category_id"] or "928016",
                    master["default_template_json"] or "{}",
                    now, local_product_id,
                ),
            )
            conn.commit()
            return cur.lastrowid

    def generate_detail(
        self,
        product_id: int,
        *,
        main_model: Optional[str] = None,
        extract_model: Optional[str] = None,
    ) -> dict[str, Any]:
        # Lock product copywriting to gpt-5.4 unless operator overrides
        # via TKSHOP_DETAIL_MODEL. The router's text_complete default
        # follows OPENAI_MODEL which can drift; pin it here.
        if not main_model:
            main_model = os.getenv("TKSHOP_DETAIL_MODEL", "gpt-5.4")
        if not extract_model:
            extract_model = os.getenv("TKSHOP_EXTRACT_MODEL", "gpt-5.4")
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                """
                SELECT pr.id, pr.pack_id, pr.shop, pr.seller_sku AS existing_seller_sku,
                       p.display_name, p.total_stickers, p.pack_uid,
                       s.style_anchor, s.palette, s.pack_archetype,
                       s.metadata_json
                  FROM tkshop_products pr
                  JOIN packs           p ON p.id = pr.pack_id
             LEFT JOIN pack_series     s ON s.id = p.series_id
                 WHERE pr.id = ?
                """,
                (product_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"tkshop_product #{product_id} not found")

        # Pull a representative, de-duplicated sample of sticker briefs from
        # the series metadata. A larger, cleaner sample makes the copy less
        # template-like and avoids leaking mojibake from older bilingual data.
        briefs_sample: list[str] = []
        try:
            md = json.loads(row["metadata_json"] or "{}")
            seen: set[str] = set()
            for pb in md.get("preview_briefs", [])[:8]:
                for raw in (pb.get("stickers", []) or [])[:8]:
                    brief = _clean_brief_text(str(raw))
                    key = brief.lower()
                    if brief and key not in seen:
                        briefs_sample.append(brief)
                        seen.add(key)
                    if len(briefs_sample) >= 18:
                        break
                if len(briefs_sample) >= 18:
                    break
        except Exception:
            pass

        # Pre-compute the default seller_sku so the AI can mirror it back.
        # Per-shop prefix keeps cross-shop SKUs from colliding (e.g. main →
        # INK1-…, inkelligentstudio → INK2-…).
        default_sku = compute_default_seller_sku(
            pack_uid=row["pack_uid"] or "",
            total_stickers=row["total_stickers"] or 0,
            pack_id=row["pack_id"],
            shop=row["shop"] or TKSHOP_DEFAULT_SHOP,
        )

        prompt = build_detail_main_prompt(
            pack_display_name=_clean_listing_context_text(row["display_name"] or "", max_len=120),
            pack_archetype=_clean_listing_context_text(row["pack_archetype"] or "", max_len=80),
            style_anchor=_clean_listing_context_text(row["style_anchor"] or "", max_len=420),
            palette=_clean_listing_context_text(row["palette"] or "", max_len=140),
            total_stickers=row["total_stickers"] or 0,
            sticker_briefs_sample=briefs_sample,
            suggested_seller_sku=default_sku,
        )

        main_text = self.router.text_complete(
            prompt,
            model=main_model,
            system=DETAIL_MAIN_SYSTEM_PROMPT,
            temperature=0.7,
            task="tkshop_detail:main",
            related_table="tkshop_products",
            related_id=product_id,
        )

        extract_ok = True
        extract_error = ""
        payload: dict[str, Any] = {}
        try:
            payload = self.router.extract_json(
                main_text,
                schema=DETAIL_EXTRACT_SCHEMA,
                instructions=DETAIL_EXTRACT_INSTRUCTIONS,
                model=extract_model,
                max_retries=1,
                task="tkshop_detail:extract",
                related_table="tkshop_products",
                related_id=product_id,
            )
        except Exception as e:
            logger.warning("tkshop detail extract failed for #%d: %s", product_id, e)
            extract_ok = False
            extract_error = str(e)[:500]

        # Decide which seller_sku to persist. Prefer the AI's value if it's
        # valid; otherwise fall back to the deterministic default.
        ai_sku = (payload.get("seller_sku") or "").strip().upper()
        if _is_valid_seller_sku(ai_sku):
            chosen_sku = ai_sku
        else:
            chosen_sku = default_sku

        with _open_db(self.db_path) as conn:
            # Guarantee no other row on the same shop uses this SKU. Scope
            # is per-shop now — two shops sharing the same pack legitimately
            # share the same SKU (operator wants cross-shop alignment).
            chosen_sku = _ensure_unique_seller_sku(
                conn, chosen_sku,
                exclude_product_id=product_id,
                shop=row["shop"] or TKSHOP_DEFAULT_SHOP,
            )
            if extract_ok:
                conn.execute(
                    """
                    UPDATE tkshop_products
                       SET detail_main_raw_text = ?, title = ?,
                           description_html = ?, selling_points = ?,
                           keywords = ?, seller_sku = ?
                     WHERE id = ?
                    """,
                    (
                        main_text,
                        _clamp_title(payload.get("title") or ""),
                        payload.get("description_html") or "",
                        json.dumps(payload.get("selling_points") or [], ensure_ascii=False),
                        json.dumps(payload.get("keywords") or [], ensure_ascii=False),
                        chosen_sku,
                        product_id,
                    ),
                )
            else:
                conn.execute(
                    """
                    UPDATE tkshop_products
                       SET detail_main_raw_text = ?, seller_sku = ?
                     WHERE id = ?
                    """,
                    (main_text, chosen_sku, product_id),
                )
            # Mirror into the local_products master so the 本地产品 tab and
            # any future listing for the same pack inherit the same content.
            # Master seller_sku has no shop prefix (per-shop prefix is applied
            # when a listing is created); we strip the listing's prefix here
            # by recomputing the unprefixed slug.
            lp_row = conn.execute(
                "SELECT pr.local_product_id, p.pack_uid "
                "FROM tkshop_products pr JOIN packs p ON p.id = pr.pack_id "
                "WHERE pr.id = ?",
                (product_id,),
            ).fetchone()
            lp_id = lp_row["local_product_id"] if lp_row else None
            if lp_id is None and lp_row:
                # Legacy row pre-migration — mint a master on the fly.
                now = int(time.time())
                conn.execute(
                    "INSERT INTO local_products (pack_id, created_at, updated_at) "
                    "SELECT pack_id, ?, ? FROM tkshop_products WHERE id = ?",
                    (now, now, product_id),
                )
                lp_id = conn.execute(
                    "SELECT id FROM local_products WHERE pack_id = "
                    "(SELECT pack_id FROM tkshop_products WHERE id = ?)",
                    (product_id,),
                ).fetchone()["id"]
                conn.execute(
                    "UPDATE tkshop_products SET local_product_id = ? WHERE id = ?",
                    (lp_id, product_id),
                )
            if lp_id is not None:
                master_sku = _slugify_for_sku(lp_row["pack_uid"] or "")
                master_sku = f"INK-{master_sku or 'P' + str(product_id)}"[:25]
                # Master SKU only needs to be unique among masters (listings
                # carry their own per-shop variant).
                while conn.execute(
                    "SELECT 1 FROM local_products "
                    "WHERE UPPER(seller_sku) = ? AND id != ?",
                    (master_sku, lp_id),
                ).fetchone():
                    master_sku = master_sku[:23] + "-2"
                if extract_ok:
                    conn.execute(
                        """
                        UPDATE local_products
                           SET detail_main_raw_text = ?, title = ?,
                               description_html = ?, selling_points = ?,
                               keywords = ?, seller_sku = ?, updated_at = ?
                         WHERE id = ?
                        """,
                        (
                            main_text,
                            _clamp_title(payload.get("title") or ""),
                            payload.get("description_html") or "",
                            json.dumps(payload.get("selling_points") or [], ensure_ascii=False),
                            json.dumps(payload.get("keywords") or [], ensure_ascii=False),
                            master_sku,
                            int(time.time()),
                            lp_id,
                        ),
                    )
                else:
                    conn.execute(
                        "UPDATE local_products "
                        "SET detail_main_raw_text = ?, seller_sku = ?, updated_at = ? "
                        "WHERE id = ?",
                        (main_text, master_sku, int(time.time()), lp_id),
                    )
            conn.commit()

        # Persist all artifacts to disk via PackStore.
        try:
            self.store.write_product_detail(
                row["pack_uid"], f"draft_{product_id}",
                title=payload.get("title") or "",
                description_html=payload.get("description_html") or "",
                selling_points=payload.get("selling_points") or [],
                keywords=payload.get("keywords") or [],
                main_raw=main_text,
            )
        except Exception as e:
            logger.warning("write_product_detail failed: %s", e)

        return {
            "product_id": product_id,
            "extract_ok": extract_ok,
            "extract_error": extract_error,
            "main_chars": len(main_text),
            "title": payload.get("title", ""),
            "selling_points_count": len(payload.get("selling_points", [])),
            "keywords_count": len(payload.get("keywords", [])),
            "seller_sku": chosen_sku,
        }

    def update_detail_manual(
        self,
        product_id: int,
        *,
        title: Optional[str] = None,
        description_html: Optional[str] = None,
        selling_points: Optional[list[str]] = None,
        keywords: Optional[list[str]] = None,
        category_id: Optional[str] = None,
        seller_sku: Optional[str] = None,
    ) -> bool:
        sets, params = [], []
        if title is not None:
            sets.append("title = ?"); params.append(_clamp_title(title))
        if description_html is not None:
            sets.append("description_html = ?"); params.append(description_html)
        if selling_points is not None:
            sets.append("selling_points = ?"); params.append(json.dumps(selling_points, ensure_ascii=False))
        if keywords is not None:
            sets.append("keywords = ?"); params.append(json.dumps(keywords, ensure_ascii=False))
        if category_id is not None:
            sets.append("category_id = ?"); params.append(category_id)
        if not sets and seller_sku is None:
            return False
        with _open_db(self.db_path) as conn:
            # Resolve SKU collision before binding it into the UPDATE so the
            # operator's manual edit can't silently shadow another row's SKU.
            if seller_sku is not None:
                clean_sku = _ensure_unique_seller_sku(
                    conn, (seller_sku or "").strip().upper()[:25],
                    exclude_product_id=product_id,
                    shop=_get_product_shop(self.db_path, product_id),
                )
                sets.append("seller_sku = ?"); params.append(clean_sku)
            cur = conn.execute(
                f"UPDATE tkshop_products SET {', '.join(sets)} WHERE id = ?",
                tuple(params) + (product_id,),
            )
            conn.commit()
            return cur.rowcount > 0

    def set_seller_sku(self, product_id: int, seller_sku: str) -> bool:
        """Tiny helper used by self-heal to update only the SKU."""
        clean = (seller_sku or "").strip().upper()[:25]
        with _open_db(self.db_path) as conn:
            clean = _ensure_unique_seller_sku(
                conn, clean, exclude_product_id=product_id,
                shop=_get_product_shop(self.db_path, product_id),
            )
            cur = conn.execute(
                "UPDATE tkshop_products SET seller_sku = ? WHERE id = ?",
                (clean, product_id),
            )
            conn.commit()
            return cur.rowcount > 0

    # ------------------------------------------------------------------
    # C.2 product images
    # ------------------------------------------------------------------

    def add_product_image(
        self,
        product_id: int,
        *,
        src_path: Path,
        role: str = "main",
        source: str = "manual",
        ai_prompt: str = "",
    ) -> int:
        """Save uploaded image into PackStore + insert tkshop_product_images row."""
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                """
                SELECT pr.id, p.pack_uid
                  FROM tkshop_products pr
                  JOIN packs           p ON p.id = pr.pack_id
                 WHERE pr.id = ?
                """,
                (product_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"product #{product_id} not found")
            sort_order = (conn.execute(
                """
                SELECT COALESCE(MAX(sort_order), 0) + 1
                  FROM tkshop_product_images
                 WHERE product_id = ? AND role = ?
                """, (product_id, role),
            ).fetchone()[0]) or 1
        ext = (src_path.suffix or ".png").lower()
        if ext not in (".png", ".jpg", ".jpeg", ".webp"):
            ext = ".png"
        filename = f"{role}_{sort_order}{ext}" if role != "main" or sort_order > 1 else f"main{ext}"
        dest = self.store.write_product_image(
            row["pack_uid"], f"draft_{product_id}", filename,
            src_path.read_bytes(),
        )
        rel = dest.as_posix()
        with _open_db(self.db_path) as conn:
            cur = conn.execute(
                """
                INSERT INTO tkshop_product_images
                    (product_id, role, source, local_path, tiktok_image_uri,
                     sort_order, ai_prompt, created_at)
                VALUES (?, ?, ?, ?, '', ?, ?, ?)
                """,
                (product_id, role, source, rel, sort_order, ai_prompt, int(time.time())),
            )
            conn.commit()
            return cur.lastrowid

    def list_product_images(self, product_id: int) -> list[dict]:
        with _open_db(self.db_path) as conn:
            rows = conn.execute(
                """
                SELECT id, role, source, local_path, tiktok_image_uri,
                       sort_order, ai_prompt, created_at
                  FROM tkshop_product_images
                 WHERE product_id = ?
                 ORDER BY role, sort_order
                """,
                (product_id,),
            ).fetchall()
        return [dict(r) for r in rows]

    def update_image(
        self,
        image_id: int,
        *,
        role: Optional[str] = None,
        sort_order: Optional[int] = None,
    ) -> dict[str, Any]:
        """Mutate role and/or sort_order of an existing product image.

        If ``role='main'`` is requested and another image already holds main
        for the same product, that other image is demoted to 'secondary'
        (TikTok allows only one main).
        """
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                "SELECT product_id, role, sort_order FROM tkshop_product_images WHERE id = ?",
                (image_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"image #{image_id} not found")

            sets, params = [], []
            if role is not None and role != row["role"]:
                if role == "main":
                    # Demote any existing main for this product.
                    conn.execute(
                        "UPDATE tkshop_product_images "
                        "SET role = 'secondary' WHERE product_id = ? AND role = 'main' AND id != ?",
                        (row["product_id"], image_id),
                    )
                sets.append("role = ?"); params.append(role)
            if sort_order is not None and int(sort_order) != row["sort_order"]:
                sets.append("sort_order = ?"); params.append(int(sort_order))
            if not sets:
                conn.commit()
                return {"changed": False, "image_id": image_id,
                        "role": row["role"], "sort_order": row["sort_order"]}
            conn.execute(
                f"UPDATE tkshop_product_images SET {', '.join(sets)} WHERE id = ?",
                tuple(params) + (image_id,),
            )
            conn.commit()
            new_row = conn.execute(
                "SELECT role, sort_order FROM tkshop_product_images WHERE id = ?",
                (image_id,),
            ).fetchone()
            product_id = row["product_id"]
            lp_row = conn.execute(
                "SELECT local_product_id FROM tkshop_products WHERE id = ?",
                (product_id,),
            ).fetchone()
            lp_id = lp_row["local_product_id"] if lp_row else None
        if role == "main" and lp_id is not None:
            self._sync_listing_main_to_local(lp_id, product_id)
        return {"changed": True, "image_id": image_id,
                "role": new_row["role"], "sort_order": new_row["sort_order"]}

    def delete_product(self, product_id: int) -> dict[str, Any]:
        """Delete a tkshop_product row and all its dependents:
        product images (rows + on-disk files), publish logs, and the
        product subdir under PackStore. Returns a small summary.

        NOTE: Does NOT touch TikTok-side data. If the product was already
        published, the listing on TikTok Shop remains live — operator
        should deactivate it via the TikTok seller center separately.
        """
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                """
                SELECT pr.id, pr.publish_status, pr.tiktok_product_id,
                       p.pack_uid
                  FROM tkshop_products pr
                  LEFT JOIN packs p ON p.id = pr.pack_id
                 WHERE pr.id = ?
                """,
                (product_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"product #{product_id} not found")
            imgs = conn.execute(
                "SELECT id, local_path FROM tkshop_product_images WHERE product_id = ?",
                (product_id,),
            ).fetchall()

        # Remove image files first (best-effort; row-cascade-by-hand below).
        for i in imgs:
            if i["local_path"]:
                try:
                    Path(i["local_path"]).unlink()
                except Exception:
                    pass

        # Remove on-disk product subfolder if the PackStore knows the pack.
        pack_uid = row["pack_uid"] or ""
        if pack_uid:
            try:
                pdir = self.store.product_dir(pack_uid, f"draft_{product_id}")
                if pdir.is_dir():
                    import shutil as _sh
                    _sh.rmtree(pdir, ignore_errors=True)
                    logger.info("Removed product dir %s", pdir)
            except Exception as e:
                logger.warning("could not remove product dir: %s", e)

        with _open_db(self.db_path) as conn:
            conn.execute("DELETE FROM tkshop_product_images WHERE product_id = ?",
                         (product_id,))
            conn.execute("DELETE FROM tkshop_publish_logs WHERE product_id = ?",
                         (product_id,))
            cur = conn.execute("DELETE FROM tkshop_products WHERE id = ?",
                               (product_id,))
            conn.commit()

        return {
            "deleted": cur.rowcount > 0,
            "product_id": product_id,
            "images_removed": len(imgs),
            "was_published": bool(row["tiktok_product_id"]),
            "tiktok_product_id": row["tiktok_product_id"] or "",
        }

    def _sync_pack_cover_from_listing_main(
        self, product_id: int, *, conn: Any = None,
    ) -> None:
        """Point packs.cover_image_path at the listing's current AI main."""

        def _run(c: Any) -> None:
            row = c.execute(
                """
                SELECT pr.pack_id, ti.local_path
                  FROM tkshop_products pr
                  JOIN tkshop_product_images ti ON ti.product_id = pr.id
                 WHERE pr.id = ? AND ti.role = 'main' AND ti.source = 'ai'
                 ORDER BY ti.sort_order ASC, ti.id ASC
                 LIMIT 1
                """,
                (product_id,),
            ).fetchone()
            if not row or not (row["local_path"] or "").strip():
                return
            c.execute(
                "UPDATE packs SET cover_image_path = ? WHERE id = ?",
                (row["local_path"], row["pack_id"]),
            )

        if conn is not None:
            _run(conn)
        else:
            with _open_db(self.db_path) as c:
                _run(c)
                c.commit()

    def _clear_pack_cover_if_matches(
        self, pack_id: int, local_path: str, conn: Any = None,
    ) -> None:
        """Drop packs.cover_image_path when it still points at a removed image."""
        path = (local_path or "").strip()
        if not path or not pack_id:
            return
        if conn is not None:
            conn.execute(
                "UPDATE packs SET cover_image_path = '' "
                "WHERE id = ? AND cover_image_path = ?",
                (pack_id, path),
            )
            return
        with _open_db(self.db_path) as c:
            c.execute(
                "UPDATE packs SET cover_image_path = '' "
                "WHERE id = ? AND cover_image_path = ?",
                (pack_id, path),
            )
            c.commit()

    def delete_image(self, image_id: int) -> bool:
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                """
                SELECT ti.local_path, ti.role, pr.pack_id
                  FROM tkshop_product_images ti
                  JOIN tkshop_products pr ON pr.id = ti.product_id
                 WHERE ti.id = ?
                """,
                (image_id,),
            ).fetchone()
            if not row:
                return False
            cur = conn.execute(
                "DELETE FROM tkshop_product_images WHERE id = ?", (image_id,),
            )
            if row["pack_id"]:
                self._clear_pack_cover_if_matches(
                    row["pack_id"], row["local_path"], conn,
                )
            conn.commit()
            ok = cur.rowcount > 0
        if ok and row["local_path"]:
            try:
                Path(row["local_path"]).unlink()
            except Exception:
                pass
        return ok

    # ------------------------------------------------------------------
    # C.2.b batch upload + pack-preview import
    # ------------------------------------------------------------------

    def add_product_images_batch(
        self,
        product_id: int,
        files: list[tuple[str, bytes]],
    ) -> dict[str, Any]:
        """Save many uploaded files at once. First file becomes 'main'
        if no main exists; remaining files are 'secondary'.
        ``files`` = list of (filename, bytes).
        """
        if not files:
            return {"created": 0, "image_ids": []}
        with _open_db(self.db_path) as conn:
            has_main = conn.execute(
                "SELECT 1 FROM tkshop_product_images WHERE product_id = ? AND role = 'main' LIMIT 1",
                (product_id,),
            ).fetchone() is not None
        created_ids: list[int] = []
        import tempfile
        tmpdir = Path(tempfile.mkdtemp())
        try:
            for i, (name, raw) in enumerate(files):
                if not raw:
                    continue
                role = "secondary" if (has_main or i > 0) else "main"
                if not has_main and i == 0:
                    has_main = True  # claimed by first file
                ext = Path(name).suffix.lower() or ".png"
                if ext not in (".png", ".jpg", ".jpeg", ".webp"):
                    ext = ".png"
                tmp = tmpdir / f"upload_{i}{ext}"
                tmp.write_bytes(raw)
                img_id = self.add_product_image(
                    product_id, src_path=tmp, role=role, source="manual",
                )
                created_ids.append(img_id)
        finally:
            try:
                for f in tmpdir.iterdir():
                    f.unlink()
                tmpdir.rmdir()
            except Exception:
                pass
        return {"created": len(created_ids), "image_ids": created_ids}

    def import_pack_previews(self, product_id: int) -> dict[str, Any]:
        """One-click: copy each ok preview from the pack's series into
        product images. First imported = 'main' if none yet, rest =
        'secondary'. Skips previews whose image_path is missing.
        """
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                """
                SELECT pr.id AS product_id, p.id AS pack_id, p.series_id, p.pack_uid
                  FROM tkshop_products pr
                  JOIN packs p ON p.id = pr.pack_id
                 WHERE pr.id = ?
                """,
                (product_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"product #{product_id} not found")
            previews = conn.execute(
                """
                SELECT id, preview_idx, image_path
                  FROM pack_previews
                 WHERE series_id = ? AND generation_status = 'ok'
                       AND COALESCE(image_path, '') != ''
                 ORDER BY preview_idx
                """,
                (row["series_id"],),
            ).fetchall() if row["series_id"] else []
        if not previews:
            return {"imported": 0, "skipped": 0, "reason": "no ok previews"}
        files: list[tuple[str, bytes]] = []
        skipped = 0
        for p in previews:
            src = Path(p["image_path"])
            if not src.is_absolute():
                src = Path(p["image_path"])
            try:
                raw = src.read_bytes()
                files.append((f"preview_{p['preview_idx']}{src.suffix or '.png'}", raw))
            except Exception as e:
                logger.warning("import_pack_previews: skip preview #%d (%s): %s",
                               p["id"], src, e)
                skipped += 1
        result = self.add_product_images_batch(product_id, files)
        return {
            "imported": result["created"],
            "skipped": skipped,
            "image_ids": result["image_ids"],
        }

    # ------------------------------------------------------------------
    # C.2.c AI-driven product image design + image-to-image generation
    # ------------------------------------------------------------------

    def collect_pack_design_context(self, product_id: int) -> dict[str, Any]:
        """Gather pack metadata + sample sticker briefs + sample preview
        prompts into a single design-context dict ready for AI synthesis.
        """
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                """
                SELECT pr.id, pr.pack_id,
                       p.display_name, p.total_stickers, p.pack_uid, p.series_id,
                       s.style_anchor, s.palette, s.pack_archetype, s.metadata_json
                  FROM tkshop_products pr
                  JOIN packs p ON p.id = pr.pack_id
             LEFT JOIN pack_series s ON s.id = p.series_id
                 WHERE pr.id = ?
                """, (product_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"product #{product_id} not found")
            previews = conn.execute(
                """
                SELECT image_path, prompt_text
                  FROM pack_previews
                 WHERE series_id = ? AND generation_status = 'ok'
                 ORDER BY preview_idx
                 LIMIT 60
                """,
                (row["series_id"],),
            ).fetchall() if row["series_id"] else []
            # High-fidelity split stickers (A.3 image-edit output) make far
            # better image-to-image references than a full sheet. Keep enough
            # of them to represent the whole pack in the main-image reference
            # grid; selected stickers only influence ordering.
            stickers = conn.execute(
                """
                SELECT ps.image_path, ps.name, ps.is_selected
                  FROM pack_stickers ps
                  JOIN pack_previews pv ON pv.id = ps.preview_id
                 WHERE pv.series_id = ?
                   AND ps.generation_status = 'ok'
                   AND ps.image_path != ''
                 ORDER BY ps.is_selected DESC, ps.id
                 LIMIT 80
                """,
                (row["series_id"],),
            ).fetchall() if row["series_id"] else []

        briefs: list[str] = []
        source_image_paths: list[str] = []
        try:
            md = json.loads(row["metadata_json"] or "{}")
            for pb in md.get("preview_briefs", [])[:3]:
                briefs.extend(pb.get("stickers", [])[:6])
            for source in md.get("source_files") or []:
                for key in ("original_preview_paths", "source_preview_paths"):
                    for p in source.get(key) or []:
                        p = str(p or "").strip()
                        if p:
                            source_image_paths.append(p)
                for key in ("source_path", "source_image_path"):
                    p = str(source.get(key) or "").strip()
                    if p:
                        source_image_paths.append(p)
                for p in source.get("preview_paths") or []:
                    p = str(p or "").strip()
                    if p:
                        source_image_paths.append(p)
            for pb in md.get("ai_source_preview_briefs") or []:
                p = str(pb.get("source_image_path") or "").strip()
                if p:
                    source_image_paths.append(p)
        except Exception:
            pass
        source_image_paths = list(dict.fromkeys(
            _normalize_disk_path(p) for p in source_image_paths if _normalize_disk_path(p)
        ))

        def _norm_paths(raw: list[str]) -> list[str]:
            return [_normalize_disk_path(p) for p in raw if _normalize_disk_path(p)]

        return {
            "product_id": product_id,
            "pack_uid": row["pack_uid"] or "",
            "display_name": row["display_name"] or "",
            "pack_archetype": row["pack_archetype"] or "",
            "style_anchor": row["style_anchor"] or "",
            "palette": row["palette"] or "",
            "total_stickers": row["total_stickers"] or 0,
            "briefs_sample": briefs,
            "preview_prompts": [p["prompt_text"] or "" for p in previews],
            "preview_image_paths": _norm_paths(
                [p["image_path"] for p in previews if p["image_path"]]
            ),
            "source_image_paths": source_image_paths,
            "sticker_image_paths": _norm_paths(
                [s["image_path"] for s in stickers if s["image_path"]]
            ),
            "selected_sticker_subjects": [
                s["name"] for s in stickers if s["is_selected"] and (s["name"] or "").strip()
            ],
        }

    @staticmethod
    def _render_white_canvas_image(source_path: Path, *, size: str = "1024x1024") -> bytes:
        try:
            w_str, h_str = (size or "1024x1024").lower().split("x")
            canvas_w, canvas_h = int(w_str), int(h_str)
        except Exception:
            canvas_w, canvas_h = 1024, 1024
        canvas_w = max(512, canvas_w)
        canvas_h = max(512, canvas_h)
        with Image.open(source_path) as src:
            image = src.convert("RGBA")
            image.thumbnail((int(canvas_w * 0.94), int(canvas_h * 0.94)), Image.Resampling.LANCZOS)
            canvas = Image.new("RGBA", (canvas_w, canvas_h), (255, 255, 255, 255))
            x = (canvas_w - image.width) // 2
            y = (canvas_h - image.height) // 2
            canvas.alpha_composite(image, (x, y))
        out = BytesIO()
        canvas.convert("RGB").save(out, "PNG", optimize=True)
        return out.getvalue()

    def synthesize_image_specs(
        self,
        product_id: int,
        *,
        secondary_count: int = 3,
        language: str = "en",
        model: Optional[str] = None,
        existing_gallery: Optional[list[dict[str, str]]] = None,
    ) -> dict[str, Any]:
        """Run gpt-5.4 (or override) on the design context to produce a
        structured ``{main, secondary[]}`` image-prompt spec.
        """
        ctx = self.collect_pack_design_context(product_id)
        prompt = build_image_design_prompt(
            pack_display_name=ctx["display_name"],
            pack_archetype=ctx["pack_archetype"],
            style_anchor=ctx["style_anchor"],
            palette=ctx["palette"],
            total_stickers=ctx["total_stickers"],
            sticker_briefs_sample=ctx["briefs_sample"],
            preview_prompt_samples=ctx["preview_prompts"],
            secondary_count=secondary_count,
            language=language,
            selected_sticker_subjects=ctx.get("selected_sticker_subjects") or [],
            existing_gallery=existing_gallery,
        )
        chosen_model = model or os.getenv("TKSHOP_IMAGE_DESIGN_MODEL", "gpt-5.4")
        main_text = self.router.text_complete(
            prompt,
            model=chosen_model,
            system=IMAGE_DESIGN_SYSTEM_PROMPT,
            temperature=0.6,
            task="tkshop_image_design:main",
            related_table="tkshop_products",
            related_id=product_id,
        )
        spec = self.router.extract_json(
            main_text,
            schema=IMAGE_DESIGN_EXTRACT_SCHEMA,
            instructions=IMAGE_DESIGN_INSTRUCTIONS,
            model=chosen_model,
            max_retries=1,
            task="tkshop_image_design:extract",
            related_table="tkshop_products",
            related_id=product_id,
        )
        return {
            "spec": spec,
            "context": ctx,
            "raw": main_text,
        }

    def synthesize_append_secondary_spec(
        self,
        product_id: int,
        *,
        existing_gallery: list[dict[str, str]],
        language: str = "en",
        model: Optional[str] = None,
    ) -> dict[str, Any]:
        """GPT: one new secondary spec that differs from existing gallery prompts."""
        ctx = self.collect_pack_design_context(product_id)
        prompt = build_append_secondary_prompt(
            pack_display_name=ctx["display_name"],
            pack_archetype=ctx["pack_archetype"],
            style_anchor=ctx["style_anchor"],
            palette=ctx["palette"],
            total_stickers=ctx["total_stickers"],
            sticker_briefs_sample=ctx["briefs_sample"],
            existing_gallery=existing_gallery,
            language=language,
        )
        chosen_model = model or os.getenv("TKSHOP_IMAGE_DESIGN_MODEL", "gpt-5.4")
        main_text = self.router.text_complete(
            prompt,
            model=chosen_model,
            system=IMAGE_DESIGN_SYSTEM_PROMPT,
            temperature=0.7,
            task="tkshop_image_design:append_secondary",
            related_table="tkshop_products",
            related_id=product_id,
        )
        spec = self.router.extract_json(
            main_text,
            schema=APPEND_SECONDARY_EXTRACT_SCHEMA,
            instructions=APPEND_SECONDARY_INSTRUCTIONS,
            model=chosen_model,
            max_retries=1,
            task="tkshop_image_design:append_extract",
            related_table="tkshop_products",
            related_id=product_id,
        )
        return {"spec": spec, "context": ctx, "raw": main_text}

    def _collect_master_gallery_for_design(
        self, product_id: int,
    ) -> tuple[list[dict[str, str]], list[int], bool]:
        """Gallery prompts + aHash fingerprints for dedup (master preferred)."""
        from src.utils.image_utils import average_hash

        asset_root = self.db_path.resolve().parent.parent
        prompt_rows: list[dict[str, str]] = []
        hashes: list[int] = []
        has_main = False
        seen_paths: set[str] = set()
        with _open_db(self.db_path) as conn:
            lp_row = conn.execute(
                "SELECT local_product_id FROM tkshop_products WHERE id = ?",
                (product_id,),
            ).fetchone()
            lp_id = lp_row["local_product_id"] if lp_row else None
            if lp_id:
                rows = conn.execute(
                    """
                    SELECT role, local_path, ai_prompt, sort_order
                      FROM local_product_images
                     WHERE local_product_id = ?
                     ORDER BY CASE role WHEN 'main' THEN 0 ELSE 1 END,
                              sort_order, id
                    """,
                    (lp_id,),
                ).fetchall()
            else:
                rows = conn.execute(
                    """
                    SELECT role, local_path, ai_prompt, sort_order
                      FROM tkshop_product_images
                     WHERE product_id = ?
                     ORDER BY CASE role WHEN 'main' THEN 0 ELSE 1 END,
                              sort_order, id
                    """,
                    (product_id,),
                ).fetchall()
        for r in rows:
            role = (r["role"] or "secondary").strip()
            if role == "main":
                has_main = True
            ai_prompt = (r["ai_prompt"] or "").strip()
            prompt_rows.append({
                "role": role,
                "ai_prompt": ai_prompt,
                "hint": ai_prompt[:200],
            })
            p = _resolve_disk_path(r["local_path"] or "", root=asset_root)
            if not p.is_file():
                continue
            key = p.as_posix()
            if key in seen_paths:
                continue
            seen_paths.add(key)
            try:
                hashes.append(average_hash(p.read_bytes()))
            except Exception:
                continue
        return prompt_rows, hashes, has_main

    @staticmethod
    def _infer_role_from_prompt(prompt: str, role: str = "") -> str:
        p = (prompt or "").lower()
        if role == "main" or (
            "main image" in p
            or ("bundle" in p and "white background" in p)
            or p.startswith("studio e-commerce main image")
        ):
            return "hero"
        if "packaging" in p or "gift box" in p or "glassine" in p or "envelope" in p:
            return "packaging"
        if any(k in p for k in ("full-set", "full set", "whole sticker set", "tidy grid", "whole set")):
            return "full_set"
        if any(k in p for k in ("peeling", "in-use", "in use", "backing", "pressed onto")):
            return "in_use"
        if "flat lay" in p or "flat_lay" in p:
            return "flat_lay"
        if "scene" in p or "still-life" in p or "still life" in p:
            return "scene"
        return "lifestyle"

    def _pick_unused_secondary_role(
        self, existing_gallery: list[dict[str, str]], *, gallery_has_main: bool,
    ) -> str:
        used = {
            self._infer_role_from_prompt(
                g.get("ai_prompt") or g.get("hint") or "", g.get("role") or "",
            )
            for g in existing_gallery
        }
        preference = ["packaging", "scene", "lifestyle", "in_use", "flat_lay", "full_set"]
        if gallery_has_main:
            preference = ["packaging", "scene", "lifestyle", "in_use"]
        for rt in preference:
            if rt not in used:
                return rt
        return "packaging"

    def _ensure_diverse_secondary(
        self,
        sec: dict[str, Any],
        existing_gallery: list[dict[str, str]],
        *,
        gallery_has_main: bool,
    ) -> dict[str, Any]:
        """Remap variety/bundle angles when gallery already covers them."""
        out = dict(sec)
        rt = (out.get("role_type") or "lifestyle").strip().lower()
        if rt == "hand_hold":
            return out
        used = {
            self._infer_role_from_prompt(
                g.get("ai_prompt") or g.get("hint") or "", g.get("role") or "",
            )
            for g in existing_gallery
        }
        if gallery_has_main and rt in self._VARIETY_ROLE_TYPES:
            rt = self._pick_unused_secondary_role(
                existing_gallery, gallery_has_main=True,
            )
            out["role_type"] = rt
        elif rt in used:
            rt = self._pick_unused_secondary_role(
                existing_gallery, gallery_has_main=gallery_has_main,
            )
            out["role_type"] = rt
        return out

    # Secondary angles that show the whole set benefit from a full SHEET as
    # the image-to-image reference; single-subject angles stage one clean
    # die-cut sticker better. role_type keys come from tkshop.prompts.
    _VARIETY_ROLE_TYPES = {"hero", "flat_lay", "full_set", "hand_hold"}

    @staticmethod
    def _inject_hand_hold_fifth_secondary(
        spec: dict[str, Any], secondary_count: int,
    ) -> None:
        """Last secondary in the batch becomes the fixed hand-hold 5th gallery image."""
        if secondary_count < 1:
            return
        hh = hand_hold_secondary_spec()
        secs = spec.get("secondary")
        if not isinstance(secs, list):
            spec["secondary"] = [hh]
            return
        padded = list(secs[:secondary_count])
        while len(padded) < secondary_count:
            padded.append({})
        padded[secondary_count - 1] = hh
        spec["secondary"] = padded

    def auto_design_images(
        self,
        product_id: int,
        *,
        secondary_count: int = 3,
        language: str = "en",
        replace_existing_ai: bool = True,
        size: str = "1024x1024",
        main_size: Optional[str] = None,
        main_quality: Optional[str] = None,
        secondary_quality: Optional[str] = None,
        mode: str = "all",
    ) -> dict[str, Any]:
        """End-to-end: design context → AI spec → image_edit per concept
        → save as tkshop_product_images rows (source='ai').

        Reference selection is per-role: "variety" angles (hero / flat_lay /
        full_set) edit a single MERGED grid. If the pack has enough preview
        sheets (default: at least 3), those previews are the preferred main
        reference; otherwise split sticker assets are used. Single-subject
        angles (lifestyle / in_use / scene / packaging) prefer one split
        sticker when available. The main/hero uses a fixed white-bg "bundle"
        prompt over that merged grid.
        The main (hero) image renders at ``main_quality`` /
        ``main_size`` (defaults: high quality, falls back to ``size``);
        secondary images use ``secondary_quality`` / ``size``.

        Each output passes a QA gate (decodable, sane dimensions, and not a
        near-identical copy of its reference); a flagged concept is retried
        once with a stronger re-staging instruction. If ``replace_existing_ai``
        is True, prior images in each role slot being regenerated (main /
        secondary) are deleted first — manual and AI alike — so new AI output
        replaces the gallery instead of appending beside old uploads.
        """
        from src.utils.image_utils import (
            average_hash,
            hash_distance,
            read_dimensions,
            compose_reference_grid,
            compress_image_bytes_for_api,
        )

        gallery_prompts, gallery_hashes, gallery_has_main = (
            self._collect_master_gallery_for_design(product_id)
        )
        gallery_threshold = int(os.getenv("TKSHOP_IMAGE_GALLERY_AHASH_DIST", "12"))
        qa_max_attempts = max(2, int(os.getenv("TKSHOP_IMAGE_QA_MAX_ATTEMPTS", "3")))

        mode_norm_early = (mode or "all").strip().lower()
        append_secondary = (
            mode_norm_early in {"all", "secondary"}
            and not replace_existing_ai
            and bool(gallery_prompts)
            and secondary_count == 1
        )
        use_hand_hold_fifth = fifth_image_style_is_hand_hold()
        if append_secondary and use_hand_hold_fifth:
            ctx = self.collect_pack_design_context(product_id)
            spec = {"main": {}, "secondary": [hand_hold_secondary_spec()]}
            synth = {"context": ctx, "spec": spec}
        elif append_secondary:
            synth = self.synthesize_append_secondary_spec(
                product_id,
                existing_gallery=gallery_prompts,
                language=language,
            )
            sec_raw = synth["spec"].get("secondary") or {}
            if not isinstance(sec_raw, dict):
                raise ValueError("append secondary spec missing secondary object")
            sec = self._ensure_diverse_secondary(
                sec_raw, gallery_prompts, gallery_has_main=gallery_has_main,
            )
            spec = {"main": {}, "secondary": [sec]}
            ctx = synth["context"]
        else:
            synth = self.synthesize_image_specs(
                product_id,
                secondary_count=secondary_count,
                language=language,
                existing_gallery=gallery_prompts or None,
            )
            spec = dict(synth["spec"])
            ctx = synth["context"]
            if (
                use_hand_hold_fifth
                and mode_norm_early in {"all", "secondary"}
                and secondary_count >= 1
                and isinstance(spec.get("secondary"), list)
            ):
                self._inject_hand_hold_fifth_secondary(spec, secondary_count)
            if gallery_prompts and isinstance(spec.get("secondary"), list):
                spec["secondary"] = [
                    self._ensure_diverse_secondary(
                        s, gallery_prompts, gallery_has_main=gallery_has_main,
                    )
                    if isinstance(s, dict) else s
                    for s in spec["secondary"][:secondary_count]
                ]

        asset_root = self.db_path.resolve().parent.parent

        def _existing_asset_paths(raw_paths: list[str]) -> list[Path]:
            seen: set[str] = set()
            out: list[Path] = []
            for raw in raw_paths:
                p = _resolve_disk_path(raw, root=asset_root)
                if not p.is_file():
                    continue
                key = p.as_posix()
                if key in seen:
                    continue
                seen.add(key)
                out.append(p)
            return out

        sheet_paths = _existing_asset_paths(ctx.get("preview_image_paths", []))
        source_paths = _existing_asset_paths(ctx.get("source_image_paths", []))
        sticker_paths = _existing_asset_paths(ctx.get("sticker_image_paths", []))
        if not sheet_paths and not sticker_paths:
            raise ValueError(
                "no preview or sticker image available to use as reference; "
                "generate at least one pack preview first"
            )

        # Resolve per-role size / quality (env-overridable).
        main_size = main_size or os.getenv("TKSHOP_MAIN_IMAGE_SIZE") or size
        # medium matches preview_gen sticker splits; high + large refs often
        # triggers JieKou transport drops (RemoteProtocolError).
        main_quality = main_quality or os.getenv("TKSHOP_MAIN_IMAGE_QUALITY", "medium")
        secondary_quality = secondary_quality or os.getenv("TKSHOP_SECONDARY_IMAGE_QUALITY", "medium")
        noop_threshold = int(os.getenv("TKSHOP_IMAGE_NOOP_AHASH_DIST", "6"))

        # JieKou gpt-image-2 only accepts a fixed size whitelist; snap anything
        # else to 1024x1024 so a bad UI/env value can't 400 the whole run.
        def _safe_size(s: str) -> str:
            s = (s or "").strip().lower()
            if s in _JIEKOU_IMAGE_SIZES:
                return s
            logger.warning("auto_design_images: unsupported size %r → 1024x1024", s)
            return "1024x1024"
        size = _safe_size(size)
        main_size = _safe_size(main_size)

        # External uploads (PDF/CDR parsed to source_previews): use the parsed
        # overview as the image_edit reference for main, same as preview grids.
        # Set TKSHOP_EXTERNAL_MAIN_DIRECT=1 to skip the model and paste the
        # source overview onto a white canvas (legacy behaviour).
        external_overview_paths = [
            p for p in source_paths
            if p.suffix.lower() in (".png", ".jpg", ".jpeg", ".webp")
        ]
        use_direct_external_main = (
            os.getenv("TKSHOP_EXTERNAL_MAIN_DIRECT", "").strip().lower()
            in ("1", "true", "yes", "on")
        )
        if external_overview_paths:
            main_bundle_paths = external_overview_paths[:1]
            main_bundle_ref_label = "external_source_overview"
            main_from_source_path: Optional[Path] = (
                main_bundle_paths[0] if use_direct_external_main else None
            )
            if use_direct_external_main:
                logger.info(
                    "auto_design_images: external overview → direct main "
                    "(TKSHOP_EXTERNAL_MAIN_DIRECT); skipping image_edit for hero"
                )
        else:
            main_from_source_path = None

        # One MERGED reference for AI-generated secondary variety angles. When
        # enough preview sheets exist they represent the whole family; otherwise
        # split sticker assets prevent a single preview from dominating.
        preview_threshold = max(1, int(os.getenv("TKSHOP_MAIN_PREVIEW_MIN_COUNT", "3")))
        if not external_overview_paths:
            if len(sheet_paths) >= preview_threshold:
                main_bundle_paths = sheet_paths
                main_bundle_ref_label = "preview_sheets_merged_grid"
            elif sticker_paths:
                main_bundle_paths = sticker_paths
                main_bundle_ref_label = "split_stickers_merged_grid"
            else:
                main_bundle_paths = sheet_paths
                main_bundle_ref_label = "preview_sheets_merged_grid"
        if len(sheet_paths) >= preview_threshold:
            secondary_bundle_paths = sheet_paths
            secondary_bundle_ref_label = "preview_sheets_merged_grid"
        elif sticker_paths:
            secondary_bundle_paths = sticker_paths
            secondary_bundle_ref_label = "split_stickers_merged_grid"
        else:
            secondary_bundle_paths = sheet_paths
            secondary_bundle_ref_label = "preview_sheets_merged_grid"
        merge_sources = [p.as_posix() for p in main_bundle_paths]
        merge_max_side = int(os.getenv("TKSHOP_MAIN_MERGE_MAX_SIDE", "1536"))
        merged_ref_bytes = compose_reference_grid(
            merge_sources,
            cell=384 if len(main_bundle_paths) >= 12 else 512,
            max_side=merge_max_side,
        )
        ref_upload_max_side = int(os.getenv("TKSHOP_MAIN_REF_UPLOAD_MAX_SIDE", "1536"))
        ref_max_bytes = int(os.getenv("TKSHOP_MAIN_REF_MAX_BYTES", "1800000"))
        before_compress = len(merged_ref_bytes)
        merged_ref_bytes = compress_image_bytes_for_api(
            merged_ref_bytes,
            max_side=ref_upload_max_side,
            max_bytes=ref_max_bytes,
        )
        merged_ref_hash = average_hash(merged_ref_bytes)
        logger.info(
            "auto_design_images: merged ref %s (%d sources, %d→%d bytes)",
            main_bundle_ref_label,
            len(main_bundle_paths),
            before_compress,
            len(merged_ref_bytes),
        )
        secondary_merge_sources = [p.as_posix() for p in secondary_bundle_paths]
        secondary_merged_ref_bytes = compose_reference_grid(
            secondary_merge_sources,
            cell=384 if len(secondary_bundle_paths) >= 12 else 512,
            max_side=merge_max_side,
        )
        secondary_before_compress = len(secondary_merged_ref_bytes)
        secondary_merged_ref_bytes = compress_image_bytes_for_api(
            secondary_merged_ref_bytes,
            max_side=ref_upload_max_side,
            max_bytes=ref_max_bytes,
        )
        secondary_merged_ref_hash = average_hash(secondary_merged_ref_bytes)
        logger.info(
            "auto_design_images: secondary merged ref %s (%d sources, %d→%d bytes)",
            secondary_bundle_ref_label,
            len(secondary_bundle_paths),
            secondary_before_compress,
            len(secondary_merged_ref_bytes),
        )
        single_source_hashes: list[tuple[str, int]] = []
        if len(main_bundle_paths) > 1:
            for p in main_bundle_paths[:12]:
                try:
                    single_source_hashes.append((p.as_posix(), average_hash(p.read_bytes())))
                except Exception:
                    continue

        reference_grid_path: Optional[Path] = None
        try:
            with _open_db(self.db_path) as conn:
                ref_row = conn.execute(
                    """
                    SELECT p.pack_uid
                      FROM tkshop_products pr
                      JOIN packs p ON p.id = pr.pack_id
                     WHERE pr.id = ?
                    """,
                    (product_id,),
                ).fetchone()
            if ref_row:
                ref_dir = self.store.product_dir(
                    ref_row["pack_uid"], f"draft_{product_id}",
                ) / "references"
                ref_dir.mkdir(parents=True, exist_ok=True)
                reference_grid_path = ref_dir / f"{main_bundle_ref_label}.png"
                reference_grid_path.write_bytes(merged_ref_bytes)
        except Exception as e:
            logger.warning("auto_design_images: could not persist reference grid: %s", e)

        # Bytes + aHash cache per reference path (read once, reused across calls).
        _ref_cache: dict[str, tuple[bytes, int]] = {}

        def _ref_for(
            role_type: str, idx: int, *, slot_role: str = "secondary",
        ) -> tuple[str, bytes, int, bool]:
            """Pick a reference for a concept. Returns (label, bytes, ahash,
            is_merged). Variety angles use the merged grid; single-subject
            angles rotate through individual split stickers. When the gallery
            already has a bundle main, new secondaries prefer single stickers
            even for flat_lay/full_set so outputs look distinct.
            """
            rt = (role_type or "hero").strip().lower()
            if rt == "hand_hold":
                return (
                    secondary_bundle_ref_label,
                    secondary_merged_ref_bytes,
                    secondary_merged_ref_hash,
                    True,
                )
            prefer_single = bool(sticker_paths) and (
                rt not in self._VARIETY_ROLE_TYPES
                or (
                    slot_role == "secondary"
                    and gallery_has_main
                    and rt in self._VARIETY_ROLE_TYPES
                )
            )
            if prefer_single:
                pick = sticker_paths[idx % len(sticker_paths)]
                key = pick.as_posix()
                if key not in _ref_cache:
                    b = pick.read_bytes()
                    _ref_cache[key] = (b, average_hash(b))
                b, h = _ref_cache[key]
                return key, b, h, False
            if rt in self._VARIETY_ROLE_TYPES or not sticker_paths:
                if external_overview_paths and rt != "hero":
                    return (
                        secondary_bundle_ref_label,
                        secondary_merged_ref_bytes,
                        secondary_merged_ref_hash,
                        True,
                    )
                return main_bundle_ref_label, merged_ref_bytes, merged_ref_hash, True
            pick = sticker_paths[idx % len(sticker_paths)]
            key = pick.as_posix()
            if key not in _ref_cache:
                b = pick.read_bytes()
                _ref_cache[key] = (b, average_hash(b))
            b, h = _ref_cache[key]
            return key, b, h, False

        def _expected_wh(size_str: str) -> Optional[tuple[int, int]]:
            try:
                w, h = size_str.lower().split("x")
                return int(w), int(h)
            except Exception:
                return None

        # ``mode`` filters which roles to actually generate:
        #   "all"       — main + secondary (original behavior)
        #   "main"      — only the hero/main image; secondaries untouched
        #   "secondary" — only the secondary set; main untouched
        # The unified spec generation upstream still runs in one GPT call
        # regardless of mode, so the operator's prompt design is consistent.
        mode_norm = (mode or "all").strip().lower()
        if mode_norm not in {"all", "main", "secondary"}:
            mode_norm = "all"
        gen_main = mode_norm in {"all", "main"}
        gen_secondary = mode_norm in {"all", "secondary"}

        if replace_existing_ai:
            roles_to_clear: set[str] = set()
            if gen_main and isinstance(spec.get("main"), dict):
                roles_to_clear.add("main")
            if gen_secondary and secondary_count > 0 and (spec.get("secondary") or []):
                roles_to_clear.add("secondary")
            if roles_to_clear:
                with _open_db(self.db_path) as conn:
                    placeholders = ",".join("?" * len(roles_to_clear))
                    rows = conn.execute(
                        f"""
                        SELECT id FROM tkshop_product_images
                         WHERE product_id = ? AND role IN ({placeholders})
                        """,
                        (product_id, *sorted(roles_to_clear)),
                    ).fetchall()
                for r in rows:
                    self.delete_image(r["id"])
                if rows:
                    logger.info(
                        "auto_design_images: cleared %d prior %s image(s) "
                        "(manual + ai) for product #%d",
                        len(rows), "/".join(sorted(roles_to_clear)), product_id,
                    )

        results: list[dict[str, Any]] = []
        concepts: list[tuple[str, dict[str, str]]] = []
        if gen_main and isinstance(spec.get("main"), dict):
            m = dict(spec["main"])
            m["role_type"] = "hero"  # main is always the hero shot
            concepts.append(("main", m))
        if gen_secondary:
            for sec in (spec.get("secondary") or [])[:secondary_count]:
                if isinstance(sec, dict):
                    concepts.append(("secondary", sec))

        # Concurrency: JieKou /v3/gpt-image-2-edit drops connections when
        # >= 2 simultaneous calls hit it (main bundle + secondary is enough).
        # Default 1; secondaries may run in parallel after main succeeds.
        max_workers = max(1, min(
            int(os.getenv("TKSHOP_IMAGE_GEN_CONCURRENCY", "1")),
            len(concepts) or 1,
        ))

        import tempfile
        from concurrent.futures import ThreadPoolExecutor
        tmpdir = Path(tempfile.mkdtemp())

        def _qa_ok(
            out_bytes: bytes,
            ref_hash: int,
            want: Optional[tuple[int, int]],
            check_noop: bool,
            single_source_guard: bool = False,
            *,
            slot_role: str = "secondary",
        ) -> Optional[str]:
            """Return None if the output passes QA, else a short reason.

            ``check_noop`` is skipped for merged-grid (bundle) references: there
            we WANT the output to resemble the real stickers, so an aHash close
            to the reference is acceptable, not a failure.
            """
            try:
                w, h = read_dimensions(out_bytes)
            except Exception as e:
                return f"undecodable ({type(e).__name__})"
            if min(w, h) < 512:
                return f"too small ({w}x{h})"
            out_hash = average_hash(out_bytes)
            if check_noop and hash_distance(out_hash, ref_hash) <= noop_threshold:
                return "near-identical to reference (no-op edit)"
            if slot_role != "main" and gallery_hashes:
                closest = min(hash_distance(out_hash, gh) for gh in gallery_hashes)
                if closest <= gallery_threshold:
                    return f"too similar to existing gallery (dist={closest})"
            if single_source_guard and single_source_hashes:
                closest_label = ""
                closest_dist = 999
                for label, src_hash in single_source_hashes:
                    dist = hash_distance(out_hash, src_hash)
                    if dist < closest_dist:
                        closest_dist = dist
                        closest_label = label
                single_threshold = int(os.getenv("TKSHOP_IMAGE_SINGLE_SOURCE_AHASH_DIST", "10"))
                if closest_dist <= single_threshold:
                    return (
                        "too close to one source image only "
                        f"({closest_dist}, {Path(closest_label).name})"
                    )
            return None

        def _gen_one(idx_role_concept: tuple[int, str, dict[str, str]]) -> dict[str, Any]:
            idx, role, c = idx_role_concept
            concept_label = (c.get("concept") or "").strip()[:80]
            role_type = (c.get("role_type") or ("hero" if role == "main" else "")).strip()
            # The main/hero is the operator-approved white-bg bundle: use the
            # fixed bundle prompt over the merged grid, not the model's free-form
            # scene prompt. hand_hold uses a fixed pile-background prompt.
            if role == "main":
                base_prompt = MAIN_BUNDLE_PROMPT
            elif role_type.lower() == "hand_hold":
                base_prompt = SECONDARY_HAND_HOLD_BUNDLE_BELOW_PROMPT
            else:
                base_prompt = (c.get("image_prompt") or "").strip()
            if not base_prompt:
                return {"idx": idx, "role": role, "ok": False,
                        "error": "empty image_prompt", "concept": concept_label}

            ref_label, ref_bytes, ref_hash, is_merged = _ref_for(
                role_type, idx, slot_role=role,
            )
            this_size = main_size if role == "main" else size
            this_quality = main_quality if role == "main" else secondary_quality
            want = _expected_wh(this_size)
            if role == "main" and main_from_source_path is not None:
                try:
                    out_bytes = self._render_white_canvas_image(main_from_source_path, size=this_size)
                    tmp = tmpdir / f"source_main_{idx}_{int(time.time()*1000)}.png"
                    tmp.write_bytes(out_bytes)
                    return {
                        "idx": idx,
                        "role": role,
                        "ok": True,
                        "tmp_path": tmp,
                        "concept": concept_label or "external source overview",
                        "image_prompt": (
                            "Local main image generated from the operator-uploaded "
                            f"source overview without AI redraw: {main_from_source_path.as_posix()}"
                        ),
                        "role_type": "hero",
                        "reference": main_from_source_path.as_posix(),
                        "reference_count": 1,
                    }
                except Exception as e:
                    return {
                        "idx": idx,
                        "role": role,
                        "ok": False,
                        "error": f"source overview render failed: {type(e).__name__}: {e}"[:300],
                        "concept": concept_label,
                        "image_prompt": base_prompt,
                    }

            attempts = qa_max_attempts
            last_reason = ""
            for attempt in range(1, attempts + 1):
                prompt = base_prompt
                # On retry, nudge single-subject angles to re-stage in a new
                # scene (their failure mode is a no-op echo of the sticker). For
                # merged bundles, retry by emphasizing multi-source coverage.
                if attempt > 1:
                    if "gallery" in last_reason or "similar" in last_reason:
                        prompt = (
                            base_prompt
                            + " IMPORTANT: The previous output looked too similar "
                            "to an existing listing photo. Use a COMPLETELY different "
                            "real-world scene — new props, surface, camera angle, "
                            "and lighting (e.g. gift packaging, themed still-life, "
                            "outdoor table) — not another white-background product pile."
                        )
                    elif is_merged:
                        prompt = (
                            base_prompt
                            + " IMPORTANT: the previous attempt used too little of "
                            "the source. Use visible designs from EVERY panel/source "
                            "in the reference grid. Do not output one sheet, one "
                            "panel, or one enlarged sticker."
                        )
                    else:
                        prompt = (
                            base_prompt
                            + " IMPORTANT: fully re-stage this in a NEW scene — the "
                            "output must look clearly different from the flat reference "
                            "(new background, real-world surface, perspective and lighting)."
                        )
                try:
                    out_bytes = self.router.image_edit(
                        ref_bytes,
                        prompt,
                        size=this_size,
                        quality=this_quality,
                        task="tkshop_image_design:edit",
                        related_table="tkshop_products",
                        related_id=product_id,
                    )
                except APIError as e:
                    return {"idx": idx, "role": role, "ok": False,
                            "error": str(e)[:300], "concept": concept_label,
                            "image_prompt": base_prompt}
                except Exception as e:
                    logger.exception(
                        "auto_design_images: unexpected image generation failure for %s/%s",
                        role, role_type or "?",
                    )
                    return {"idx": idx, "role": role, "ok": False,
                            "error": f"{type(e).__name__}: {e}"[:300],
                            "concept": concept_label,
                            "image_prompt": base_prompt}
                reason = _qa_ok(
                    out_bytes, ref_hash, want,
                    check_noop=not is_merged,
                    single_source_guard=(role == "main" and is_merged),
                    slot_role=role,
                )
                if reason is None:
                    tmp = tmpdir / f"ai_{role}_{idx}_{int(time.time()*1000)}.png"
                    tmp.write_bytes(out_bytes)
                    return {"idx": idx, "role": role, "ok": True,
                            "tmp_path": tmp, "concept": concept_label,
                            "image_prompt": base_prompt, "role_type": role_type,
                            "reference": ref_label,
                            "reference_count": len(main_bundle_paths) if is_merged else 1}
                last_reason = reason
                logger.warning(
                    "auto_design_images: QA rejected %s/%s (attempt %d/%d): %s",
                    role, role_type or "?", attempt, attempts, reason,
                )
            return {"idx": idx, "role": role, "ok": False,
                    "error": f"failed QA: {last_reason}", "concept": concept_label,
                    "image_prompt": base_prompt}

        try:
            t0 = time.time()
            tasks_packed = [(i, r, c) for i, (r, c) in enumerate(concepts)]
            main_tasks = [t for t in tasks_packed if t[1] == "main"]
            other_tasks = [t for t in tasks_packed if t[1] != "main"]
            logger.info(
                "auto_design_images: dispatching %d concepts "
                "(main-first, then %d secondary with workers=%d)",
                len(tasks_packed), len(other_tasks), max_workers,
            )
            gen_results: list[dict[str, Any] | None] = [None] * len(tasks_packed)
            for t in main_tasks:
                gen_results[t[0]] = _gen_one(t)
            if other_tasks:
                with ThreadPoolExecutor(max_workers=max_workers) as ex:
                    for t, r in zip(other_tasks, ex.map(_gen_one, other_tasks)):
                        gen_results[t[0]] = r
            gen_results = [r for r in gen_results if r is not None]
            logger.info(
                "auto_design_images: image gen done in %.1fs", time.time() - t0,
            )

            # Direct-mode only: if image_edit main failed, fall back to pasting
            # the parsed source overview onto a white canvas so publish is not
            # blocked. Default (model) path does not use this fallback.
            if (
                use_direct_external_main
                and main_from_source_path is not None
                and not any(r.get("ok") and r.get("role") == "main" for r in gen_results)
            ):
                try:
                    out_bytes = self._render_white_canvas_image(
                        main_from_source_path, size=main_size,
                    )
                    tmp = tmpdir / f"fallback_source_main_{int(time.time() * 1000)}.png"
                    tmp.write_bytes(out_bytes)
                    gen_results.insert(0, {
                        "idx": -1,
                        "role": "main",
                        "ok": True,
                        "tmp_path": tmp,
                        "concept": "external source overview",
                        "image_prompt": (
                            "Local main image generated from the operator-uploaded "
                            f"source overview without AI redraw: "
                            f"{main_from_source_path.as_posix()}"
                        ),
                        "role_type": "hero",
                        "reference": main_from_source_path.as_posix(),
                        "reference_count": 1,
                    })
                    logger.info(
                        "auto_design_images: recovered main from external source "
                        "overview for product #%d",
                        product_id,
                    )
                except Exception as e:
                    logger.warning(
                        "auto_design_images: external overview main fallback "
                        "failed for product #%d: %s",
                        product_id, e,
                    )

            # DB writes sequential to keep sort_order monotonic per role.
            for r in gen_results:
                if not r.get("ok"):
                    results.append({k: v for k, v in r.items()
                                    if k not in ("tmp_path", "image_prompt")})
                    continue
                img_id = self.add_product_image(
                    product_id, src_path=r["tmp_path"],
                    role=r["role"], source="ai",
                    ai_prompt=r["image_prompt"],
                )
                results.append({"role": r["role"], "ok": True,
                                "image_id": img_id, "concept": r["concept"],
                                "role_type": r.get("role_type", ""),
                                "reference": r.get("reference", ""),
                                "reference_count": r.get("reference_count", 0)})
        finally:
            try:
                for f in tmpdir.iterdir():
                    f.unlink()
                tmpdir.rmdir()
            except Exception:
                pass

        generated_ok_roles = {r["role"] for r in results if r.get("ok")}
        new_image_ids = [
            int(r["image_id"]) for r in results
            if r.get("ok") and r.get("image_id")
        ]
        if generated_ok_roles:
            if append_secondary:
                self.mirror_listing_ai_to_local_master(
                    product_id,
                    append_only=True,
                    image_ids=new_image_ids or None,
                )
            else:
                self.mirror_listing_ai_to_local_master(
                    product_id, roles=generated_ok_roles,
                )

        return {
            "product_id": product_id,
            "generated": sum(1 for r in results if r.get("ok")),
            "failed": sum(1 for r in results if not r.get("ok")),
            "results": results,
            "reference_preview": (
                reference_grid_path.as_posix()
                if reference_grid_path else main_bundle_paths[0].as_posix()
            ),
            "reference_count": len(main_bundle_paths),
            "reference_kind": main_bundle_ref_label,
            "spec": spec,
        }

    # ------------------------------------------------------------------
    # Batch: create + AI-fill (no publish) from pack manager multi-select
    # ------------------------------------------------------------------

    def batch_create_products_from_packs(
        self,
        pack_ids: list[int],
        *,
        shop: Optional[str] = None,
        generate_detail: bool = True,
        generate_images: bool = True,
        secondary_count: int = 3,
        language: str = "en",
        size: str = "1024x1024",
        max_workers: int = 3,
        skip_if_any_product_exists: bool = True,
    ) -> dict[str, Any]:
        """Bulk pipeline: for each pack id, create a draft tkshop_products row
        in ``shop`` and (optionally) run ``generate_detail`` then
        ``auto_design_images`` to fill in title/description/SKU/main image/
        secondary images. **Does not publish** — leaves rows in
        ``publish_status='draft'`` so the operator can review before sending
        to TikTok.

        Concurrency: up to ``max_workers`` packs are processed in parallel via a
        ThreadPoolExecutor (default 3 to stay under OpenAI rate limits).
        Within a single pack the steps are serial — detail → images — so a
        single rate-limit hit only stalls that pack.

        ``skip_if_any_product_exists`` matches the pack-manager UI rule: once
        a pack has any tkshop_products row it disappears from the
        batch-generate candidate set, so re-running this with the same input
        is idempotent.

        Returns ``{created:[{pack_id, product_id, detail_ok, images_ok,
        error}], skipped:[{pack_id, reason}], elapsed_s}``.
        """
        from concurrent.futures import ThreadPoolExecutor, as_completed

        target_shop = (shop or TKSHOP_DEFAULT_SHOP).strip() or TKSHOP_DEFAULT_SHOP
        t0 = time.time()
        # De-dup while preserving order — operators sometimes hit submit twice.
        seen: set[int] = set()
        ordered: list[int] = []
        for pid in pack_ids:
            try:
                p = int(pid)
            except (TypeError, ValueError):
                continue
            if p in seen:
                continue
            seen.add(p)
            ordered.append(p)

        # Pre-filter: skip packs that already have any product row. Doing this
        # up-front (instead of inside each worker) lets the UI summary count
        # be honest even if the worker pool serializes badly.
        to_process: list[int] = []
        skipped: list[dict] = []
        with _open_db(self.db_path) as conn:
            for pid in ordered:
                pack_exists = conn.execute(
                    "SELECT 1 FROM packs WHERE id = ?", (pid,),
                ).fetchone()
                if not pack_exists:
                    skipped.append({"pack_id": pid, "reason": "pack not found"})
                    continue
                if skip_if_any_product_exists:
                    has = conn.execute(
                        "SELECT 1 FROM tkshop_products WHERE pack_id = ? LIMIT 1",
                        (pid,),
                    ).fetchone()
                    if has:
                        skipped.append({"pack_id": pid, "reason": "product already exists"})
                        continue
                to_process.append(pid)

        created: list[dict] = []

        def _one(pack_id: int) -> dict:
            entry: dict = {"pack_id": pack_id, "product_id": None,
                            "detail_ok": False, "images_ok": False, "error": ""}
            try:
                product_id = self.create_product_from_pack(pack_id, shop=target_shop)
                entry["product_id"] = product_id
            except Exception as e:
                entry["error"] = f"create failed: {e}"[:300]
                return entry
            if generate_detail:
                try:
                    self.generate_detail(product_id)
                    entry["detail_ok"] = True
                except Exception as e:
                    entry["error"] = f"detail failed: {e}"[:300]
                    # Continue to images anyway — they don't depend on text.
            if generate_images:
                try:
                    self.auto_design_images(
                        product_id,
                        secondary_count=secondary_count,
                        language=language,
                        size=size,
                        replace_existing_ai=True,
                    )
                    entry["images_ok"] = True
                except Exception as e:
                    prev = entry["error"]
                    entry["error"] = (prev + " | " if prev else "") + f"images failed: {e}"[:300]
            return entry

        if to_process:
            workers = max(1, min(max_workers, len(to_process)))
            with ThreadPoolExecutor(max_workers=workers,
                                     thread_name_prefix="tkshop-batch") as ex:
                futures = {ex.submit(_one, pid): pid for pid in to_process}
                for fut in as_completed(futures):
                    pack_id = futures[fut]
                    try:
                        created.append(fut.result())
                    except Exception as e:
                        created.append({
                            "pack_id": pack_id, "product_id": None,
                            "detail_ok": False, "images_ok": False,
                            "error": f"worker crashed: {e}"[:300],
                        })

        elapsed = round(time.time() - t0, 1)
        logger.info(
            "batch_create_products_from_packs: shop=%s requested=%d processed=%d skipped=%d in %.1fs",
            target_shop, len(ordered), len(created), len(skipped), elapsed,
        )
        return {"created": created, "skipped": skipped, "elapsed_s": elapsed,
                "shop": target_shop}

    # ------------------------------------------------------------------
    # C.3 publish — POST to multi-channel-api sticker_publish endpoint
    # ------------------------------------------------------------------

    def publish(
        self,
        product_id: int,
        *,
        dry_run: bool = False,
        auto_activate: bool = True,
    ) -> dict[str, Any]:
        """POST a sticker_publish payload to multi-channel-api.

        Endpoint: ``POST {TKSHOP_SERVER_URL}/api/v1/tiktok/products/sticker_publish``

        Request shape (FIXED contract — see the multi-channel-api repo)::

          {
            "title": str,
            "description_html": str,
            "category_id": str,
            "seller_sku": str,
            "quantity": int,
            "image_urls": [str, ...]
          }

        Success response::
            {"ok": true, "product_id": "...", "sku_id": "...",
             "seller_sku": "...", "status": "LIVE"}

        Failure response::
            {"ok": false, "error_code": "TT_xxxx",
             "error_message": "...", "field_hints": ["title", ...]}

        ``dry_run=True`` builds the payload + writes the publish_log row
        but does NOT make the HTTP call.
        """
        product = self._build_publish_payload(product_id)
        product["auto_activate"] = bool(auto_activate)

        attempt_idx = 1
        with _open_db(self.db_path) as conn:
            attempt_idx = (conn.execute(
                "SELECT COALESCE(MAX(attempt_idx), 0) + 1 FROM tkshop_publish_logs WHERE product_id = ?",
                (product_id,),
            ).fetchone()[0]) or 1

        shop_name = _get_product_shop(self.db_path, product_id)
        endpoint = f"{TKSHOP_SERVER_URL.rstrip('/')}/api/v1/tiktok/products/sticker_publish"
        if dry_run:
            self._log_publish_attempt(
                product_id, attempt_idx, endpoint, product,
                response={"_dry_run": True, "_shop": shop_name},
                success=True,
            )
            return {"ok": True, "dry_run": True, "endpoint": endpoint,
                    "payload": product, "shop": shop_name}

        resp_json: dict[str, Any] = {}
        success = False
        error_code = ""
        error_message = ""
        field_hints: list[str] = []
        tiktok_product_id = ""
        tiktok_sku_id = ""
        tiktok_status = ""
        with _open_db(self.db_path) as conn:
            conn.execute(
                "UPDATE tkshop_products SET publish_status = 'publishing' WHERE id = ?",
                (product_id,),
            )
            conn.commit()
        try:
            resp = requests.post(endpoint, json=product,
                                 params={"shop": shop_name},
                                 timeout=TKSHOP_SERVER_TIMEOUT)
            try:
                resp_json = resp.json()
            except Exception:
                resp_json = {"_raw_text": resp.text[:500]}
            # multi-channel-api wraps business payload in
            # {success, message, data: {...}}. Unwrap so downstream code
            # reads the inner contract directly.
            if isinstance(resp_json, dict) and "data" in resp_json and isinstance(resp_json["data"], dict) and "ok" in resp_json["data"]:
                resp_json = resp_json["data"]
            if 200 <= resp.status_code < 300 and resp_json.get("ok") is True:
                success = True
                tiktok_product_id = str(resp_json.get("product_id") or "")
                tiktok_sku_id = str(resp_json.get("sku_id") or "")
                tiktok_status = str(resp_json.get("status") or "").strip()
                # multi-channel-api now returns ok=true even when activate
                # fails (product was created → not a payload bug → DON'T
                # let self-heal duplicate the listing). Surface the activate
                # error via error_message so the UI can show it.
                ae = resp_json.get("activate_error") or {}
                if isinstance(ae, dict) and ae:
                    error_code = str(ae.get("code") or "TT_ACTIVATE_FAILED")
                    error_message = (
                        f"商品已创建（id={tiktok_product_id}）但 activate 失败 — "
                        f"{ae.get('note') or ''} 详情：{ae.get('message') or ''}"
                    )[:500]
            else:
                # Either non-2xx HTTP or {ok:false} body. Surface whichever
                # error info we have.
                error_code = str(resp_json.get("error_code") or resp.status_code)
                error_message = str(
                    resp_json.get("error_message")
                    or resp_json.get("error")
                    or resp.text[:200]
                )
                hints_raw = resp_json.get("field_hints") or []
                if isinstance(hints_raw, list):
                    field_hints = [str(h) for h in hints_raw]
        except requests.RequestException as e:
            error_code = "NETWORK"
            error_message = str(e)[:300]
        except Exception as e:
            error_code = "INTERNAL"
            error_message = f"{type(e).__name__}: {e}"[:300]

        # Persist any field_hints into the response body so the publish
        # log row carries them through (response_payload is queried by UI).
        if not success and field_hints and isinstance(resp_json, dict):
            resp_json.setdefault("field_hints", field_hints)

        self._log_publish_attempt(
            product_id, attempt_idx, endpoint, product,
            response=resp_json, success=success,
            error_code=error_code, error_message=error_message,
        )

        with _open_db(self.db_path) as conn:
            if success:
                # Honor TikTok-side status. Single mapping shared with sync.
                # Fall back to the raw value (or 'pending') if unmapped, since a
                # successful publish must still land in a non-empty status.
                local_status = (self._local_status_from_tiktok(tiktok_status or "")
                                or (tiktok_status or "").strip().lower() or "pending")
                conn.execute(
                    """
                    UPDATE tkshop_products
                       SET publish_status = ?,
                           tiktok_product_id = ?,
                           tiktok_sku_id = ?,
                           published_at = ?
                     WHERE id = ?
                    """,
                    (local_status, tiktok_product_id, tiktok_sku_id,
                     int(time.time()), product_id),
                )
            else:
                conn.execute(
                    "UPDATE tkshop_products SET publish_status = 'failed' WHERE id = ?",
                    (product_id,),
                )
            conn.commit()

        # After a successful create, auto-apply the standing discount + default
        # free shipping. Both best-effort — never flip a successful publish to
        # failed.
        discount_result = None
        free_shipping_result = None
        if success and tiktok_product_id:
            discount_result = self.apply_discount(product_id)
            if TKSHOP_DEFAULT_FREE_SHIPPING:
                free_shipping_result = self.apply_free_shipping(product_id)

        return {
            "ok": success,
            "tiktok_product_id": tiktok_product_id,
            "tiktok_sku_id": tiktok_sku_id,
            "tiktok_status": tiktok_status,
            "error_code": error_code,
            "error_message": error_message,
            "field_hints": field_hints,
            "attempt_idx": attempt_idx,
            "discount": discount_result,
            "free_shipping": free_shipping_result,
        }

    def apply_free_shipping(self, product_id: int) -> dict[str, Any]:
        """Best-effort: put a published product on a SHIPPING_DISCOUNT (免邮)
        promotion via the middle layer. NEVER raises on a remote failure (the
        listing stays live); only raises if the product row is missing.

        NOTE: the middle layer's SHIPPING_DISCOUNT create currently needs the
        discount_threshold schema confirmed (see mca promotions.py); until then
        this returns {applied: False, reason: 'rejected'} and logs the TikTok
        error — publish is unaffected.
        """
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                "SELECT tiktok_product_id, title, shop FROM tkshop_products WHERE id = ?",
                (product_id,),
            ).fetchone()
        if not row:
            raise ValueError(f"product #{product_id} not found")
        tt_id = (row["tiktok_product_id"] or "").strip()
        if not tt_id:
            return {"applied": False, "reason": "not_on_platform"}
        shop = row["shop"] or TKSHOP_DEFAULT_SHOP
        endpoint = (
            f"{TKSHOP_SERVER_URL.rstrip('/')}"
            f"/api/v1/tiktok/products/{tt_id}/free-shipping"
        )
        try:
            resp = requests.post(endpoint, json={}, params={"shop": shop},
                                 timeout=TKSHOP_SERVER_TIMEOUT)
            rj = resp.json() if resp.text else {}
        except requests.RequestException as e:
            logger.warning("free-shipping network error for product #%s: %s", product_id, e)
            return {"applied": False, "reason": "network", "error": str(e)[:200]}
        except Exception as e:
            return {"applied": False, "reason": "internal", "error": f"{type(e).__name__}: {e}"[:200]}
        if isinstance(rj, dict) and isinstance(rj.get("data"), dict) and "ok" in rj["data"]:
            rj = rj["data"]
        if rj.get("ok") is True or rj.get("success") is True:
            logger.info("free-shipping applied: product #%s tiktok=%s activity=%s",
                        product_id, tt_id, rj.get("activity_id") or "")
            return {"applied": True, "activity_id": rj.get("activity_id") or ""}
        msg = str(rj.get("error_message") or rj.get("message") or "")
        logger.warning("free-shipping rejected for product #%s: %s", product_id, msg[:200])
        return {"applied": False, "reason": "rejected", "error": msg[:200]}

    def _create_product_discount_remote(
        self,
        tiktok_product_id: str,
        *,
        percent: float,
        shop: str,
        sku_id: str = "",
        title: str = "",
    ) -> dict[str, Any]:
        """POST a discount request to multi-channel-api.

        Endpoint (repo B must implement — see
        docs/tiktok_product_shipping_discount_contract.md). repo B turns this
        into the TikTok Promotion two-step (create activity + add product)::

            POST {TKSHOP_SERVER_URL}/api/v1/tiktok/products/{id}/discount?shop=

        Body::
            {"discount_percent": 40.0, "sku_id": "...", "product_title": "...",
             "activity_type": "DIRECT_DISCOUNT", "begin_time": <epoch>,
             "end_time": <epoch>}
        Success:: {"ok": true, "activity_id": "...", "discount_percent": 40.0}
        Failure:: {"ok": false, "error_code": "...", "error_message": "..."}
        """
        now = int(time.time())
        endpoint = (
            f"{TKSHOP_SERVER_URL.rstrip('/')}"
            f"/api/v1/tiktok/products/{tiktok_product_id}/discount"
        )
        body = {
            "discount_percent": float(percent),
            "sku_id": sku_id or "",
            "product_title": title or "",
            # Activity context so repo B can build the Promotion activity. A long
            # window emulates a "standing" discount.
            "activity_type": TKSHOP_DISCOUNT_ACTIVITY_TYPE,
            "begin_time": now,
            "end_time": now + TKSHOP_DISCOUNT_WINDOW_DAYS * 86400,
        }
        resp = requests.post(
            endpoint, json=body, params={"shop": shop},
            timeout=TKSHOP_SERVER_TIMEOUT,
        )
        try:
            rj = resp.json()
        except Exception:
            rj = {"_raw_text": resp.text[:500]}
        # Unwrap the {success, message, data:{ok:...}} envelope when present.
        if (isinstance(rj, dict) and isinstance(rj.get("data"), dict)
                and "ok" in rj["data"]):
            rj = rj["data"]
        rj.setdefault("_http_status", resp.status_code)
        return rj

    def _set_discount_state(
        self,
        product_id: int,
        *,
        status: str,
        percent: Optional[float] = None,
        activity_id: Optional[str] = None,
        applied_at: Optional[int] = None,
    ) -> None:
        """Persist discount tracking columns. Defensive: a missing-column error
        (migration 022 not applied) is logged, not raised, so publish/discount
        never breaks on an un-migrated DB."""
        sets = ["discount_status = ?"]
        params: list[Any] = [status]
        if percent is not None:
            sets.append("discount_percent = ?"); params.append(float(percent))
        if activity_id is not None:
            sets.append("discount_activity_id = ?"); params.append(activity_id)
        if applied_at is not None:
            sets.append("discount_applied_at = ?"); params.append(applied_at)
        params.append(product_id)
        try:
            with _open_db(self.db_path) as conn:
                conn.execute(
                    f"UPDATE tkshop_products SET {', '.join(sets)} WHERE id = ?",
                    tuple(params),
                )
                conn.commit()
        except sqlite3.OperationalError as e:
            logger.warning(
                "discount state persist skipped (run migration 022?): %s", e)

    def apply_discount(
        self,
        product_id: int,
        *,
        percent: Optional[float] = None,
        force: bool = False,
    ) -> dict[str, Any]:
        """Put a published product on an N% discount via the middle layer.

        Resolves the product's TikTok id / shop / sku from the DB, calls the
        middle layer, and persists the outcome to the discount_* columns.
        NEVER raises on a remote failure (the listing stays live); only raises
        if the product row itself is missing.

        - ``percent``: override; defaults to ``TKSHOP_AUTO_DISCOUNT_PERCENT``.
        - ``force``: re-apply even if already marked applied (idempotent skip
          otherwise).
        """
        pct = TKSHOP_AUTO_DISCOUNT_PERCENT if percent is None else float(percent)
        if pct <= 0:
            return {"applied": False, "reason": "disabled", "percent": pct}

        with _open_db(self.db_path) as conn:
            row = conn.execute(
                "SELECT tiktok_product_id, tiktok_sku_id, title, shop, "
                "discount_status FROM tkshop_products WHERE id = ?",
                (product_id,),
            ).fetchone()
        if not row:
            raise ValueError(f"product #{product_id} not found")

        tt_id = (row["tiktok_product_id"] or "").strip()
        if not tt_id:
            return {"applied": False, "reason": "not_on_platform", "percent": pct}
        if not force and (row["discount_status"] or "") == "applied":
            return {"applied": True, "reason": "already_applied",
                    "percent": pct, "skipped": True}

        shop = row["shop"] or TKSHOP_DEFAULT_SHOP
        self._set_discount_state(product_id, status="pending", percent=pct)

        try:
            rj = self._create_product_discount_remote(
                tt_id, percent=pct, shop=shop,
                sku_id=row["tiktok_sku_id"] or "", title=row["title"] or "",
            )
        except requests.RequestException as e:
            self._set_discount_state(product_id, status="failed", percent=pct)
            logger.warning("discount network error for product #%s: %s",
                           product_id, e)
            return {"applied": False, "reason": "network",
                    "error": str(e)[:200], "percent": pct}
        except Exception as e:  # pragma: no cover - defensive
            self._set_discount_state(product_id, status="failed", percent=pct)
            logger.warning("discount failed for product #%s: %s", product_id, e)
            return {"applied": False, "reason": "internal",
                    "error": f"{type(e).__name__}: {e}"[:200], "percent": pct}

        ok = bool(rj.get("ok") is True or rj.get("success") is True)
        if ok:
            activity_id = str(rj.get("activity_id") or "")
            self._set_discount_state(
                product_id, status="applied", percent=pct,
                activity_id=activity_id, applied_at=int(time.time()),
            )
            logger.info("discount applied: product #%s tiktok=%s %.0f%% activity=%s",
                        product_id, tt_id, pct, activity_id)
            return {"applied": True, "percent": pct, "activity_id": activity_id}

        msg = str(rj.get("error_message") or rj.get("message")
                  or rj.get("_raw_text") or "")
        self._set_discount_state(product_id, status="failed", percent=pct)
        logger.warning("discount rejected for product #%s: %s", product_id, msg[:200])
        return {"applied": False, "reason": "rejected",
                "error": msg[:200], "percent": pct}

    def publish_with_self_heal(
        self,
        product_id: int,
        *,
        max_retries: Optional[int] = None,
        dry_run: bool = False,
        auto_activate: bool = True,
    ) -> dict[str, Any]:
        """Publish, and if it fails ask the AI to fix the payload + retry.

        ``max_retries`` is the total attempt cap (default
        ``TKSHOP_PUBLISH_MAX_RETRIES``). The first attempt counts as 1, so
        the AI fix loop runs at most ``max_retries - 1`` times.

        Returns the last attempt's result dict, plus an ``attempts`` list
        carrying every individual try (so the UI can show what AI changed
        and why).
        """
        if not TKSHOP_PUBLISH_AUTO_FIX_DEFAULT and max_retries is None:
            # Auto-fix disabled at the env layer and caller didn't override.
            # Behave exactly like a single publish().
            single = self.publish(product_id, dry_run=dry_run, auto_activate=auto_activate)
            single["attempts"] = [single]
            return single

        cap = max_retries if max_retries is not None else TKSHOP_PUBLISH_MAX_RETRIES_DEFAULT
        cap = max(1, int(cap))

        # Guard: if this product already has a tiktok_product_id, the
        # listing exists on TikTok — never let self-heal POST sticker_publish
        # again, that would create a duplicate. Use update_on_platform / activate
        # instead.
        with _open_db(self.db_path) as conn:
            existing_tt = (conn.execute(
                "SELECT tiktok_product_id FROM tkshop_products WHERE id = ?",
                (product_id,),
            ).fetchone() or [""])[0] or ""
        if existing_tt and not dry_run:
            logger.warning(
                "publish_with_self_heal: product #%d already has tiktok_product_id=%s; "
                "skipping create-loop. Use /update_on_platform or /activate instead.",
                product_id, existing_tt,
            )
            return {
                "ok": True,
                "tiktok_product_id": existing_tt,
                "message": "product already exists on TikTok — no new create issued",
                "attempts": [],
                "skipped_create": True,
            }

        attempts: list[dict[str, Any]] = []
        last: dict[str, Any] = {}
        for i in range(cap):
            last = self.publish(product_id, dry_run=dry_run, auto_activate=auto_activate)
            attempts.append(last)
            if last.get("ok"):
                break
            # If publish() returned a tiktok_product_id (rare, but possible if
            # the response shape carried it on a "soft" failure), treat as
            # done — we've created the listing, don't recreate.
            if last.get("tiktok_product_id"):
                logger.info(
                    "publish_with_self_heal: stopping retry — product was created "
                    "(tiktok_product_id=%s) even though ok=false",
                    last.get("tiktok_product_id"),
                )
                last["skipped_further_retries"] = True
                break
            # NETWORK / timeout / non-payload errors → DO NOT retry. The
            # product may have been created server-side; recreating would
            # duplicate. Operator should sync platform status manually.
            ec = (last.get("error_code") or "").upper()
            if ec in ("NETWORK", "INTERNAL", "TT_IMAGE_UPLOAD") or ec.isdigit() and int(ec) >= 500:
                logger.warning(
                    "publish_with_self_heal: stopping retry — error_code=%s "
                    "is not a payload-fixable issue (likely transport / upstream).",
                    ec,
                )
                last["skipped_further_retries"] = True
                last["skip_reason"] = f"non-retryable error_code={ec}"
                break
            if i == cap - 1:
                # No retry budget left.
                break
            # Ask the AI to patch the payload, then loop.
            try:
                fix = self._ai_fix_payload(
                    product_id,
                    error_code=str(last.get("error_code") or ""),
                    error_message=str(last.get("error_message") or ""),
                    field_hints=list(last.get("field_hints") or []),
                )
            except Exception as e:
                logger.warning(
                    "self-heal AI fix failed for product #%d attempt %d: %s",
                    product_id, i + 1, e,
                )
                last["self_heal_error"] = str(e)[:300]
                break
            if not fix.get("changed_fields"):
                # AI didn't propose any change; pointless to retry the
                # exact same payload.
                last["self_heal_error"] = "AI returned no changes"
                break

        last["attempts"] = attempts
        return last

    def _ai_fix_payload(
        self,
        product_id: int,
        *,
        error_code: str,
        error_message: str,
        field_hints: list[str],
    ) -> dict[str, Any]:
        """Call the AI to rewrite the failing fields, then apply the patch.

        Returns ``{"changed_fields": [...], "rationale": "..."}``.
        """
        failed_payload = self._build_publish_payload(product_id)
        prompt = build_self_heal_prompt(
            failed_payload=failed_payload,
            error_code=error_code,
            error_message=error_message,
            field_hints=field_hints,
        )
        # extract_json is happy to do a one-shot JSON-mode call; the
        # "source text" is just the prompt itself (the AI rewrites the
        # fields based on the failure context).
        patch = self.router.extract_json(
            prompt,
            schema=SELF_HEAL_EXTRACT_SCHEMA,
            instructions=SELF_HEAL_INSTRUCTIONS + "\n\n" + SELF_HEAL_SYSTEM_PROMPT,
            max_retries=1,
            task="tkshop_publish:self_heal",
            related_table="tkshop_products",
            related_id=product_id,
        )

        new_title = (patch.get("title") or "").strip()
        new_desc = patch.get("description_html") or ""
        new_sku = (patch.get("seller_sku") or "").strip().upper()
        rationale = (patch.get("rationale") or "").strip()

        changed: list[str] = []
        if new_title:
            self.update_detail_manual(product_id, title=_clamp_title(new_title))
            changed.append("title")
        if new_desc:
            self.update_detail_manual(product_id, description_html=new_desc)
            changed.append("description_html")
        if new_sku and _is_valid_seller_sku(new_sku):
            self.set_seller_sku(product_id, new_sku)
            changed.append("seller_sku")

        # Append a one-line diff entry so operators can audit what AI did.
        diff_line = (
            f"[{int(time.time())}] code={error_code or '?'} "
            f"changed={','.join(changed) or '(none)'} :: {rationale[:200]}"
        )
        with _open_db(self.db_path) as conn:
            conn.execute(
                """
                UPDATE tkshop_products
                   SET auto_fix_attempts = COALESCE(auto_fix_attempts, 0) + 1,
                       last_fix_diff = CASE
                           WHEN COALESCE(last_fix_diff, '') = '' THEN ?
                           ELSE last_fix_diff || char(10) || ?
                       END
                 WHERE id = ?
                """,
                (diff_line, diff_line, product_id),
            )
            conn.commit()

        return {
            "changed_fields": changed,
            "rationale": rationale,
            "diff_line": diff_line,
        }

    def _build_publish_payload(self, product_id: int) -> dict[str, Any]:
        """Build the multi-channel-api ``sticker_publish`` request body.

        Shape (do not change without the multi-channel-api side):
            {
              "title": str,
              "description_html": str,
              "category_id": str,
              "seller_sku": str,
              "quantity": int,
              "image_urls": [str, ...],
              "image_paths": [str, ...],
              "auto_activate": bool
            }
        """
        with _open_db(self.db_path) as conn:
            pr = conn.execute(
                """
                SELECT pr.id, pr.shop, pr.title, pr.description_html, pr.category_id,
                       pr.seller_sku, pr.tiktok_product_id, pr.tiktok_sku_id,
                       p.display_name, p.pack_uid, p.total_stickers,
                       lp.default_quantity AS master_default_quantity
                  FROM tkshop_products pr
                  JOIN packs p ON p.id = pr.pack_id
             LEFT JOIN local_products lp ON lp.id = pr.local_product_id
                 WHERE pr.id = ?
                """, (product_id,),
            ).fetchone()
            if not pr:
                raise ValueError(f"product #{product_id} not found")
            imgs = conn.execute(
                """
                SELECT role, local_path, sort_order
                  FROM tkshop_product_images
                 WHERE product_id = ?
                 ORDER BY role = 'main' DESC, sort_order ASC, id ASC
                """, (product_id,),
            ).fetchall()

        # Build image_urls AND image_paths. Same-host deployment: mca prefers
        # image_paths (zero HTTP roundtrip — no ReadError on busy sticker app).
        # image_urls are kept as fallback for off-box deployments.
        image_urls: list[str] = []
        image_paths: list[str] = []
        for i in imgs:
            local = i["local_path"] or ""
            if not local:
                continue
            url = _local_path_to_public_url(local)
            if url:
                image_urls.append(url)
            # Absolute path so mca can resolve regardless of its own cwd.
            # Skip non-ASCII paths: mca's open() chokes on them ("ReadError"),
            # and it does not fall back to image_urls. Letting the wrapper
            # download via URL works for Unicode paths.
            abs_path = local if os.path.isabs(local) else os.path.abspath(local)
            staged = _stage_path_for_mca(abs_path)
            if staged and os.path.isfile(staged):
                image_paths.append(staged)

        # Quantity (initial inventory written to TikTok). Prefer the master's
        # per-product default_quantity (set via the 本地产品 batch-stock tool);
        # fall back to the env-wide TKSHOP_DEFAULT_QUANTITY when unset (NULL).
        quantity = int(TKSHOP_DEFAULT_QUANTITY)
        with _open_db(self.db_path) as conn:
            mq = conn.execute(
                "SELECT lp.default_quantity FROM local_products lp "
                "JOIN tkshop_products pr ON pr.local_product_id = lp.id "
                "WHERE pr.id = ?",
                (product_id,),
            ).fetchone()
        if mq is not None and mq["default_quantity"] is not None:
            try:
                quantity = max(0, int(mq["default_quantity"]))
            except (TypeError, ValueError):
                pass

        seller_sku = pr["seller_sku"] or ""
        if not seller_sku:
            with _open_db(self.db_path) as conn:
                # Fallback if AI/extract step never populated it. Prefer the
                # master's SKU so this listing matches the local product (and
                # any sibling shop's listing); only compute a fresh default if
                # the master is also blank.
                master_sku_row = conn.execute(
                    "SELECT lp.seller_sku FROM local_products lp "
                    "JOIN tkshop_products pr ON pr.local_product_id = lp.id "
                    "WHERE pr.id = ?",
                    (product_id,),
                ).fetchone()
                seller_sku = (
                    (master_sku_row["seller_sku"] or "").strip()
                    if master_sku_row else ""
                )
                if not seller_sku:
                    seller_sku = compute_default_seller_sku(
                        pack_uid=pr["pack_uid"] or "",
                        total_stickers=pr["total_stickers"] or 0,
                        pack_id=product_id,
                        shop=pr["shop"] or TKSHOP_DEFAULT_SHOP,
                    )
                    seller_sku = _ensure_unique_seller_sku(
                        conn, seller_sku, exclude_product_id=product_id,
                        shop=pr["shop"] or TKSHOP_DEFAULT_SHOP,
                    )
                # Persist so future retries/self-heal see the same value.
                conn.execute(
                    "UPDATE tkshop_products SET seller_sku = ? WHERE id = ?",
                    (seller_sku, product_id),
                )
                conn.commit()

        # Defensive clamp: TikTok / multi-channel-api enforce 100-char title.
        return {
            "title": _clamp_title(pr["title"] or ""),
            "description_html": pr["description_html"] or "",
            "category_id": pr["category_id"] or "928016",
            "seller_sku": seller_sku,
            "quantity": quantity,
            "image_urls": image_urls,
            "image_paths": image_paths,
            "auto_activate": True,  # caller may override
        }

    def _log_publish_attempt(
        self,
        product_id: int,
        attempt_idx: int,
        endpoint: str,
        request_payload: dict,
        *,
        response: dict,
        success: bool,
        error_code: str = "",
        error_message: str = "",
    ) -> None:
        with _open_db(self.db_path) as conn:
            conn.execute(
                """
                INSERT INTO tkshop_publish_logs
                    (product_id, attempt_idx, api_endpoint,
                     request_payload, response_payload, success,
                     error_code, error_message, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    product_id, attempt_idx, endpoint,
                    json.dumps(request_payload, ensure_ascii=False),
                    json.dumps(response, ensure_ascii=False),
                    1 if success else 0,
                    error_code, error_message[:500],
                    int(time.time()),
                ),
            )
            conn.commit()

    # ------------------------------------------------------------------
    # Platform browse + push update (already-published products)
    # ------------------------------------------------------------------

    def list_platform_products(
        self,
        *,
        page_size: int = 20,
        page_token: str = "",
        status: Optional[str] = None,
    ) -> dict[str, Any]:
        """Search products that exist on TikTok Shop via multi-channel-api
        ``POST /api/v1/tiktok/products/search``. Returns unwrapped payload.

        The multi-channel-api response is the ``ApiResponse`` envelope
        ``{success, message, data, detail?}``. ``success=false`` typically
        means the upstream TikTok call failed (e.g. expired access token).
        We surface ``message`` + ``detail`` as ``error`` so the UI shows
        a real reason instead of "未知错误".
        """
        endpoint = f"{TKSHOP_SERVER_URL.rstrip('/')}/api/v1/tiktok/products/search"
        body: dict[str, Any] = {"page_size": int(page_size), "page_token": page_token or ""}
        if status:
            body["status"] = status
        try:
            resp = requests.post(endpoint, json=body, timeout=TKSHOP_SERVER_TIMEOUT)
            data = resp.json() if resp.text else {}
        except requests.RequestException as e:
            return {"ok": False, "error": f"{type(e).__name__}: {e}",
                    "products": [], "endpoint": endpoint}

        envelope_success = bool(data.get("success", True)) if isinstance(data, dict) else True
        message = (data.get("message") or "") if isinstance(data, dict) else ""
        detail  = (data.get("detail")  or "") if isinstance(data, dict) else ""

        # Envelope-level failure (HTTP 200 but success=false, often expired token)
        if not envelope_success or not resp.ok:
            err_parts = [message or f"HTTP {resp.status_code}"]
            if detail:
                err_parts.append(detail[:300])
            return {
                "ok": False,
                "error": " | ".join(err_parts),
                "products": [],
                "endpoint": endpoint,
            }

        inner = data.get("data") if isinstance(data, dict) else None
        if not isinstance(inner, dict):
            inner = data if isinstance(data, dict) else {}
        return {
            "ok": True,
            "raw": inner,
            "products": inner.get("products") or inner.get("product_list") or [],
            "next_page_token": inner.get("next_page_token") or "",
            "total_count": inner.get("total_count") or 0,
            "endpoint": endpoint,
        }

    def get_platform_product(self, tiktok_product_id: str,
                              shop: str | None = None) -> dict[str, Any]:
        """Fetch a product from TikTok via the wrapper. ``shop`` selects the
        credential set when present; otherwise the wrapper's default shop is
        used (single-shop deployments don't need to pass it)."""
        endpoint = (
            f"{TKSHOP_SERVER_URL.rstrip('/')}"
            f"/api/v1/tiktok/products/{tiktok_product_id}"
        )
        params = {"shop": shop} if shop else None
        try:
            resp = requests.get(endpoint, params=params, timeout=TKSHOP_SERVER_TIMEOUT)
            data = resp.json() if resp.text else {}
        except requests.RequestException as e:
            return {"ok": False, "error": f"{type(e).__name__}: {e}"}
        inner = data.get("data") if isinstance(data, dict) else None
        if not isinstance(inner, dict):
            inner = data if isinstance(data, dict) else {}
        return {"ok": resp.ok, "data": inner, "endpoint": endpoint}

    def update_on_platform(self, product_id: int) -> dict[str, Any]:
        """Push local edits to TikTok for an already-published product.
        Reuses the same payload builder as ``publish``.

        Preserves the current platform status:
          * 'published' / activate aliases → save_mode=LISTING (stay live)
          * 'draft_on_platform'            → save_mode=AS_DRAFT (stay draft)
          * 'seller_deactivated'           → save_mode=LISTING then re-deactivate
            (TikTok's PUT cannot keep a product suspended; we PUT live then
            immediately call deactivate to restore the suspended state.)
        """
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                "SELECT tiktok_product_id, publish_status FROM tkshop_products WHERE id = ?",
                (product_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"product #{product_id} not found")
            tt_id = (row["tiktok_product_id"] or "").strip()
            prev_status = (row["publish_status"] or "").lower()
        if not tt_id:
            return {"ok": False, "error_code": "NOT_PUBLISHED",
                    "error_message": "product has no tiktok_product_id; publish first"}

        payload = self._build_publish_payload(product_id)
        # Map current local status → desired auto_activate (= save_mode).
        payload["auto_activate"] = (prev_status != "draft_on_platform")
        shop_name = _get_product_shop(self.db_path, product_id)
        endpoint = (
            f"{TKSHOP_SERVER_URL.rstrip('/')}"
            f"/api/v1/tiktok/products/{tt_id}/sticker_update"
        )

        attempt_idx = 1
        with _open_db(self.db_path) as conn:
            attempt_idx = (conn.execute(
                "SELECT COALESCE(MAX(attempt_idx), 0) + 1 FROM tkshop_publish_logs WHERE product_id = ?",
                (product_id,),
            ).fetchone()[0]) or 1

        resp_json: dict[str, Any] = {}
        success = False
        error_code = ""
        error_message = ""
        try:
            resp = requests.post(endpoint, json=payload,
                                 params={"shop": shop_name},
                                 timeout=TKSHOP_SERVER_TIMEOUT)
            try:
                resp_json = resp.json()
            except Exception:
                resp_json = {"_raw_text": resp.text[:500]}
            if isinstance(resp_json, dict) and "data" in resp_json and isinstance(resp_json["data"], dict) and "ok" in resp_json["data"]:
                resp_json = resp_json["data"]
            if 200 <= resp.status_code < 300 and resp_json.get("ok") is True:
                success = True
            else:
                error_code = str(resp_json.get("error_code") or resp.status_code)
                error_message = str(resp_json.get("error_message") or resp.text[:200])
        except requests.RequestException as e:
            error_code = "NETWORK"; error_message = str(e)[:300]
        except Exception as e:
            error_code = "INTERNAL"; error_message = f"{type(e).__name__}: {e}"[:300]

        self._log_publish_attempt(
            product_id, attempt_idx, endpoint, payload,
            response=resp_json, success=success,
            error_code=error_code, error_message=error_message,
        )
        # Re-deactivate after a successful update if the product was
        # 'seller_deactivated' before — TikTok's PUT bumps it back to LIVE.
        if success and prev_status == "seller_deactivated":
            try:
                self._post_platform_action(product_id, "deactivate")
                with _open_db(self.db_path) as conn:
                    conn.execute(
                        "UPDATE tkshop_products SET publish_status = 'seller_deactivated' WHERE id = ?",
                        (product_id,),
                    )
                    conn.commit()
            except Exception as e:  # noqa: BLE001
                logger.warning(
                    "post-update re-deactivate failed for product #%d: %s — "
                    "product is LIVE on TikTok now; operator can manually deactivate.",
                    product_id, e,
                )

        return {
            "ok": success,
            "tiktok_product_id": tt_id,
            "error_code": error_code,
            "error_message": error_message,
            "attempt_idx": attempt_idx,
        }

    def _post_platform_action(self, product_id: int, action: str) -> dict[str, Any]:
        """Call POST /api/v1/tiktok/products/{tt_id}/{action} (activate /
        deactivate / recover). Returns ``{ok, error_code?, error_message?}``.
        """
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                "SELECT tiktok_product_id FROM tkshop_products WHERE id = ?",
                (product_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"product #{product_id} not found")
            tt_id = (row["tiktok_product_id"] or "").strip()
        if not tt_id:
            return {"ok": False, "error_code": "NOT_PUBLISHED",
                    "error_message": "product has no tiktok_product_id"}
        shop_name = _get_product_shop(self.db_path, product_id)
        endpoint = (
            f"{TKSHOP_SERVER_URL.rstrip('/')}"
            f"/api/v1/tiktok/products/{tt_id}/{action}"
        )
        try:
            resp = requests.post(endpoint, params={"shop": shop_name},
                                 timeout=TKSHOP_SERVER_TIMEOUT)
            data = resp.json() if resp.text else {}
        except requests.RequestException as e:
            return {"ok": False, "error_code": "NETWORK",
                    "error_message": f"{type(e).__name__}: {e}"}
        env_ok = bool(data.get("success", True)) if isinstance(data, dict) else True
        msg = (data.get("message") or "") if isinstance(data, dict) else ""
        det = (data.get("detail")  or "") if isinstance(data, dict) else ""
        if not env_ok or not resp.ok:
            raw = (msg or "") + (f" | {det}" if det else "")
            friendly = _friendly_platform_error(action, raw)
            return {"ok": False,
                    "error_code": str(resp.status_code),
                    "error_message": friendly,
                    "raw_error": raw}
        return {"ok": True, "data": data.get("data") or {}}

    def activate_on_platform(self, product_id: int) -> dict[str, Any]:
        result = self._post_platform_action(product_id, "activate")
        if result.get("ok"):
            with _open_db(self.db_path) as conn:
                conn.execute(
                    "UPDATE tkshop_products SET publish_status = 'published' WHERE id = ?",
                    (product_id,),
                )
                conn.commit()
        return result

    def deactivate_on_platform(self, product_id: int) -> dict[str, Any]:
        result = self._post_platform_action(product_id, "deactivate")
        if result.get("ok"):
            with _open_db(self.db_path) as conn:
                conn.execute(
                    "UPDATE tkshop_products SET publish_status = 'seller_deactivated' WHERE id = ?",
                    (product_id,),
                )
                conn.commit()
        return result

    def reset_for_republish(self, product_id: int) -> None:
        """Clear platform-side IDs so the next publish() creates a fresh
        listing instead of being short-circuited by the
        publish_with_self_heal duplicate-guard. Used when the operator wants
        to re-publish a product whose platform listing was deleted (or for
        any reason should be re-created).
        """
        with _open_db(self.db_path) as conn:
            conn.execute(
                """
                UPDATE tkshop_products
                   SET tiktok_product_id = '',
                       tiktok_sku_id     = '',
                       publish_status    = 'draft',
                       published_at      = NULL
                 WHERE id = ?
                """,
                (product_id,),
            )
            conn.commit()

    def clone_to_shop(self, product_id: int, target_shop: str) -> int:
        """Duplicate a product row + image rows into a new tkshop_products
        entry bound to ``target_shop``. Resets platform IDs and status so the
        clone shows up as a fresh local draft ready to publish to the new
        shop. Pack association and on-disk artifacts (images) are shared with
        the source — duplicating those would balloon storage for marginal
        benefit.

        Returns the new product id.
        """
        if not target_shop:
            raise ValueError("target_shop is required")
        # Normalize to the canonical shop key up front so a legacy alias and
        # its canonical name are never treated as two different stores.
        target_shop = canonical_shop_key(target_shop)
        if target_shop not in TKSHOP_SHOPS:
            raise ValueError(
                f"unknown shop {target_shop!r}; configured shops: {TKSHOP_SHOPS}"
            )
        now = int(time.time())
        with _open_db(self.db_path) as conn:
            src = conn.execute(
                "SELECT pr.*, p.pack_uid, p.total_stickers "
                "FROM tkshop_products pr "
                "LEFT JOIN packs p ON p.id = pr.pack_id "
                "WHERE pr.id = ?",
                (product_id,),
            ).fetchone()
            if not src:
                raise ValueError(f"product #{product_id} not found")
            if canonical_shop_key(src["shop"] or TKSHOP_DEFAULT_SHOP) == target_shop:
                raise ValueError(
                    f"product #{product_id} is already on shop {target_shop!r}"
                )
            # Don't fork a second listing if this pack already has one on the
            # target shop (under any alias) — return that existing listing.
            for row in conn.execute(
                "SELECT id, shop FROM tkshop_products WHERE pack_id = ? AND id != ?",
                (src["pack_id"], product_id),
            ).fetchall():
                if canonical_shop_key(row["shop"]) == target_shop:
                    return row["id"]
            # Unified SKU: every shop's listing for a pack shares one SKU.
            # Prefer the master's SKU so a manually-edited master propagates;
            # fall back to the computed default. Uniqueness is scoped to the
            # target shop; a pre-existing row squatting that SKU gets a -2.
            master_sku = ""
            if src["local_product_id"]:
                m = conn.execute(
                    "SELECT seller_sku FROM local_products WHERE id = ?",
                    (src["local_product_id"],),
                ).fetchone()
                master_sku = (m["seller_sku"] or "").strip() if m else ""
            target_sku = master_sku or compute_default_seller_sku(
                pack_uid=src["pack_uid"] or "",
                total_stickers=src["total_stickers"] or 0,
                pack_id=src["pack_id"],
            )
            target_sku = _ensure_unique_seller_sku(
                conn, target_sku, shop=target_shop,
            )
            # Carry the source's local_product_id over to the clone so the new
            # listing still points at the same master — that's what makes a
            # multi-shop pack reuse one catalog entry instead of forking copies.
            cur = conn.execute(
                """
                INSERT INTO tkshop_products
                    (pack_id, shop, tiktok_product_id, tiktok_sku_id,
                     detail_main_raw_text, title, description_html,
                     selling_points, keywords, seller_sku, category_id,
                     default_template_json, publish_status,
                     created_at, published_at, auto_fix_attempts, last_fix_diff,
                     local_product_id)
                VALUES (?, ?, '', '', ?, ?, ?, ?, ?, ?, ?, ?, 'draft', ?, NULL, 0, '', ?)
                """,
                (
                    src["pack_id"], target_shop,
                    src["detail_main_raw_text"] or "",
                    src["title"] or "", src["description_html"] or "",
                    src["selling_points"] or "[]", src["keywords"] or "[]",
                    target_sku, src["category_id"] or "928016",
                    src["default_template_json"] or "{}", now,
                    src["local_product_id"],
                ),
            )
            new_id = cur.lastrowid
            # Replicate image rows so the new product has the same images.
            # local_path stays the same — files are shared on disk.
            # tiktok_image_uri is reset because the new product is a fresh
            # listing on the target shop and will get its own URIs on publish.
            conn.execute(
                """
                INSERT INTO tkshop_product_images
                    (product_id, role, source, local_path,
                     tiktok_image_uri, sort_order, ai_prompt, created_at)
                SELECT ?, role, source, local_path,
                       '', sort_order, ai_prompt, ?
                  FROM tkshop_product_images
                 WHERE product_id = ?
                """,
                (new_id, now, product_id),
            )
            conn.commit()
        logger.info(
            "cloned product #%d → #%d (shop=%s → %s)",
            product_id, new_id,
            (src["shop"] or TKSHOP_DEFAULT_SHOP), target_shop,
        )
        return int(new_id)

    def batch_publish_local_to_shop(
        self,
        local_product_ids: list[int],
        target_shop: str,
        *,
        auto_activate: bool = True,
        auto_fix: bool = True,
        max_workers: int = 2,
    ) -> dict[str, Any]:
        """Master-first variant of ``batch_publish_to_shop``.

        Input is local_product ids (rows from the 本地产品 tab). For each:
          - resolve the pack
          - find or create a tkshop_products listing for (pack, target_shop)
            via ``create_product_from_pack`` (which seeds the listing from
            master + carries the local_product_id FK)
          - skip if the listing is already on TikTok
          - else publish the listing

        Same summary shape as ``batch_publish_to_shop`` so the UI consumes
        either uniformly.
        """
        from concurrent.futures import ThreadPoolExecutor, as_completed
        if not target_shop or target_shop not in TKSHOP_SHOPS:
            raise ValueError(
                f"unknown shop {target_shop!r}; configured shops: {TKSHOP_SHOPS}"
            )
        t0 = time.time()
        seen: set[int] = set()
        clean_ids: list[int] = []
        for lp in local_product_ids:
            try:
                v = int(lp)
            except (TypeError, ValueError):
                continue
            if v in seen:
                continue
            seen.add(v)
            clean_ids.append(v)
        published, skipped, failed = [], [], []
        # Resolve master → listing in target shop, building a "what to publish"
        # plan in a single DB pass to avoid worker-pool contention.
        plan: list[tuple[int, int]] = []  # (listing_id, local_product_id)
        with _open_db(self.db_path) as conn:
            for lp_id in clean_ids:
                row = conn.execute(
                    "SELECT id, pack_id, title, category_id FROM local_products WHERE id = ?",
                    (lp_id,),
                ).fetchone()
                if not row:
                    skipped.append({"local_product_id": lp_id, "reason": "master not found"})
                    continue
                if not (row["title"] or "").strip():
                    skipped.append({"local_product_id": lp_id, "reason": "incomplete: missing title"})
                    continue
                if not (row["category_id"] or "").strip():
                    skipped.append({"local_product_id": lp_id, "reason": "incomplete: missing category"})
                    continue
                # Need at least one master main image. If the master gallery
                # is missing one but the target-shop listing already has a
                # main row (e.g. operator promoted a listing image), sync it
                # back so publish can proceed.
                has_img = conn.execute(
                    "SELECT 1 FROM local_product_images "
                    "WHERE local_product_id = ? AND role = 'main' LIMIT 1",
                    (lp_id,),
                ).fetchone()
                listing_for_sync = conn.execute(
                    "SELECT id FROM tkshop_products "
                    "WHERE local_product_id = ? AND shop = ? "
                    "ORDER BY id DESC LIMIT 1",
                    (lp_id, target_shop),
                ).fetchone()
                if not has_img and listing_for_sync:
                    if self._sync_listing_main_to_local(
                        lp_id, listing_for_sync["id"],
                    ):
                        has_img = True
                if not has_img:
                    skipped.append({"local_product_id": lp_id, "reason": "incomplete: no main image"})
                    continue
                # If a listing on target shop already has a TikTok id, this is
                # a redundant request — surface the existing id and skip.
                already = conn.execute(
                    "SELECT id FROM tkshop_products "
                    "WHERE local_product_id = ? AND shop = ? "
                    "AND tiktok_product_id != ''",
                    (lp_id, target_shop),
                ).fetchone()
                if already:
                    skipped.append({
                        "local_product_id": lp_id,
                        "reason": "already listed on shop",
                        "existing_product_id": already["id"],
                    })
                    continue
                # Find or create the listing. Reuse a draft listing on the
                # target shop if one exists; otherwise let create_product_from_pack
                # mint a fresh one (it's idempotent for (pack, shop)).
                listing = conn.execute(
                    "SELECT id FROM tkshop_products "
                    "WHERE local_product_id = ? AND shop = ? "
                    "ORDER BY id DESC LIMIT 1",
                    (lp_id, target_shop),
                ).fetchone()
                if listing:
                    plan.append((listing["id"], lp_id))
                else:
                    # Defer creation to the worker so DB writes don't pile up
                    # before publish kicks off — but we need pack_id here.
                    plan.append((-1 * row["pack_id"], lp_id))  # negative = needs-create
        publish_fn = self.publish_with_self_heal if auto_fix else self.publish

        def _one(item: tuple[int, int]) -> dict:
            raw, lp_id = item
            try:
                if raw < 0:
                    listing_id = self.create_product_from_pack(-raw, shop=target_shop)
                else:
                    listing_id = raw
                # Local product is the source of truth: refresh the listing's
                # text fields + SKU and images from the master right before
                # publishing, so a reused draft never ships stale content.
                self._sync_master_to_listing(lp_id, listing_id)
                self.sync_master_images_to_listing(lp_id, listing_id, replace=True)
                with _open_db(self.db_path) as conn:
                    conn.execute(
                        "UPDATE tkshop_products SET publish_status = 'publishing' "
                        "WHERE id = ?",
                        (listing_id,),
                    )
                    conn.commit()
                kwargs: dict[str, Any] = {"auto_activate": auto_activate}
                if not auto_fix:
                    kwargs["dry_run"] = False
                publish_fn(listing_id, **kwargs)
                try:
                    self.sync_one_status(listing_id)
                except Exception:
                    pass
                return {"ok": True, "local_product_id": lp_id, "product_id": listing_id}
            except Exception as e:
                return {"ok": False, "local_product_id": lp_id,
                        "error": f"{e}"[:300]}

        if plan:
            workers = max(1, min(max_workers, len(plan)))
            with ThreadPoolExecutor(max_workers=workers,
                                     thread_name_prefix="tkshop-publish-master") as ex:
                futures = [ex.submit(_one, item) for item in plan]
                for fut in as_completed(futures):
                    res = fut.result()
                    if res.get("ok"):
                        published.append({k: v for k, v in res.items() if k != "ok"})
                    else:
                        failed.append({"local_product_id": res["local_product_id"],
                                       "error": res.get("error", "")})
        elapsed = round(time.time() - t0, 1)
        logger.info(
            "batch_publish_local_to_shop: target=%s requested=%d published=%d skipped=%d failed=%d in %.1fs",
            target_shop, len(clean_ids), len(published), len(skipped), len(failed), elapsed,
        )
        return {"published": published, "skipped": skipped, "failed": failed,
                "elapsed_s": elapsed, "target_shop": target_shop}

    def mirror_listing_ai_to_local_master(
        self,
        listing_id: int,
        *,
        roles: Optional[set[str]] = None,
        append_only: bool = False,
        image_ids: Optional[list[int]] = None,
    ) -> int:
        """Copy listing AI images into ``local_product_images``.

        When ``roles`` is omitted, sync every AI row on the listing (full
        repair). Otherwise only replace the given roles (``main`` /
        ``secondary``) before inserting the listing's AI rows.

        When ``append_only`` is True, existing master rows are kept and only
        the selected listing image rows are inserted (used for +1 append).

        Returns the number of rows inserted into the master gallery.
        """
        with _open_db(self.db_path) as conn:
            lp_id_row = conn.execute(
                "SELECT local_product_id FROM tkshop_products WHERE id = ?",
                (listing_id,),
            ).fetchone()
            if not lp_id_row or lp_id_row["local_product_id"] is None:
                return 0
            lp_id = int(lp_id_row["local_product_id"])
            if append_only:
                query = """
                    SELECT role, source, local_path, sort_order, ai_prompt, created_at
                      FROM tkshop_product_images
                     WHERE product_id = ? AND source = 'ai'
                """
                params: list[Any] = [listing_id]
                if image_ids:
                    placeholders = ",".join("?" * len(image_ids))
                    query += f" AND id IN ({placeholders})"
                    params.extend(int(i) for i in image_ids)
                rows = conn.execute(query, tuple(params)).fetchall()
                inserted = 0
                for r in rows:
                    conn.execute(
                        """
                        INSERT INTO local_product_images
                            (local_product_id, role, source, local_path,
                             sort_order, ai_prompt, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            lp_id, r["role"], r["source"], r["local_path"],
                            r["sort_order"], r["ai_prompt"], r["created_at"],
                        ),
                    )
                    inserted += 1
                if inserted and any(r["role"] == "main" for r in rows):
                    self._sync_pack_cover_from_listing_main(listing_id, conn=conn)
                if inserted:
                    conn.execute(
                        "UPDATE local_products SET updated_at = ? WHERE id = ?",
                        (int(time.time()), lp_id),
                    )
                conn.commit()
                return inserted
            sync_roles = roles
            if sync_roles is None:
                sync_roles = {
                    r["role"] for r in conn.execute(
                        """
                        SELECT DISTINCT role FROM tkshop_product_images
                         WHERE product_id = ? AND source = 'ai'
                        """,
                        (listing_id,),
                    )
                }
            if "main" in sync_roles:
                conn.execute(
                    "DELETE FROM local_product_images "
                    "WHERE local_product_id = ? AND role = 'main'",
                    (lp_id,),
                )
            if "secondary" in sync_roles:
                conn.execute(
                    "DELETE FROM local_product_images "
                    "WHERE local_product_id = ? AND role = 'secondary'",
                    (lp_id,),
                )
            conn.execute(
                "DELETE FROM local_product_images "
                "WHERE local_product_id = ? AND source = 'ai'",
                (lp_id,),
            )
            cur = conn.execute(
                """
                INSERT INTO local_product_images
                    (local_product_id, role, source, local_path,
                     sort_order, ai_prompt, created_at)
                SELECT ?, role, source, local_path, sort_order, ai_prompt, created_at
                  FROM tkshop_product_images
                 WHERE product_id = ? AND source = 'ai'
                """,
                (lp_id, listing_id),
            )
            inserted = cur.rowcount
            conn.execute(
                "UPDATE local_products SET updated_at = ? WHERE id = ?",
                (int(time.time()), lp_id),
            )
            if "main" in sync_roles:
                self._sync_pack_cover_from_listing_main(listing_id, conn=conn)
            conn.commit()
            return inserted

    def _sync_master_to_listing(
        self, local_product_id: int, listing_id: int,
    ) -> bool:
        """Overwrite a listing's text fields from its master (local product).

        The local product is the single source of truth, so right before a
        push we copy the master's title/description/selling_points/keywords/
        detail/category and SKU onto the listing. Platform IDs and
        publish_status are left untouched. Returns False if either row is
        missing. Only call this for unpublished listings — the master-first
        publish path already skips listings that are live on TikTok.
        """
        with _open_db(self.db_path) as conn:
            m = conn.execute(
                "SELECT title, description_html, selling_points, keywords, "
                "       detail_main_raw_text, category_id, "
                "       default_template_json, seller_sku "
                "FROM local_products WHERE id = ?",
                (local_product_id,),
            ).fetchone()
            if not m:
                return False
            cur = conn.execute(
                """
                UPDATE tkshop_products
                   SET title                = ?,
                       description_html      = ?,
                       selling_points        = ?,
                       keywords              = ?,
                       detail_main_raw_text  = ?,
                       category_id           = ?,
                       default_template_json = ?,
                       seller_sku            = ?
                 WHERE id = ?
                """,
                (
                    m["title"] or "",
                    m["description_html"] or "",
                    m["selling_points"] or "[]",
                    m["keywords"] or "[]",
                    m["detail_main_raw_text"] or "",
                    m["category_id"] or "928016",
                    m["default_template_json"] or "{}",
                    m["seller_sku"] or "",
                    listing_id,
                ),
            )
            conn.commit()
            return cur.rowcount > 0

    def _sync_listing_main_to_local(
        self, local_product_id: int, listing_id: int,
    ) -> bool:
        """Copy the listing's main image into ``local_product_images`` when
        the master gallery has no main row yet."""
        with _open_db(self.db_path) as conn:
            has_local_main = conn.execute(
                "SELECT 1 FROM local_product_images "
                "WHERE local_product_id = ? AND role = 'main' LIMIT 1",
                (local_product_id,),
            ).fetchone()
            if has_local_main:
                return True
            listing_main = conn.execute(
                """
                SELECT role, source, local_path, sort_order, ai_prompt, created_at
                  FROM tkshop_product_images
                 WHERE product_id = ? AND role = 'main'
                 ORDER BY sort_order ASC, id ASC
                 LIMIT 1
                """,
                (listing_id,),
            ).fetchone()
            if not listing_main or not (listing_main["local_path"] or "").strip():
                return False
            conn.execute(
                "DELETE FROM local_product_images "
                "WHERE local_product_id = ? AND role = 'main'",
                (local_product_id,),
            )
            conn.execute(
                """
                INSERT INTO local_product_images
                    (local_product_id, role, source, local_path,
                     sort_order, ai_prompt, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    local_product_id,
                    listing_main["role"],
                    listing_main["source"],
                    listing_main["local_path"],
                    listing_main["sort_order"],
                    listing_main["ai_prompt"],
                    listing_main["created_at"],
                ),
            )
            conn.commit()
            return True

    def _snapshot_master_images_to_listing(
        self, local_product_id: int, listing_id: int,
    ) -> int:
        """Replicate master image rows into the listing's
        ``tkshop_product_images``. Only fills if the listing has no images
        yet (typical for a freshly-minted listing); existing listing images
        are left alone so a published listing's tiktok_image_uri isn't lost.
        Returns the number of rows inserted.
        """
        with _open_db(self.db_path) as conn:
            has_imgs = conn.execute(
                "SELECT 1 FROM tkshop_product_images WHERE product_id = ? LIMIT 1",
                (listing_id,),
            ).fetchone()
            if has_imgs:
                return 0
            cur = conn.execute(
                """
                INSERT INTO tkshop_product_images
                    (product_id, role, source, local_path,
                     tiktok_image_uri, sort_order, ai_prompt, created_at)
                SELECT ?, role, source, local_path,
                       '', sort_order, ai_prompt, ?
                  FROM local_product_images
                 WHERE local_product_id = ?
                """,
                (listing_id, int(time.time()), local_product_id),
            )
            conn.commit()
            return cur.rowcount

    def pad_master_gallery_to_target(
        self, local_product_id: int, *, target_total: int = 5,
    ) -> int:
        """Pad the master gallery to ``target_total`` by duplicating existing
        secondary images (no AI). Returns rows inserted."""
        target_total = max(2, int(target_total))
        with _open_db(self.db_path) as conn:
            rows = conn.execute(
                """
                SELECT id, role, local_path, sort_order, ai_prompt
                  FROM local_product_images
                 WHERE local_product_id = ?
                 ORDER BY role = 'main' DESC, sort_order ASC, id ASC
                """,
                (local_product_id,),
            ).fetchall()
            if len(rows) >= target_total:
                return 0
            secondaries = [
                r for r in rows
                if r["role"] == "secondary" and (r["local_path"] or "").strip()
            ]
            if not secondaries:
                raise ValueError(
                    f"local_product #{local_product_id} has no secondary image to duplicate"
                )
            now = int(time.time())
            max_sort = max(int(r["sort_order"] or 0) for r in rows)
            added = 0
            idx = 0
            while len(rows) + added < target_total:
                src = secondaries[idx % len(secondaries)]
                max_sort += 1
                conn.execute(
                    """
                    INSERT INTO local_product_images
                        (local_product_id, role, source, local_path,
                         sort_order, ai_prompt, created_at)
                    VALUES (?, 'secondary', 'duplicate', ?, ?, ?, ?)
                    """,
                    (
                        local_product_id,
                        src["local_path"],
                        max_sort,
                        src["ai_prompt"] or "",
                        now,
                    ),
                )
                added += 1
                idx += 1
            if added:
                conn.execute(
                    "UPDATE local_products SET updated_at = ? WHERE id = ?",
                    (now, local_product_id),
                )
                conn.commit()
            return added

    def remove_duplicate_gallery_rows(
        self, local_product_ids: Optional[list[int]] = None,
    ) -> dict[str, int]:
        """Drop ``source='duplicate'`` rows from master + listing galleries.

        Does not unlink files on disk — duplicate rows point at the same
        ``local_path`` as the originals they were copied from.
        """
        lp_ids = [int(x) for x in (local_product_ids or []) if int(x) > 0]
        with _open_db(self.db_path) as conn:
            if lp_ids:
                ph = ",".join("?" * len(lp_ids))
                master_rows = conn.execute(
                    f"""
                    SELECT id FROM local_product_images
                     WHERE source = 'duplicate' AND local_product_id IN ({ph})
                    """,
                    tuple(lp_ids),
                ).fetchall()
                listing_rows = conn.execute(
                    f"""
                    SELECT ti.id FROM tkshop_product_images ti
                      JOIN tkshop_products t ON t.id = ti.product_id
                     WHERE ti.source = 'duplicate' AND t.local_product_id IN ({ph})
                    """,
                    tuple(lp_ids),
                ).fetchall()
            else:
                master_rows = conn.execute(
                    "SELECT id FROM local_product_images WHERE source = 'duplicate'",
                ).fetchall()
                listing_rows = conn.execute(
                    "SELECT id FROM tkshop_product_images WHERE source = 'duplicate'",
                ).fetchall()
            for row in master_rows:
                conn.execute(
                    "DELETE FROM local_product_images WHERE id = ?", (row["id"],),
                )
            for row in listing_rows:
                conn.execute(
                    "DELETE FROM tkshop_product_images WHERE id = ?", (row["id"],),
                )
            if lp_ids:
                now = int(time.time())
                for lp_id in lp_ids:
                    conn.execute(
                        "UPDATE local_products SET updated_at = ? WHERE id = ?",
                        (now, lp_id),
                    )
            conn.commit()
        return {
            "deleted_master": len(master_rows),
            "deleted_listing": len(listing_rows),
        }

    @staticmethod
    def _is_packaging_secondary_prompt(prompt: str) -> bool:
        p = (prompt or "").lower()
        return any(k in p for k in (
            "envelope", "glassine", "kraft sleeve", "kraft paper sl",
            "pack partly", "pack half", "partially tucked", "partly tucked",
            "packaging photo", "glassine sleeve", "kraft pouch",
        ))

    def _merged_reference_bytes_for_product(
        self, product_id: int, ctx: Optional[dict[str, Any]] = None,
    ) -> tuple[bytes, str]:
        from src.utils.image_utils import compose_reference_grid, compress_image_bytes_for_api

        ctx = ctx or self.collect_pack_design_context(product_id)
        asset_root = self.db_path.resolve().parent.parent

        def _paths(key: str) -> list[Path]:
            out: list[Path] = []
            seen: set[str] = set()
            for raw in ctx.get(key, []):
                p = _resolve_disk_path(raw, root=asset_root)
                if not p.is_file():
                    continue
                k = p.as_posix()
                if k in seen:
                    continue
                seen.add(k)
                out.append(p)
            return out

        sheet_paths = _paths("preview_image_paths")
        sticker_paths = _paths("sticker_image_paths")
        preview_threshold = max(1, int(os.getenv("TKSHOP_MAIN_PREVIEW_MIN_COUNT", "3")))
        if len(sheet_paths) >= preview_threshold:
            bundle_paths, label = sheet_paths, "preview_sheets_merged_grid"
        elif sticker_paths:
            bundle_paths, label = sticker_paths, "split_stickers_merged_grid"
        else:
            bundle_paths, label = sheet_paths, "preview_sheets_merged_grid"
        if not bundle_paths:
            raise ValueError(f"product #{product_id}: no preview/sticker reference assets")
        raw = compose_reference_grid(
            [p.as_posix() for p in bundle_paths],
            cell=384 if len(bundle_paths) >= 12 else 512,
            max_side=int(os.getenv("TKSHOP_MAIN_MERGE_MAX_SIDE", "1536")),
        )
        compressed = compress_image_bytes_for_api(
            raw,
            max_side=int(os.getenv("TKSHOP_MAIN_REF_UPLOAD_MAX_SIDE", "1536")),
            max_bytes=int(os.getenv("TKSHOP_MAIN_REF_MAX_BYTES", "1800000")),
        )
        return compressed, label

    def delete_gallery_fifth_image(self, local_product_id: int) -> dict[str, Any]:
        """Remove the 5th gallery image (highest sort_order secondary)."""
        with _open_db(self.db_path) as conn:
            rows = conn.execute(
                """
                SELECT id, sort_order, local_path
                  FROM local_product_images
                 WHERE local_product_id = ? AND role = 'secondary'
                 ORDER BY sort_order DESC, id DESC
                """,
                (local_product_id,),
            ).fetchall()
        if not rows:
            return {"deleted": False, "reason": "no secondary rows"}
        target = rows[0]
        self.delete_local_product_image(int(target["id"]))
        return {
            "deleted": True,
            "image_id": int(target["id"]),
            "sort_order": int(target["sort_order"] or 0),
            "local_path": target["local_path"] or "",
        }

    def delete_latest_packaging_secondary(self, local_product_id: int) -> dict[str, Any]:
        """Remove the newest packaging/envelope secondary from the master gallery."""
        with _open_db(self.db_path) as conn:
            rows = conn.execute(
                """
                SELECT id, ai_prompt, sort_order
                  FROM local_product_images
                 WHERE local_product_id = ? AND role = 'secondary'
                 ORDER BY sort_order DESC, id DESC
                """,
                (local_product_id,),
            ).fetchall()
        if not rows:
            return {"deleted": False, "reason": "no secondary rows"}
        packaging = [
            r for r in rows
            if self._is_packaging_secondary_prompt(r["ai_prompt"] or "")
        ]
        target = packaging[0] if packaging else rows[0]
        self.delete_local_product_image(int(target["id"]))
        return {
            "deleted": True,
            "image_id": int(target["id"]),
            "sort_order": int(target["sort_order"] or 0),
            "was_packaging": bool(packaging),
        }

    def generate_hand_hold_secondary_for_local(
        self, local_product_id: int, *, sort_order: Optional[int] = None,
    ) -> dict[str, Any]:
        """Generate one hand-hold + pile-background secondary on the master gallery."""
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                """
                SELECT p.pack_uid, t.id AS listing_id
                  FROM local_products lp
                  JOIN packs p ON p.id = lp.pack_id
                  JOIN tkshop_products t ON t.local_product_id = lp.id
                 WHERE lp.id = ?
                 ORDER BY t.id
                 LIMIT 1
                """,
                (local_product_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"local_product #{local_product_id} not found")
            listing_id = int(row["listing_id"])
            pack_uid = row["pack_uid"]

        ref_bytes, ref_label = self._merged_reference_bytes_for_product(listing_id)
        out_bytes = self.router.image_edit(
            ref_bytes,
            hand_hold_secondary_spec()["image_prompt"],
            size="1024x1024",
            quality=os.getenv("TKSHOP_SECONDARY_IMAGE_QUALITY", "medium"),
            task="tkshop_image_design:hand_hold_secondary",
            related_table="local_products",
            related_id=local_product_id,
        )
        filename = f"secondary_hand_hold_{int(time.time() * 1000)}.png"
        dest = self.store.write_product_image(
            pack_uid, f"draft_{listing_id}", filename, out_bytes,
        )
        hh_spec = hand_hold_secondary_spec()
        img_id = self.add_local_product_image(
            local_product_id,
            local_path=dest.as_posix(),
            role="secondary",
            source="ai",
            ai_prompt=hh_spec["image_prompt"],
            sort_order=sort_order,
        )
        return {
            "ok": True,
            "local_product_id": local_product_id,
            "image_id": img_id,
            "local_path": dest.as_posix(),
            "reference": ref_label,
        }

    def batch_replace_hand_hold_and_push(
        self,
        local_product_ids: list[int],
        *,
        workers: int = 3,
        push: bool = True,
        dry_run: bool = False,
    ) -> dict[str, Any]:
        """Replace packaging/envelope 5th secondaries with hand-hold pile shots.

        Only touches the listed ``local_product_ids``. Master stays at 5 images
        (delete one packaging secondary, add one hand-hold). Then syncs every
        published listing for each master and optionally pushes to TikTok.
        """
        from concurrent.futures import ThreadPoolExecutor, as_completed

        ids = [int(x) for x in local_product_ids if int(x) > 0]
        workers = max(1, min(int(workers), len(ids) or 1))
        t0 = time.time()
        replaced, synced, pushed, failed, skipped = [], [], [], [], []

        if dry_run:
            preview = []
            for lp_id in ids:
                with _open_db(self.db_path) as conn:
                    total = conn.execute(
                        "SELECT COUNT(*) FROM local_product_images WHERE local_product_id = ?",
                        (lp_id,),
                    ).fetchone()[0]
                    listings = [
                        dict(r) for r in conn.execute(
                            """
                            SELECT id, shop, tiktok_product_id
                              FROM tkshop_products
                             WHERE local_product_id = ?
                               AND COALESCE(tiktok_product_id, '') != ''
                             ORDER BY shop, id
                            """,
                            (lp_id,),
                        ).fetchall()
                    ]
                preview.append({
                    "local_product_id": lp_id,
                    "master_total": total,
                    "listings": listings,
                })
            return {
                "dry_run": True,
                "workers": workers,
                "local_product_ids": ids,
                "preview": preview,
            }

        def _one(lp_id: int) -> dict[str, Any]:
            entry: dict[str, Any] = {"local_product_id": lp_id}
            try:
                deleted = self.delete_latest_packaging_secondary(lp_id)
                entry["deleted"] = deleted
                if not deleted.get("deleted"):
                    return {**entry, "ok": False, "error": deleted.get("reason", "delete failed")}
                gen = self.generate_hand_hold_secondary_for_local(
                    lp_id, sort_order=deleted.get("sort_order"),
                )
                entry.update(gen)
                return entry
            except Exception as e:
                return {**entry, "ok": False, "error": str(e)[:300]}

        with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="hand-hold") as ex:
            futures = {ex.submit(_one, lp_id): lp_id for lp_id in ids}
            for fut in as_completed(futures):
                lp_id = futures[fut]
                try:
                    result = fut.result()
                except Exception as e:
                    result = {"local_product_id": lp_id, "ok": False, "error": str(e)[:300]}
                if result.get("ok"):
                    replaced.append(result)
                else:
                    failed.append(result)

        for item in replaced:
            lp_id = int(item["local_product_id"])
            with _open_db(self.db_path) as conn:
                listings = [
                    dict(r) for r in conn.execute(
                        """
                        SELECT id, shop, tiktok_product_id
                          FROM tkshop_products
                         WHERE local_product_id = ?
                           AND COALESCE(tiktok_product_id, '') != ''
                         ORDER BY shop, id
                        """,
                        (lp_id,),
                    ).fetchall()
                ]
            for listing in listings:
                lid = int(listing["id"])
                try:
                    n = self.sync_master_images_to_listing(lp_id, lid, replace=True)
                    synced.append({
                        "local_product_id": lp_id,
                        "listing_id": lid,
                        "shop": listing["shop"],
                        "images": n,
                    })
                    if push:
                        res = self.update_on_platform(lid)
                        rec = {
                            "local_product_id": lp_id,
                            "listing_id": lid,
                            "shop": listing["shop"],
                            "ok": bool(res.get("ok")),
                            "error": res.get("error_message") or "",
                        }
                        (pushed if rec["ok"] else failed).append(rec)
                except Exception as exc:
                    failed.append({
                        "local_product_id": lp_id,
                        "listing_id": lid,
                        "shop": listing.get("shop", ""),
                        "ok": False,
                        "error": str(exc)[:300],
                    })

        return {
            "dry_run": False,
            "workers": workers,
            "local_product_ids": ids,
            "replaced": replaced,
            "synced": synced,
            "pushed": pushed,
            "skipped": skipped,
            "failed": failed,
            "elapsed_s": round(time.time() - t0, 1),
        }

    def batch_hand_hold_gallery_mixed(
        self,
        replace_fifth_ids: list[int],
        append_fifth_ids: list[int],
        *,
        workers: int = 3,
        push: bool = True,
        dry_run: bool = False,
    ) -> dict[str, Any]:
        """Hand-hold 5th image: replace group deletes 5th first; append group adds when short."""
        from concurrent.futures import ThreadPoolExecutor, as_completed

        replace_ids = [int(x) for x in replace_fifth_ids if int(x) > 0]
        append_ids = [int(x) for x in append_fifth_ids if int(x) > 0]
        all_ids = replace_ids + append_ids
        workers = max(1, min(int(workers), len(all_ids) or 1))
        replace_set = set(replace_ids)
        t0 = time.time()
        generated, synced, pushed, failed = [], [], [], []

        if dry_run:
            preview = []
            for lp_id in all_ids:
                with _open_db(self.db_path) as conn:
                    total = conn.execute(
                        "SELECT COUNT(*) FROM local_product_images WHERE local_product_id = ?",
                        (lp_id,),
                    ).fetchone()[0]
                    listings = [
                        dict(r) for r in conn.execute(
                            """
                            SELECT id, shop, tiktok_product_id
                              FROM tkshop_products
                             WHERE local_product_id = ?
                               AND COALESCE(tiktok_product_id, '') != ''
                             ORDER BY shop, id
                            """,
                            (lp_id,),
                        ).fetchall()
                    ]
                preview.append({
                    "local_product_id": lp_id,
                    "mode": "replace_fifth" if lp_id in replace_set else "append_fifth",
                    "master_total": total,
                    "listings": listings,
                })
            return {
                "dry_run": True,
                "workers": workers,
                "replace_fifth_ids": replace_ids,
                "append_fifth_ids": append_ids,
                "preview": preview,
            }

        def _one(lp_id: int) -> dict[str, Any]:
            entry: dict[str, Any] = {
                "local_product_id": lp_id,
                "mode": "replace_fifth" if lp_id in replace_set else "append_fifth",
            }
            try:
                sort_order: Optional[int] = None
                if lp_id in replace_set:
                    deleted = self.delete_gallery_fifth_image(lp_id)
                    entry["deleted"] = deleted
                    if not deleted.get("deleted"):
                        return {**entry, "ok": False, "error": deleted.get("reason", "delete failed")}
                    sort_order = deleted.get("sort_order")
                else:
                    with _open_db(self.db_path) as conn:
                        total = conn.execute(
                            "SELECT COUNT(*) FROM local_product_images WHERE local_product_id = ?",
                            (lp_id,),
                        ).fetchone()[0]
                    entry["master_total_before"] = int(total)
                    if total >= 5:
                        deleted = self.delete_gallery_fifth_image(lp_id)
                        entry["deleted"] = deleted
                        if deleted.get("deleted"):
                            sort_order = deleted.get("sort_order")
                    else:
                        entry["deleted"] = {"deleted": False, "skipped": True}
                gen = self.generate_hand_hold_secondary_for_local(lp_id, sort_order=sort_order)
                entry.update(gen)
                return entry
            except Exception as e:
                return {**entry, "ok": False, "error": str(e)[:300]}

        with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="hand-hold") as ex:
            futures = {ex.submit(_one, lp_id): lp_id for lp_id in all_ids}
            for fut in as_completed(futures):
                lp_id = futures[fut]
                try:
                    result = fut.result()
                except Exception as e:
                    result = {"local_product_id": lp_id, "ok": False, "error": str(e)[:300]}
                if result.get("ok"):
                    generated.append(result)
                else:
                    failed.append(result)

        for item in generated:
            lp_id = int(item["local_product_id"])
            with _open_db(self.db_path) as conn:
                listings = [
                    dict(r) for r in conn.execute(
                        """
                        SELECT id, shop, tiktok_product_id
                          FROM tkshop_products
                         WHERE local_product_id = ?
                           AND COALESCE(tiktok_product_id, '') != ''
                         ORDER BY shop, id
                        """,
                        (lp_id,),
                    ).fetchall()
                ]
            for listing in listings:
                lid = int(listing["id"])
                try:
                    n = self.sync_master_images_to_listing(lp_id, lid, replace=True)
                    synced.append({
                        "local_product_id": lp_id,
                        "listing_id": lid,
                        "shop": listing["shop"],
                        "images": n,
                    })
                    if push:
                        res = self.update_on_platform(lid)
                        rec = {
                            "local_product_id": lp_id,
                            "listing_id": lid,
                            "shop": listing["shop"],
                            "ok": bool(res.get("ok")),
                            "error": res.get("error_message") or "",
                        }
                        (pushed if rec["ok"] else failed).append(rec)
                except Exception as exc:
                    failed.append({
                        "local_product_id": lp_id,
                        "listing_id": lid,
                        "shop": listing.get("shop", ""),
                        "ok": False,
                        "error": str(exc)[:300],
                    })

        return {
            "dry_run": False,
            "workers": workers,
            "replace_fifth_ids": replace_ids,
            "append_fifth_ids": append_ids,
            "generated": generated,
            "synced": synced,
            "pushed": pushed,
            "failed": failed,
            "elapsed_s": round(time.time() - t0, 1),
        }

    def sync_master_images_to_listing(
        self, local_product_id: int, listing_id: int, *, replace: bool = False,
    ) -> int:
        """Copy ``local_product_images`` rows onto a shop listing.

        When ``replace`` is False, behaves like ``_snapshot_master_images_to_listing``
        (no-op if the listing already has any image rows). When ``replace`` is True,
        wipes the listing gallery and re-inserts from the master so every shop sees
        the same 1 main + N secondary set before a platform update.
        """
        with _open_db(self.db_path) as conn:
            lp_row = conn.execute(
                "SELECT local_product_id FROM tkshop_products WHERE id = ?",
                (listing_id,),
            ).fetchone()
            if not lp_row or int(lp_row["local_product_id"] or 0) != int(local_product_id):
                raise ValueError(
                    f"listing #{listing_id} is not linked to local_product #{local_product_id}"
                )
            if not replace:
                return self._snapshot_master_images_to_listing(local_product_id, listing_id)
            conn.execute(
                "DELETE FROM tkshop_product_images WHERE product_id = ?",
                (listing_id,),
            )
            now = int(time.time())
            cur = conn.execute(
                """
                INSERT INTO tkshop_product_images
                    (product_id, role, source, local_path,
                     tiktok_image_uri, sort_order, ai_prompt, created_at)
                SELECT ?, role, source, local_path,
                       '', sort_order, ai_prompt, ?
                  FROM local_product_images
                 WHERE local_product_id = ?
                 ORDER BY role = 'main' DESC, sort_order ASC, id ASC
                """,
                (listing_id, now, local_product_id),
            )
            conn.commit()
            return cur.rowcount

    def batch_fill_gallery_and_push(
        self,
        local_product_ids: Optional[list[int]] = None,
        *,
        target_total: int = 5,
        target_secondary: int = 4,
        dry_run: bool = False,
        fill: bool = True,
        push: bool = True,
        fill_mode: str = "ai",
    ) -> dict[str, Any]:
        """Ensure master + published listings have ``target_total`` images
        (1 main + ``target_secondary`` secondaries), then push updates.

        ``fill_mode``:
          - ``ai`` — generate missing images via AI (default)
          - ``duplicate`` — pad by copying existing secondaries (no AI)
          - ``ai_then_duplicate`` — AI first, duplicate pad if still short
          - ``regenerate`` — delete duplicate rows, AI-generate replacements
        """
        fill_mode = (fill_mode or "ai").strip().lower()
        if fill_mode not in {"ai", "duplicate", "ai_then_duplicate", "regenerate"}:
            fill_mode = "ai"
        target_total = max(2, int(target_total))
        target_secondary = max(1, min(int(target_secondary), target_total - 1))
        t0 = time.time()
        filled, synced, pushed, skipped, failed = [], [], [], [], []

        with _open_db(self.db_path) as conn:
            if local_product_ids:
                ids = [int(x) for x in local_product_ids]
            else:
                ids = [
                    int(r[0])
                    for r in conn.execute(
                        """
                        SELECT DISTINCT lp.id
                          FROM local_products lp
                          JOIN tkshop_products t ON t.local_product_id = lp.id
                         WHERE COALESCE(t.tiktok_product_id, '') != ''
                         ORDER BY lp.id
                        """
                    ).fetchall()
                ]

        removed_dupes: dict[str, int] = {}
        if fill_mode == "regenerate" and fill and not dry_run:
            removed_dupes = self.remove_duplicate_gallery_rows(ids)

        for lp_id in ids:
            with _open_db(self.db_path) as conn:
                master_total = conn.execute(
                    "SELECT COUNT(*) FROM local_product_images WHERE local_product_id = ?",
                    (lp_id,),
                ).fetchone()[0]
                master_sec = conn.execute(
                    "SELECT COUNT(*) FROM local_product_images "
                    "WHERE local_product_id = ? AND role = 'secondary'",
                    (lp_id,),
                ).fetchone()[0]
                dup_count = conn.execute(
                    "SELECT COUNT(*) FROM local_product_images "
                    "WHERE local_product_id = ? AND source = 'duplicate'",
                    (lp_id,),
                ).fetchone()[0]
                has_main = conn.execute(
                    "SELECT 1 FROM local_product_images "
                    "WHERE local_product_id = ? AND role = 'main' LIMIT 1",
                    (lp_id,),
                ).fetchone() is not None
                listings = [
                    dict(r)
                    for r in conn.execute(
                        """
                        SELECT id, shop, tiktok_product_id,
                               (SELECT COUNT(*) FROM tkshop_product_images ti
                                 WHERE ti.product_id = t.id) AS img_count
                          FROM tkshop_products t
                         WHERE t.local_product_id = ? AND COALESCE(t.tiktok_product_id, '') != ''
                         ORDER BY t.shop, t.id
                        """,
                        (lp_id,),
                    ).fetchall()
                ]

            need_sec = max(0, target_secondary - max(0, int(master_sec) - int(dup_count)))
            need_main = not has_main
            real_total = max(0, int(master_total) - int(dup_count))
            needs_work = (
                dup_count > 0
                or real_total < target_total
                or any(int(l["img_count"]) < target_total for l in listings)
            )
            push_only = push and not fill
            if not needs_work and not push_only:
                skipped.append({"local_product_id": lp_id, "reason": "already at target"})
                continue

            item = {
                "local_product_id": lp_id,
                "master_total": master_total,
                "master_total_after_delete": real_total,
                "duplicate_rows": int(dup_count),
                "need_main": need_main,
                "need_secondary": need_sec,
                "listings": [
                    {"id": l["id"], "shop": l["shop"], "img_count": l["img_count"]}
                    for l in listings
                ],
            }
            if dry_run:
                filled.append(item)
                continue

            try:
                if fill and (need_main or need_sec > 0 or master_total < target_total or dup_count > 0):
                    if fill_mode in {"ai", "ai_then_duplicate", "regenerate"} and (need_main or need_sec > 0):
                        if need_main:
                            self.auto_design_images_for_local_product(
                                lp_id, mode="main", replace_existing_ai=False,
                            )
                        if need_sec > 0:
                            self.auto_design_images_for_local_product(
                                lp_id,
                                mode="secondary",
                                secondary_count=need_sec,
                                replace_existing_ai=False,
                            )
                    if fill_mode in {"duplicate", "ai_then_duplicate"}:
                        with _open_db(self.db_path) as conn:
                            cur_total = conn.execute(
                                "SELECT COUNT(*) FROM local_product_images WHERE local_product_id = ?",
                                (lp_id,),
                            ).fetchone()[0]
                        if cur_total < target_total:
                            dup = self.pad_master_gallery_to_target(
                                lp_id, target_total=target_total,
                            )
                            item["duplicated"] = dup
                    with _open_db(self.db_path) as conn:
                        item["master_total_after"] = conn.execute(
                            "SELECT COUNT(*) FROM local_product_images WHERE local_product_id = ?",
                            (lp_id,),
                        ).fetchone()[0]
                    filled.append(item)

                for listing in listings:
                    lid = int(listing["id"])
                    try:
                        n = self.sync_master_images_to_listing(
                            lp_id, lid, replace=True,
                        )
                        synced.append(
                            {"local_product_id": lp_id, "listing_id": lid,
                             "shop": listing["shop"], "images": n}
                        )
                        if push:
                            res = self.update_on_platform(lid)
                            rec = {
                                "local_product_id": lp_id,
                                "listing_id": lid,
                                "shop": listing["shop"],
                                "ok": bool(res.get("ok")),
                                "error": res.get("error_message") or "",
                            }
                            (pushed if rec["ok"] else failed).append(rec)
                    except Exception as exc:
                        failed.append({
                            "local_product_id": lp_id,
                            "listing_id": lid,
                            "shop": listing["shop"],
                            "error": str(exc)[:300],
                        })
            except Exception as exc:
                failed.append({
                    "local_product_id": lp_id,
                    "error": str(exc)[:300],
                })

        return {
            "dry_run": dry_run,
            "fill_mode": fill_mode,
            "removed_duplicates": removed_dupes,
            "target_total": target_total,
            "target_secondary": target_secondary,
            "filled": filled,
            "synced": synced,
            "pushed": pushed,
            "skipped": skipped,
            "failed": failed,
            "elapsed_s": round(time.time() - t0, 1),
        }

    def batch_publish_to_shop(
        self,
        product_ids: list[int],
        target_shop: str,
        *,
        auto_activate: bool = True,
        auto_fix: bool = True,
        max_workers: int = 2,
    ) -> dict[str, Any]:
        """Bulk-publish a list of local draft products to ``target_shop``.

        For each input product:
          - If the pack already has a tkshop_products row on ``target_shop``
            that's been pushed to TikTok (``tiktok_product_id != ''``) → skip
            with ``reason='pack already listed on shop'`` so re-clicking the
            batch button doesn't double-publish.
          - If the product is already on ``target_shop`` (same row) but still
            a local draft → publish in place.
          - Else clone to ``target_shop`` (which regenerates the SKU with
            that shop's prefix and replicates images) and publish the clone.

        Health-check: products missing title / main image / category are
        skipped with reason='incomplete'; the operator sees the count in the
        summary and can fix them in the detail page.

        Returns ``{published:[{product_id, pack_id, source_id}],
                   skipped:[{product_id, reason}],
                   failed:[{product_id, error}],
                   elapsed_s}``.
        """
        from concurrent.futures import ThreadPoolExecutor, as_completed

        if not target_shop or target_shop not in TKSHOP_SHOPS:
            raise ValueError(
                f"unknown shop {target_shop!r}; configured shops: {TKSHOP_SHOPS}"
            )
        t0 = time.time()

        seen: set[int] = set()
        ordered: list[int] = []
        for pid in product_ids:
            try:
                p = int(pid)
            except (TypeError, ValueError):
                continue
            if p in seen:
                continue
            seen.add(p)
            ordered.append(p)

        published: list[dict] = []
        skipped: list[dict] = []
        failed: list[dict] = []

        # Pre-flight: bucket each input id into publish-in-place / clone-then-
        # publish / skip. Doing this in one DB pass keeps the worker pool's
        # SQL contention low.
        plan: list[tuple[int, int, Optional[int]]] = []  # (source_id, pack_id, publish_id_or_None_means_clone_first)
        with _open_db(self.db_path) as conn:
            for pid in ordered:
                src = conn.execute(
                    "SELECT id, pack_id, shop, title, category_id, tiktok_product_id "
                    "FROM tkshop_products WHERE id = ?",
                    (pid,),
                ).fetchone()
                if not src:
                    skipped.append({"product_id": pid, "reason": "product not found"})
                    continue
                pack_id = src["pack_id"]
                # Health: title + category are mandatory to even attempt publish.
                # Missing fields would either 400 immediately or land as broken
                # listings — better to surface "incomplete N" up-front.
                if not (src["title"] or "").strip():
                    skipped.append({"product_id": pid, "reason": "incomplete: missing title"})
                    continue
                if not (src["category_id"] or "").strip():
                    skipped.append({"product_id": pid, "reason": "incomplete: missing category"})
                    continue
                main_img = conn.execute(
                    "SELECT 1 FROM tkshop_product_images "
                    "WHERE product_id = ? AND role = 'main' LIMIT 1",
                    (pid,),
                ).fetchone()
                if not main_img:
                    skipped.append({"product_id": pid, "reason": "incomplete: no main image"})
                    continue
                # Already-listed guard: if any row for this pack on the target
                # shop has a TikTok id, treat this as a duplicate request.
                already = conn.execute(
                    "SELECT id FROM tkshop_products "
                    "WHERE pack_id = ? AND shop = ? AND tiktok_product_id != ''",
                    (pack_id, target_shop),
                ).fetchone()
                if already:
                    skipped.append({
                        "product_id": pid, "reason": "pack already listed on shop",
                        "existing_product_id": already["id"],
                    })
                    continue
                if (src["shop"] or TKSHOP_DEFAULT_SHOP) == target_shop:
                    plan.append((pid, pack_id, pid))  # publish in place
                else:
                    plan.append((pid, pack_id, None))  # clone first, publish clone

        publish_fn = self.publish_with_self_heal if auto_fix else self.publish

        def _one(item: tuple[int, int, Optional[int]]) -> dict:
            source_id, pack_id, publish_id = item
            try:
                if publish_id is None:
                    publish_id = self.clone_to_shop(source_id, target_shop)
                # Mark publishing so the products list reflects in-flight state
                # right away, mirroring the single-product publish route.
                with _open_db(self.db_path) as conn:
                    conn.execute(
                        "UPDATE tkshop_products SET publish_status = 'publishing' "
                        "WHERE id = ?",
                        (publish_id,),
                    )
                    conn.commit()
                kwargs: dict[str, Any] = {"auto_activate": auto_activate}
                if not auto_fix:
                    kwargs["dry_run"] = False
                publish_fn(publish_id, **kwargs)
                # Best-effort status reconcile in case the wrapper returned
                # without updating publish_status (some self-heal paths).
                try:
                    self.sync_one_status(publish_id)
                except Exception:
                    pass
                return {"ok": True, "product_id": publish_id,
                        "pack_id": pack_id, "source_id": source_id}
            except Exception as e:
                return {"ok": False, "product_id": publish_id or source_id,
                        "pack_id": pack_id, "error": f"{e}"[:300]}

        if plan:
            workers = max(1, min(max_workers, len(plan)))
            with ThreadPoolExecutor(max_workers=workers,
                                     thread_name_prefix="tkshop-publish") as ex:
                futures = {ex.submit(_one, item): item for item in plan}
                for fut in as_completed(futures):
                    res = fut.result()
                    if res.get("ok"):
                        published.append({k: v for k, v in res.items() if k != "ok"})
                    else:
                        failed.append({"product_id": res["product_id"],
                                       "error": res.get("error", "")})

        elapsed = round(time.time() - t0, 1)
        logger.info(
            "batch_publish_to_shop: target=%s requested=%d published=%d skipped=%d failed=%d in %.1fs",
            target_shop, len(ordered), len(published), len(skipped), len(failed), elapsed,
        )
        return {
            "published": published, "skipped": skipped, "failed": failed,
            "elapsed_s": elapsed, "target_shop": target_shop,
        }

    @staticmethod
    def _local_status_from_tiktok(remote: str) -> str:
        """Map TikTok-side status to our local vocab.

        Critically: TikTok 'DRAFT' must NOT map to local 'draft' — local 'draft'
        means "never sent to platform" (UI shows the create-product buttons).
        A product that exists on TikTok as a draft is 'draft_on_platform'.
        Mirrors the mapping in publish() so sync_one_status and the publish
        path agree.
        """
        ts = (remote or "").upper()
        if ts == "DRAFT":
            return "draft_on_platform"
        # TikTok uses several names for "under audit": INITIAL (just-created,
        # pre-review), PENDING / PENDING_REVIEW / REVIEWING (active review).
        if ts in ("INITIAL", "PENDING", "PENDING_REVIEW", "REVIEWING"):
            return "pending"
        # LIVE is the documented "on sale" value; ACTIVATE / ACTIVE show up
        # in real responses too — fold them all into 'published'.
        if ts in ("LIVE", "ACTIVATE", "ACTIVE"):
            return "published"
        # Audit rejection surfaces as FAILED / AUDIT_FAILED / REJECTED — fold
        # them into local 'failed' so the UI surfaces a fix-rejection action.
        if ts in ("FAILED", "AUDIT_FAILED", "REJECTED"):
            return "failed"
        # TikTok lumps seller-initiated deactivation and platform violation
        # suspension into the same SUSPENDED bucket — pick the friendlier
        # interpretation since both are recoverable via the activate API.
        if ts == "SUSPENDED":
            return "seller_deactivated"
        if ts == "DELETED":
            return "deleted"
        # Unrecognized / empty -> "" so sync callers know NOT to overwrite the
        # local status with junk (e.g. when the platform response omits the
        # status field). Callers that need a value supply their own fallback.
        return ""

    @staticmethod
    def _extract_audit_reasons(tt_data: dict) -> list[str]:
        """Pull human-readable rejection reasons out of a TikTok product
        detail payload. The wrapper's response shape isn't fully documented,
        so try several field names and flatten what we find.

        Returns an empty list when no audit-rejection info is present.
        """
        if not isinstance(tt_data, dict):
            return []
        out: list[str] = []
        # Common single-string field names.
        for k in ("audit_failed_reasons", "audit_fail_reason", "fail_reason",
                  "reject_reason", "rejection_reason", "qa_failed_reason"):
            v = tt_data.get(k)
            if isinstance(v, str) and v.strip():
                out.append(v.strip())
        # List-of-objects / list-of-strings field names. Items may be
        # ``{"reason": "...", "field": "title"}`` or just strings.
        for k in ("audit_failed_reasons", "audit_failed_list",
                  "reject_reasons", "rejection_reasons"):
            v = tt_data.get(k)
            if isinstance(v, list):
                for item in v:
                    if isinstance(item, str) and item.strip():
                        out.append(item.strip())
                    elif isinstance(item, dict):
                        msg = (item.get("reason") or item.get("message")
                               or item.get("desc") or item.get("text") or "").strip()
                        field = (item.get("field") or item.get("attribute") or "").strip()
                        if msg and field:
                            out.append(f"[{field}] {msg}")
                        elif msg:
                            out.append(msg)
        # Dedup while keeping order.
        seen: set[str] = set()
        uniq: list[str] = []
        for r in out:
            if r in seen:
                continue
            seen.add(r); uniq.append(r)
        return uniq

    def fix_rejection_and_resubmit(
        self,
        product_id: int,
        *,
        max_retries: int = 2,
    ) -> dict[str, Any]:
        """Self-heal for audit-rejected listings on the platform side.

        Flow:
          1. Fetch TikTok product detail and extract audit_failed_reasons.
          2. If reasons exist, feed them into the existing self-heal AI
             prompt as error_message + field_hints.
          3. Apply the AI patch to the local listing fields.
          4. Push the corrected content to TikTok via update_on_platform.
          5. Re-fetch detail to confirm the audit cleared; if still failing
             and we have retries left, loop with the latest reasons.

        Returns a summary dict with the trail of attempts so the UI can
        show what AI changed each round.
        """
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                "SELECT tiktok_product_id, shop, publish_status "
                "FROM tkshop_products WHERE id = ?",
                (product_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"product #{product_id} not found")
            tt_id = (row["tiktok_product_id"] or "").strip()
            shop_name = row["shop"] or TKSHOP_DEFAULT_SHOP
        if not tt_id:
            return {
                "ok": False,
                "error_message": "no tiktok_product_id — product never reached the platform",
            }

        cap = max(1, int(max_retries))
        attempts: list[dict[str, Any]] = []
        last_reasons: list[str] = []
        for i in range(cap):
            # 1) Pull current platform state.
            platform = self.get_platform_product(tt_id, shop=shop_name)
            data = platform.get("data") or {}
            reasons = self._extract_audit_reasons(data)
            remote_status = (data.get("status") or "").strip()
            attempts.append({
                "round": i + 1,
                "remote_status": remote_status,
                "reasons": reasons,
            })
            last_reasons = reasons
            if not reasons:
                # Either the listing isn't actually rejected, or the platform
                # doesn't surface reasons in this response. Bail out so we
                # don't pointlessly rewrite a passing listing.
                attempts[-1]["note"] = "no audit reasons returned — nothing to fix"
                break
            # 2) Ask AI to rewrite the failing fields.
            try:
                fix = self._ai_fix_payload(
                    product_id,
                    error_code="AUDIT_REJECTED",
                    error_message=" | ".join(reasons[:6])[:1200],
                    field_hints=[],
                )
            except Exception as exc:
                attempts[-1]["ai_error"] = str(exc)[:300]
                logger.warning(
                    "fix_rejection_and_resubmit: AI fix failed for #%d round %d: %s",
                    product_id, i + 1, exc,
                )
                break
            attempts[-1]["changed_fields"] = fix.get("changed_fields") or []
            attempts[-1]["rationale"] = fix.get("rationale") or ""
            if not fix.get("changed_fields"):
                attempts[-1]["note"] = "AI proposed no changes — stopping"
                break
            # 3) Push the patched listing to TikTok.
            try:
                push = self.update_on_platform(product_id)
                attempts[-1]["push_ok"] = bool(push.get("ok"))
                attempts[-1]["push_error"] = push.get("error_message") or ""
            except Exception as exc:
                attempts[-1]["push_error"] = str(exc)[:300]
                logger.warning(
                    "fix_rejection_and_resubmit: update_on_platform failed for #%d: %s",
                    product_id, exc,
                )
                break

        return {
            "ok": bool(attempts) and not last_reasons,
            "product_id": product_id,
            "attempts": attempts,
            "remaining_reasons": last_reasons,
        }

    def sync_one_status(self, product_id: int) -> dict[str, Any]:
        """Refresh the publish_status for a single product from TikTok."""
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                "SELECT tiktok_product_id, publish_status FROM tkshop_products WHERE id = ?",
                (product_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"product #{product_id} not found")
            tt_id = (row["tiktok_product_id"] or "").strip()
        if not tt_id:
            return {"ok": False, "error_message": "no tiktok_product_id"}
        shop_name = _get_product_shop(self.db_path, product_id)
        result = self.get_platform_product(tt_id, shop=shop_name)
        if not result.get("ok"):
            return {"ok": False, "error_message": result.get("error", "fetch failed")}
        data = result.get("data") or {}
        remote = (data.get("status") or "").strip()
        local = self._local_status_from_tiktok(remote)
        if not local:
            # Platform response had no recognizable status field — DON'T
            # overwrite the local value with junk; surface the raw shape so we
            # can see what the wrapper actually returned.
            logger.warning(
                "sync_one_status #%s: empty/unmapped remote status %r; raw=%s",
                product_id, remote, str(data)[:300],
            )
            return {"ok": False, "remote_status": remote,
                    "local_status": row["publish_status"], "previous": row["publish_status"],
                    "error_message": f"平台返回的状态无法识别（status={remote or '空'}）；已保留本地状态"}
        if local != (row["publish_status"] or ""):
            with _open_db(self.db_path) as conn:
                conn.execute(
                    "UPDATE tkshop_products SET publish_status = ? WHERE id = ?",
                    (local, product_id),
                )
                conn.commit()
        return {"ok": True, "remote_status": remote, "local_status": local,
                "previous": row["publish_status"]}

    def set_listing_live_status(
        self, product_id: int, *, activate: bool,
    ) -> dict[str, Any]:
        """上架 / 下架 a single listing on its shop via the middle layer.

        Calls ``POST {TKSHOP_SERVER_URL}/api/v1/tiktok/products/{id}/{activate
        |deactivate}?shop=``. Only works for a listing already on the platform
        (has tiktok_product_id). On success, refreshes publish_status from
        TikTok (falls back to an optimistic local value if the sync fails).
        Returns ``{ok, ...}``.
        """
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                "SELECT tiktok_product_id, shop, publish_status "
                "FROM tkshop_products WHERE id = ?",
                (product_id,),
            ).fetchone()
        if not row:
            return {"ok": False, "error_message": f"listing #{product_id} not found"}
        tt_id = (row["tiktok_product_id"] or "").strip()
        if not tt_id:
            return {"ok": False, "error_message": "尚未上架到平台(无 tiktok_product_id)"}
        shop_name = canonical_shop_key(row["shop"] or TKSHOP_DEFAULT_SHOP)
        action = "activate" if activate else "deactivate"
        endpoint = (
            f"{TKSHOP_SERVER_URL.rstrip('/')}"
            f"/api/v1/tiktok/products/{tt_id}/{action}"
        )
        try:
            resp = requests.post(
                endpoint, params={"shop": shop_name},
                timeout=TKSHOP_SERVER_TIMEOUT,
            )
        except requests.RequestException as e:
            return {"ok": False, "error_message": f"{type(e).__name__}: {e}"[:200]}
        try:
            rj = resp.json() if resp.text else {}
        except Exception:
            rj = {"_raw_text": resp.text[:500]}
        # ApiResponse envelope: {success, message, data}. Treat success=False
        # (or non-2xx) as failure.
        ok = bool(resp.ok)
        if isinstance(rj, dict) and "success" in rj:
            ok = ok and bool(rj.get("success"))
        if not ok:
            msg = ""
            if isinstance(rj, dict):
                msg = str(rj.get("message") or rj.get("error_message") or rj)
            return {"ok": False, "_http_status": resp.status_code,
                    "error_message": msg[:300] or f"HTTP {resp.status_code}"}
        # Refresh status from the platform; on failure, set optimistically so
        # the badge reflects the action immediately.
        try:
            self.sync_one_status(product_id)
        except Exception:
            with _open_db(self.db_path) as conn:
                conn.execute(
                    "UPDATE tkshop_products SET publish_status = ? WHERE id = ?",
                    ("activate" if activate else "seller_deactivated", product_id),
                )
                conn.commit()
        return {"ok": True, "product_id": product_id, "shop": shop_name,
                "action": action}

    def batch_set_shop_listing_status(
        self, local_product_ids: list[int], shop: str, *, activate: bool,
    ) -> dict[str, Any]:
        """一键 上架 / 下架 selected masters' listings on one shop.

        For each local product, find its listing on ``shop`` (canonical-key
        aware) that is already on the platform, and activate/deactivate it.
        Masters with no platform listing on that shop are skipped. Returns a
        summary ``{ok, action, shop, done, skipped, failed}``.
        """
        target_shop = canonical_shop_key(shop)
        if target_shop not in TKSHOP_SHOPS:
            raise ValueError(
                f"unknown shop {shop!r}; configured shops: {TKSHOP_SHOPS}"
            )
        action = "activate" if activate else "deactivate"
        ids: list[int] = []
        seen: set[int] = set()
        for v in local_product_ids:
            try:
                iv = int(v)
            except (TypeError, ValueError):
                continue
            if iv not in seen:
                seen.add(iv)
                ids.append(iv)
        done, skipped, failed = [], [], []
        for lp_id in ids:
            with _open_db(self.db_path) as conn:
                rows = conn.execute(
                    "SELECT id, shop, tiktok_product_id FROM tkshop_products "
                    "WHERE local_product_id = ? AND tiktok_product_id != '' "
                    "ORDER BY id ASC",
                    (lp_id,),
                ).fetchall()
            listing_id = None
            for r in rows:
                if canonical_shop_key(r["shop"]) == target_shop:
                    listing_id = r["id"]
                    break
            if listing_id is None:
                skipped.append({"local_product_id": lp_id,
                                "reason": "该店铺无已上架 listing"})
                continue
            res = self.set_listing_live_status(listing_id, activate=activate)
            if res.get("ok"):
                done.append({"local_product_id": lp_id, "product_id": listing_id})
            else:
                failed.append({"local_product_id": lp_id, "product_id": listing_id,
                               "error": res.get("error_message", "")})
        logger.info(
            "batch_set_shop_listing_status: shop=%s action=%s done=%d skipped=%d failed=%d",
            target_shop, action, len(done), len(skipped), len(failed),
        )
        return {"ok": True, "action": action, "shop": target_shop,
                "done": done, "skipped": skipped, "failed": failed}

    def batch_shelf_on_shop(
        self, local_product_ids: list[int], shop: str,
        *, auto_fix: bool = True,
    ) -> dict[str, Any]:
        """Smart 上架 for one shop — the single entry point that puts products
        live, whether or not they already exist on the platform:

          - master has NO platform listing on the shop  -> publish (create +
            activate) via ``batch_publish_local_to_shop``
          - has a listing that is on-platform but not live (deactivated/draft)
            -> activate it
          - already live -> skip

        Merges the old separate "批量发布" into 上架. Returns a summary
        ``{ok, action:'shelf_on', shop, published, activated, skipped, failed}``.
        """
        target_shop = canonical_shop_key(shop)
        if target_shop not in TKSHOP_SHOPS:
            raise ValueError(
                f"unknown shop {shop!r}; configured shops: {TKSHOP_SHOPS}"
            )
        ids: list[int] = []
        seen: set[int] = set()
        for v in local_product_ids:
            try:
                iv = int(v)
            except (TypeError, ValueError):
                continue
            if iv not in seen:
                seen.add(iv)
                ids.append(iv)

        to_publish: list[int] = []          # no platform listing yet
        to_activate: list[tuple[int, int]] = []  # (lp_id, listing_id) deactivated/draft
        skipped: list[dict] = []
        for lp_id in ids:
            with _open_db(self.db_path) as conn:
                rows = conn.execute(
                    "SELECT id, shop, tiktok_product_id, publish_status "
                    "FROM tkshop_products WHERE local_product_id = ? "
                    "ORDER BY id ASC",
                    (lp_id,),
                ).fetchall()
            platform_listing = None
            for r in rows:
                if canonical_shop_key(r["shop"]) == target_shop \
                        and (r["tiktok_product_id"] or "").strip():
                    platform_listing = r
                    break
            if platform_listing is None:
                # No live listing on the platform yet -> first-time publish.
                to_publish.append(lp_id)
            elif (platform_listing["publish_status"] or "") == "activate":
                skipped.append({"local_product_id": lp_id, "reason": "已在线销售中"})
            else:
                to_activate.append((lp_id, platform_listing["id"]))

        published, activated, failed = [], [], []
        # First-time publishes go through the master-first publish pipeline
        # (creates listing, snapshots images, publishes + activates).
        if to_publish:
            pub = self.batch_publish_local_to_shop(
                to_publish, target_shop, auto_activate=True, auto_fix=auto_fix,
            )
            published = pub.get("published", [])
            for s in pub.get("skipped", []):
                skipped.append(s)
            for f in pub.get("failed", []):
                failed.append(f)
        # Re-activate existing deactivated/draft listings.
        for lp_id, listing_id in to_activate:
            res = self.set_listing_live_status(listing_id, activate=True)
            if res.get("ok"):
                activated.append({"local_product_id": lp_id, "product_id": listing_id})
            else:
                failed.append({"local_product_id": lp_id, "product_id": listing_id,
                               "error": res.get("error_message", "")})
        logger.info(
            "batch_shelf_on_shop: shop=%s published=%d activated=%d skipped=%d failed=%d",
            target_shop, len(published), len(activated), len(skipped), len(failed),
        )
        return {"ok": True, "action": "shelf_on", "shop": target_shop,
                "published": published, "activated": activated,
                "skipped": skipped, "failed": failed}

    def sync_statuses(self) -> dict[str, Any]:
        """C.4 — query multi-channel-api for status of every published product.

        Hits ``GET {TKSHOP_SERVER_URL}/api/v1/tiktok/products/{tiktok_product_id}``
        which returns an ``ApiResponse`` envelope ``{success, message, data}``.
        TikTok status (``data.status``) values: DRAFT, PENDING, LIVE,
        SUSPENDED, DELETED. We normalize to lowercase and write to
        ``publish_status``.

        Returns ``{'checked': n, 'updated': n, 'errors': [...]}``.
        """
        with _open_db(self.db_path) as conn:
            rows = conn.execute(
                """
                SELECT id, tiktok_product_id, publish_status, shop
                  FROM tkshop_products
                 WHERE tiktok_product_id != ''
                """,
            ).fetchall()
            # Any product that exists on the platform (has a tiktok_product_id)
            # is syncable — including ones locally marked 'failed', which may
            # actually have gone live on TikTok (the old filter skipped them).
        if not rows:
            return {"checked": 0, "updated": 0, "errors": []}

        updated, errors = 0, []
        for r in rows:
            shop_name = (r["shop"] or "").strip() or TKSHOP_DEFAULT_SHOP
            url = (
                f"{TKSHOP_SERVER_URL.rstrip('/')}"
                f"/api/v1/tiktok/products/{r['tiktok_product_id']}"
            )
            try:
                resp = requests.get(url, params={"shop": shop_name},
                                    timeout=TKSHOP_SERVER_TIMEOUT)
                if not resp.ok:
                    errors.append({"product_id": r["id"], "error": f"HTTP {resp.status_code}"})
                    continue
                body = resp.json() if resp.text else {}
                # ApiResponse envelope: {success, message, data: {...}}
                data = body.get("data") if isinstance(body, dict) else None
                if not isinstance(data, dict):
                    data = body if isinstance(body, dict) else {}
                remote = (data.get("status") or "").strip()
                local = self._local_status_from_tiktok(remote)
                if not local:
                    logger.warning(
                        "sync_statuses #%s: empty/unmapped remote status %r; raw=%s",
                        r["id"], remote, str(data)[:200],
                    )
                    errors.append({"product_id": r["id"],
                                   "error": f"status unrecognized: {remote or 'empty'}"})
                    continue
                if local != (r["publish_status"] or ""):
                    with _open_db(self.db_path) as conn:
                        conn.execute(
                            "UPDATE tkshop_products SET publish_status = ? WHERE id = ?",
                            (local, r["id"]),
                        )
                        conn.commit()
                    updated += 1
            except requests.RequestException as e:
                errors.append({"product_id": r["id"], "error": f"{type(e).__name__}: {e}"[:200]})
        return {"checked": len(rows), "updated": updated, "errors": errors}

    def list_publish_logs(self, product_id: int) -> list[dict]:
        with _open_db(self.db_path) as conn:
            rows = conn.execute(
                """
                SELECT id, attempt_idx, api_endpoint, success,
                       error_code, error_message, response_payload, created_at
                  FROM tkshop_publish_logs
                 WHERE product_id = ?
                 ORDER BY id DESC
                """, (product_id,),
            ).fetchall()
        out = []
        for r in rows:
            d = dict(r)
            # Best-effort parse of response_payload so the UI can read field_hints.
            try:
                d["response_obj"] = json.loads(d.get("response_payload") or "{}")
            except Exception:
                d["response_obj"] = {}
            out.append(d)
        return out

    # ------------------------------------------------------------------
    # Read
    # ------------------------------------------------------------------

    def list_products(
        self,
        *,
        pack_id: Optional[int] = None,
        publish_status: Optional[str] = None,
        q: Optional[str] = None,
        shop: Optional[str] = None,
        local_only: bool = False,
        on_platform_only: bool = False,
        limit: int = 100,
    ) -> tuple[list[dict], int]:
        clauses, params = [], []
        if pack_id is not None:
            clauses.append("pr.pack_id = ?"); params.append(pack_id)
        if publish_status and publish_status != "all":
            clauses.append("pr.publish_status = ?"); params.append(publish_status)
        if shop and shop != "all":
            clauses.append("pr.shop = ?"); params.append(shop)
        # local_only: rows that never made it to TikTok (no tiktok_product_id).
        # This powers the new top-level 「本地产品」tab — the draft pool that
        # operators batch-publish from. on_platform_only is the inverse for
        # the per-shop tabs.
        if local_only:
            clauses.append("(pr.tiktok_product_id IS NULL OR pr.tiktok_product_id = '')")
        elif on_platform_only:
            clauses.append("pr.tiktok_product_id IS NOT NULL AND pr.tiktok_product_id != ''")
        if q:
            kw = f"%{q.strip()}%"
            clauses.append(
                "(pr.title LIKE ? OR p.display_name LIKE ? "
                "OR pr.seller_sku LIKE ? OR pr.tiktok_product_id LIKE ?)"
            )
            params.extend([kw, kw, kw, kw])
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        with _open_db(self.db_path) as conn:
            total = conn.execute(
                f"""SELECT COUNT(*) FROM tkshop_products pr
                    LEFT JOIN packs p ON p.id = pr.pack_id
                    {where}""",
                tuple(params),
            ).fetchone()[0]
            rows = conn.execute(
                f"""
                SELECT pr.id, pr.pack_id, pr.shop, pr.tiktok_product_id, pr.tiktok_sku_id,
                       pr.seller_sku, pr.title,
                       pr.publish_status, pr.created_at, pr.published_at,
                       pr.category_id,
                       p.display_name AS pack_name, p.cover_image_path AS pack_cover
                  FROM tkshop_products pr
             LEFT JOIN packs           p ON p.id = pr.pack_id
                  {where}
                 ORDER BY pr.id DESC
                 LIMIT ?
                """,
                tuple(params) + (limit,),
            ).fetchall()
        return [dict(r) for r in rows], total

    def get_pack_shop_listings(
        self, pack_ids: list[int],
    ) -> dict[int, dict[str, dict[str, Any]]]:
        """For each pack id, return a per-shop summary of its products.

        Output: ``{pack_id: {shop_name: {product_id, publish_status,
                                          tiktok_product_id}}}``.

        Used by the 「本地产品」 tab to render shop-listing badges next to
        each draft row (so the operator sees "this pack is already on shop A
        as #123, still missing from shop B" without leaving the page).
        Packs absent from the result have no products at all.
        """
        if not pack_ids:
            return {}
        # De-dup + coerce to ints — the caller may pass strings.
        clean: list[int] = []
        seen: set[int] = set()
        for p in pack_ids:
            try:
                i = int(p)
            except (TypeError, ValueError):
                continue
            if i in seen:
                continue
            seen.add(i)
            clean.append(i)
        if not clean:
            return {}
        ph = ",".join(["?"] * len(clean))
        out: dict[int, dict[str, dict[str, Any]]] = {}
        with _open_db(self.db_path) as conn:
            rows = conn.execute(
                f"""
                SELECT pack_id, shop, id, publish_status, tiktok_product_id
                  FROM tkshop_products
                 WHERE pack_id IN ({ph})
                 ORDER BY id ASC
                """,
                tuple(clean),
            ).fetchall()
        for r in rows:
            pid = r["pack_id"]
            shop = (r["shop"] or TKSHOP_DEFAULT_SHOP)
            out.setdefault(pid, {})[shop] = {
                "product_id": r["id"],
                "publish_status": r["publish_status"] or "draft",
                "tiktok_product_id": r["tiktok_product_id"] or "",
            }
        return out

    def export_products_rows(self, product_ids: list[int]) -> list[dict]:
        """Pull export-ready dicts for the given product ids. One row per
        product, with images joined as a list of local_path strings.
        Returns rows in input order (skips ids that don't exist).
        """
        if not product_ids:
            return []
        ph = ",".join(["?"] * len(product_ids))
        with _open_db(self.db_path) as conn:
            prods = conn.execute(
                f"""
                SELECT pr.id, pr.pack_id, pr.shop, pr.tiktok_product_id, pr.tiktok_sku_id,
                       pr.seller_sku, pr.title, pr.description_html,
                       pr.selling_points, pr.keywords, pr.category_id,
                       pr.default_template_json,
                       pr.publish_status, pr.created_at, pr.published_at,
                       p.display_name AS pack_name, p.pack_uid, p.total_stickers
                  FROM tkshop_products pr
             LEFT JOIN packs           p ON p.id = pr.pack_id
                 WHERE pr.id IN ({ph})
                """,
                tuple(product_ids),
            ).fetchall()
            order = {pid: i for i, pid in enumerate(product_ids)}
            prods = sorted([dict(r) for r in prods], key=lambda r: order.get(r["id"], 1e9))
            for pr in prods:
                imgs = conn.execute(
                    """SELECT id, role, source, local_path, sort_order, ai_prompt
                         FROM tkshop_product_images
                        WHERE product_id = ?
                        ORDER BY role = 'main' DESC, sort_order ASC, id ASC""",
                    (pr["id"],),
                ).fetchall()
                pr["images"] = [dict(i) for i in imgs]
                # Decode JSON columns for spreadsheet display
                try:
                    pr["selling_points_list"] = json.loads(pr["selling_points"] or "[]")
                except Exception:
                    pr["selling_points_list"] = []
                try:
                    pr["keywords_list"] = json.loads(pr["keywords"] or "[]")
                except Exception:
                    pr["keywords_list"] = []
        self._enrich_export_commerce_fields(prods)
        return prods

    def _enrich_export_commerce_fields(self, rows: list[dict]) -> None:
        """Fill sale/discount price, stock, and 后台关键词 for spreadsheet export."""
        for pr in rows:
            pr["search_keywords"] = _format_search_keywords(
                pr.get("keywords_list") or [],
            )
            tpl_sale, tpl_discount = _price_from_default_template(
                pr.get("default_template_json") or "{}",
            )
            pr["sale_price"] = tpl_sale or TKSHOP_DEFAULT_SALE_PRICE
            pr["discount_price"] = (
                tpl_discount or TKSHOP_DEFAULT_DISCOUNT_PRICE
            )
            pr["currency"] = "USD"
            pr["stock"] = TKSHOP_DEFAULT_QUANTITY

            tt_id = (pr.get("tiktok_product_id") or "").strip()
            if not tt_id:
                continue
            shop = (pr.get("shop") or "").strip() or TKSHOP_DEFAULT_SHOP
            try:
                result = self.get_platform_product(tt_id, shop=shop)
                if not result.get("ok"):
                    continue
                meta = _sku_meta_from_platform_data(result.get("data") or {})
                if meta.get("sale_price"):
                    pr["sale_price"] = meta["sale_price"]
                if meta.get("discount_price"):
                    pr["discount_price"] = meta["discount_price"]
                if meta.get("currency"):
                    pr["currency"] = meta["currency"]
                if meta.get("stock") is not None:
                    pr["stock"] = meta["stock"]
            except Exception as e:  # noqa: BLE001
                logger.debug(
                    "export commerce fetch failed for #%s: %s",
                    pr.get("id"), e,
                )

    def export_local_products_rows(self, local_product_ids: list[int]) -> list[dict]:
        """Export-ready dicts for master-catalog (local_products) rows.

        Mirrors the shape of :meth:`export_products_rows` so the same XLSX/ZIP
        builders work, but pulls from ``local_products`` / ``local_product_images``.
        Platform-specific columns (tiktok ids, publish status) are aggregated
        across each master's shop listings, since the master is shop-agnostic.
        Returns rows in input order (skips ids that don't exist).
        """
        if not local_product_ids:
            return []
        ph = ",".join(["?"] * len(local_product_ids))
        with _open_db(self.db_path) as conn:
            masters = conn.execute(
                f"""
                SELECT lp.id, lp.pack_id, lp.seller_sku, lp.title,
                       lp.description_html, lp.selling_points, lp.keywords,
                       lp.category_id, lp.default_template_json, lp.created_at,
                       p.display_name AS pack_name, p.pack_uid, p.total_stickers
                  FROM local_products lp
             LEFT JOIN packs         p ON p.id = lp.pack_id
                 WHERE lp.id IN ({ph})
                """,
                tuple(local_product_ids),
            ).fetchall()
            order = {pid: i for i, pid in enumerate(local_product_ids)}
            masters = sorted(
                [dict(r) for r in masters],
                key=lambda r: order.get(r["id"], 1e9),
            )
            for pr in masters:
                imgs = conn.execute(
                    """SELECT id, role, source, local_path, sort_order, ai_prompt
                         FROM local_product_images
                        WHERE local_product_id = ?
                        ORDER BY role = 'main' DESC, sort_order ASC, id ASC""",
                    (pr["id"],),
                ).fetchall()
                pr["images"] = [dict(i) for i in imgs]
                # Aggregate the per-shop listings into the platform columns.
                listings = conn.execute(
                    """SELECT shop, tiktok_product_id, tiktok_sku_id,
                              publish_status, published_at
                         FROM tkshop_products
                        WHERE local_product_id = ?
                        ORDER BY shop, id""",
                    (pr["id"],),
                ).fetchall()
                pr["tiktok_product_id"] = ", ".join(
                    (ls["tiktok_product_id"] or "").strip()
                    for ls in listings if (ls["tiktok_product_id"] or "").strip()
                )
                pr["tiktok_sku_id"] = ", ".join(
                    (ls["tiktok_sku_id"] or "").strip()
                    for ls in listings if (ls["tiktok_sku_id"] or "").strip()
                )
                pr["publish_status"] = "; ".join(
                    f"{ls['shop']}:{ls['publish_status'] or 'draft'}"
                    for ls in listings
                ) or "no_listing"
                pub_times = [int(ls["published_at"]) for ls in listings if ls["published_at"]]
                pr["published_at"] = max(pub_times) if pub_times else None
                try:
                    pr["selling_points_list"] = json.loads(pr["selling_points"] or "[]")
                except Exception:
                    pr["selling_points_list"] = []
                try:
                    pr["keywords_list"] = json.loads(pr["keywords"] or "[]")
                except Exception:
                    pr["keywords_list"] = []
        self._enrich_export_commerce_fields(masters)
        return masters

    def export_products_xlsx(
        self, product_ids: Optional[list[int]] = None, *, rows: Optional[list[dict]] = None,
    ) -> bytes:
        """XLSX with embedded main image + all metadata. One row per product.

        Pass ``rows`` to export precomputed dicts (e.g. master-catalog rows from
        :meth:`export_local_products_rows`); otherwise rows are pulled from
        ``tkshop_products`` for the given ``product_ids``.
        """
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.drawing.image import Image as XLImage

        rows = rows if rows is not None else self.export_products_rows(product_ids or [])
        wb = Workbook()
        ws = wb.active
        ws.title = "TKShop Products"

        headers = [
            "ID", "Main Image", "Title", "Seller SKU",
            "TikTok Product ID", "TikTok SKU ID", "Status",
            "Pack Name", "Pack UID", "Total Stickers",
            "Category ID", "Selling Points", "Keywords",
            "后台关键词", "零售价", "折扣价", "货币", "库存",
            "Image Count", "All Image Paths",
            "Created", "Published",
        ]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F46E5")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Reasonable column widths
        widths = [6, 22, 60, 22, 22, 22, 14, 32, 32, 12, 12, 60, 50, 50, 10, 10, 8, 8, 8, 60, 18, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

        for r_idx, pr in enumerate(rows, start=2):
            ws.row_dimensions[r_idx].height = 110
            ws.cell(row=r_idx, column=1, value=pr["id"])
            ws.cell(row=r_idx, column=3, value=pr.get("title") or "")
            ws.cell(row=r_idx, column=4, value=pr.get("seller_sku") or "")
            ws.cell(row=r_idx, column=5, value=pr.get("tiktok_product_id") or "")
            ws.cell(row=r_idx, column=6, value=pr.get("tiktok_sku_id") or "")
            ws.cell(row=r_idx, column=7, value=pr.get("publish_status") or "")
            ws.cell(row=r_idx, column=8, value=pr.get("pack_name") or "")
            ws.cell(row=r_idx, column=9, value=pr.get("pack_uid") or "")
            ws.cell(row=r_idx, column=10, value=pr.get("total_stickers") or 0)
            ws.cell(row=r_idx, column=11, value=pr.get("category_id") or "")
            ws.cell(row=r_idx, column=12,
                    value="\n".join(pr.get("selling_points_list") or []))
            ws.cell(row=r_idx, column=13,
                    value=", ".join(pr.get("keywords_list") or []))
            ws.cell(row=r_idx, column=14, value=pr.get("search_keywords") or "")
            ws.cell(row=r_idx, column=15, value=pr.get("sale_price") or "")
            ws.cell(row=r_idx, column=16, value=pr.get("discount_price") or "")
            ws.cell(row=r_idx, column=17, value=pr.get("currency") or "")
            ws.cell(row=r_idx, column=18, value=pr.get("stock") or "")
            imgs = pr.get("images") or []
            ws.cell(row=r_idx, column=19, value=len(imgs))
            ws.cell(row=r_idx, column=20,
                    value="\n".join(i.get("local_path") or "" for i in imgs))

            # Created / Published timestamps as readable strings
            from datetime import datetime as _dt
            ts = pr.get("created_at")
            ws.cell(row=r_idx, column=21,
                    value=_dt.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M") if ts else "")
            ts = pr.get("published_at")
            ws.cell(row=r_idx, column=22,
                    value=_dt.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M") if ts else "")

            # Wrap text cells
            for col in (3, 12, 13, 14, 20):
                ws.cell(row=r_idx, column=col).alignment = Alignment(
                    wrap_text=True, vertical="top",
                )

            # Embed main image (column B = 2)
            main_img = next((i for i in imgs if (i.get("role") or "") == "main"),
                            imgs[0] if imgs else None)
            if main_img and main_img.get("local_path"):
                ip = Path(main_img["local_path"])
                if ip.is_file():
                    try:
                        xl_img = XLImage(str(ip))
                        # Resize to fit within ~140x140 px in cell
                        target = 140
                        xl_img.width = target
                        xl_img.height = target
                        anchor_cell = f"B{r_idx}"
                        ws.add_image(xl_img, anchor_cell)
                    except Exception as e:  # noqa: BLE001
                        ws.cell(row=r_idx, column=2,
                                value=f"(image embed failed: {type(e).__name__})")
                else:
                    ws.cell(row=r_idx, column=2,
                            value=f"(missing: {main_img['local_path']})")
            else:
                ws.cell(row=r_idx, column=2, value="(no image)")

        # Header freeze
        ws.freeze_panes = "C2"

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def export_products_zip(
        self, product_ids: Optional[list[int]] = None, *, rows: Optional[list[dict]] = None,
    ) -> bytes:
        """ZIP containing products.csv + images/ folder with all product images.
        Image filenames: ``{seller_sku or product_id}__{role}_{sort_order}{ext}``.

        Pass ``rows`` to export precomputed dicts (e.g. master-catalog rows);
        otherwise rows are pulled from ``tkshop_products`` for ``product_ids``.
        """
        import csv
        from io import BytesIO, StringIO
        import zipfile

        rows = rows if rows is not None else self.export_products_rows(product_ids or [])
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            # CSV
            sio = StringIO()
            writer = csv.writer(sio)
            writer.writerow([
                "id", "title", "seller_sku", "tiktok_product_id", "tiktok_sku_id",
                "publish_status", "pack_name", "pack_uid", "total_stickers",
                "category_id", "selling_points", "keywords", "search_keywords",
                "sale_price", "discount_price", "currency", "stock",
                "image_files", "created_at", "published_at",
            ])
            for pr in rows:
                imgs = pr.get("images") or []
                key = (pr.get("seller_sku") or f"P{pr['id']}").strip()
                file_names = []
                for i in imgs:
                    src = i.get("local_path") or ""
                    if not src:
                        continue
                    sp = Path(src)
                    if not sp.is_file():
                        continue
                    ext = sp.suffix or ".png"
                    role = i.get("role") or "img"
                    so = i.get("sort_order") or 0
                    out_name = f"images/{key}__{role}_{so}{ext}"
                    try:
                        zf.write(str(sp), arcname=out_name)
                        file_names.append(out_name)
                    except Exception:  # noqa: BLE001
                        pass
                writer.writerow([
                    pr["id"], pr.get("title") or "", pr.get("seller_sku") or "",
                    pr.get("tiktok_product_id") or "", pr.get("tiktok_sku_id") or "",
                    pr.get("publish_status") or "", pr.get("pack_name") or "",
                    pr.get("pack_uid") or "", pr.get("total_stickers") or 0,
                    pr.get("category_id") or "",
                    " | ".join(pr.get("selling_points_list") or []),
                    ", ".join(pr.get("keywords_list") or []),
                    pr.get("search_keywords") or "",
                    pr.get("sale_price") or "",
                    pr.get("discount_price") or "",
                    pr.get("currency") or "",
                    pr.get("stock") or "",
                    " | ".join(file_names),
                    pr.get("created_at") or "", pr.get("published_at") or "",
                ])
            zf.writestr("products.csv", sio.getvalue())
        return buf.getvalue()

    def get_product(self, product_id: int) -> Optional[dict]:
        with _open_db(self.db_path) as conn:
            r = conn.execute(
                """
                SELECT pr.id, pr.pack_id, pr.shop, pr.tiktok_product_id, pr.tiktok_sku_id,
                       pr.seller_sku, pr.auto_fix_attempts, pr.last_fix_diff,
                       pr.detail_main_raw_text, pr.title,
                       pr.description_html, pr.selling_points,
                       pr.keywords, pr.category_id,
                       pr.default_template_json, pr.publish_status,
                       pr.created_at, pr.published_at,
                       pr.discount_percent, pr.discount_status,
                       pr.discount_activity_id, pr.discount_applied_at,
                       p.display_name AS pack_name, p.cover_image_path,
                       p.pack_uid, p.total_stickers
                  FROM tkshop_products pr
             LEFT JOIN packs           p ON p.id = pr.pack_id
                 WHERE pr.id = ?
                """, (product_id,),
            ).fetchone()
            if not r:
                return None
            d = dict(r)
            try:
                d["selling_points"] = json.loads(d.get("selling_points") or "[]")
            except Exception:
                d["selling_points"] = []
            try:
                d["keywords"] = json.loads(d.get("keywords") or "[]")
            except Exception:
                d["keywords"] = []
            d["images"] = self.list_product_images(product_id)
            d["publish_logs"] = self.list_publish_logs(product_id)
            d["auto_fix_default"] = TKSHOP_PUBLISH_AUTO_FIX_DEFAULT
        return d

    # ==================================================================
    # Local products (master catalog) — one row per pack, shop-agnostic.
    # tkshop_products rows are "listings" pointing back via
    # local_product_id and carry the shop-specific platform IDs + a
    # snapshot of what was last pushed to TikTok.
    # ==================================================================

    def get_or_create_local_product(self, pack_id: int) -> int:
        """Idempotent: return the local_products row id for ``pack_id``,
        creating an empty draft row if none exists. The master is created
        with empty fields; ``generate_local_product_detail`` fills them.
        """
        with _open_db(self.db_path) as conn:
            existing = conn.execute(
                "SELECT id FROM local_products WHERE pack_id = ?",
                (pack_id,),
            ).fetchone()
            if existing:
                return existing["id"]
            if not conn.execute(
                "SELECT 1 FROM packs WHERE id = ?", (pack_id,),
            ).fetchone():
                raise ValueError(f"pack #{pack_id} not found")
            now = int(time.time())
            cur = conn.execute(
                """
                INSERT INTO local_products
                    (pack_id, title, description_html, selling_points, keywords,
                     detail_main_raw_text, seller_sku, category_id,
                     default_template_json, created_at, updated_at)
                VALUES (?, '', '', '[]', '[]', '', '', '928016', '{}', ?, ?)
                """,
                (pack_id, now, now),
            )
            conn.commit()
            return int(cur.lastrowid)

    def get_local_product(self, local_product_id: int) -> Optional[dict]:
        """Master + pack join + parsed JSON columns + image list + listing
        summary (which shops it's already on)."""
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                """
                SELECT lp.*, p.pack_uid, p.display_name AS pack_name,
                       p.total_stickers, p.cover_image_path AS pack_cover,
                       s.style_anchor, s.palette, s.pack_archetype
                  FROM local_products lp
                  JOIN packs p          ON p.id = lp.pack_id
             LEFT JOIN pack_series s    ON s.id = p.series_id
                 WHERE lp.id = ?
                """,
                (local_product_id,),
            ).fetchone()
            if not row:
                return None
            d = dict(row)
            try:
                d["selling_points"] = json.loads(d.get("selling_points") or "[]")
            except Exception:
                d["selling_points"] = []
            try:
                d["keywords"] = json.loads(d.get("keywords") or "[]")
            except Exception:
                d["keywords"] = []
            d["images"] = self.list_local_product_images(local_product_id)
            # Listing summary: which shops have a tkshop_products row, and
            # whether each is published.
            listings = conn.execute(
                """
                SELECT id, shop, publish_status, tiktok_product_id,
                       tiktok_sku_id, seller_sku, created_at, published_at
                  FROM tkshop_products
                 WHERE local_product_id = ?
                 ORDER BY id ASC
                """,
                (local_product_id,),
            ).fetchall()
            d["listings"] = [dict(r) for r in listings]
            master_updated = int(d.get("updated_at") or 0)
            live_ids = [
                ls["id"] for ls in d["listings"]
                if (ls.get("tiktok_product_id") or "").strip()
            ]
            push_map = _fetch_last_success_push_by_product(conn, live_ids)
            needs_any = False
            for ls in d["listings"]:
                last_push = push_map.get(ls["id"], 0)
                ls["last_pushed_at"] = last_push or None
                ls["needs_push"] = _listing_needs_platform_push(
                    master_updated_at=master_updated,
                    tiktok_product_id=ls.get("tiktok_product_id") or "",
                    last_success_push_at=last_push,
                    published_at=int(ls.get("published_at") or 0),
                )
                if ls["needs_push"]:
                    needs_any = True
            d["needs_push"] = needs_any
        return d

    def list_local_products(
        self,
        *,
        pack_id: Optional[int] = None,
        q: Optional[str] = None,
        limit: int = 100,
    ) -> tuple[list[dict], int]:
        """Master list. Each row carries a ``shops`` dict
        ``{shop: {product_id, publish_status, tiktok_product_id}}`` for the
        badges on the 本地产品 tab."""
        clauses, params = [], []
        if pack_id is not None:
            clauses.append("lp.pack_id = ?"); params.append(pack_id)
        if q:
            kw = f"%{q.strip()}%"
            clauses.append(
                "(lp.title LIKE ? OR p.display_name LIKE ? "
                "OR lp.seller_sku LIKE ?)"
            )
            params.extend([kw, kw, kw])
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        with _open_db(self.db_path) as conn:
            total = conn.execute(
                f"""SELECT COUNT(*) FROM local_products lp
                    LEFT JOIN packs p ON p.id = lp.pack_id
                    {where}""",
                tuple(params),
            ).fetchone()[0]
            rows = conn.execute(
                f"""
                SELECT lp.id, lp.pack_id, lp.title, lp.seller_sku,
                       lp.category_id, lp.default_quantity, lp.default_price,
                       lp.created_at, lp.updated_at,
                       p.display_name AS pack_name, p.pack_uid,
                       p.cover_image_path AS pack_cover,
                       (SELECT COUNT(*) FROM tkshop_products t
                          WHERE t.local_product_id = lp.id) AS listing_count,
                       (SELECT COUNT(*) FROM local_product_images li
                          WHERE li.local_product_id = lp.id) AS image_count,
                       (SELECT sp.sync_status FROM shopify_products sp
                          WHERE sp.local_product_id = lp.id) AS shopify_sync_status
                  FROM local_products lp
             LEFT JOIN packs           p ON p.id = lp.pack_id
                  {where}
                 ORDER BY lp.updated_at DESC, lp.id DESC
                 LIMIT ?
                """,
                tuple(params) + (limit,),
            ).fetchall()
            out: list[dict] = []
            ids = [r["id"] for r in rows]
            # Bulk-fetch listing summaries so we don't N+1 the badge query.
            listings_by_lp: dict[int, dict[str, dict[str, Any]]] = {}
            if ids:
                ph = ",".join(["?"] * len(ids))
                listing_rows = conn.execute(
                    f"""SELECT local_product_id, shop, id, publish_status,
                              tiktok_product_id, published_at
                         FROM tkshop_products
                        WHERE local_product_id IN ({ph})""",
                    tuple(ids),
                ).fetchall()
                live_product_ids = [
                    r["id"] for r in listing_rows
                    if (r["tiktok_product_id"] or "").strip()
                ]
                push_map = _fetch_last_success_push_by_product(conn, live_product_ids)
                updated_by_id = {r["id"]: int(r["updated_at"] or 0) for r in rows}
                for r in listing_rows:
                    lp_id = r["local_product_id"]
                    shop = r["shop"] or TKSHOP_DEFAULT_SHOP
                    last_push = push_map.get(r["id"], 0)
                    needs_push = _listing_needs_platform_push(
                        master_updated_at=updated_by_id.get(lp_id, 0),
                        tiktok_product_id=r["tiktok_product_id"] or "",
                        last_success_push_at=last_push,
                        published_at=int(r["published_at"] or 0),
                    )
                    listings_by_lp.setdefault(lp_id, {})[shop] = {
                        "product_id": r["id"],
                        "publish_status": r["publish_status"] or "draft",
                        "tiktok_product_id": r["tiktok_product_id"] or "",
                        "needs_push": needs_push,
                        "last_pushed_at": last_push or None,
                    }
            for r in rows:
                d = dict(r)
                d["shops"] = listings_by_lp.get(d["id"], {})
                d["needs_push"] = any(
                    info.get("needs_push") for info in d["shops"].values()
                )
                d["has_live_listing"] = any(
                    (info.get("tiktok_product_id") or "").strip()
                    for info in d["shops"].values()
                )
                out.append(d)
        return out, total

    def update_local_product_detail(
        self,
        local_product_id: int,
        *,
        title: Optional[str] = None,
        description_html: Optional[str] = None,
        selling_points: Optional[list[str]] = None,
        keywords: Optional[list[str]] = None,
        category_id: Optional[str] = None,
        seller_sku: Optional[str] = None,
        detail_main_raw_text: Optional[str] = None,
    ) -> bool:
        """Patch master fields. Bumps ``updated_at`` so listings can show
        "master is newer than last push"."""
        sets, params = [], []
        if title is not None:
            sets.append("title = ?"); params.append(_clamp_title(title))
        if description_html is not None:
            sets.append("description_html = ?"); params.append(description_html)
        if selling_points is not None:
            sets.append("selling_points = ?")
            params.append(json.dumps(selling_points, ensure_ascii=False))
        if keywords is not None:
            sets.append("keywords = ?")
            params.append(json.dumps(keywords, ensure_ascii=False))
        if category_id is not None:
            sets.append("category_id = ?"); params.append(category_id)
        if detail_main_raw_text is not None:
            sets.append("detail_main_raw_text = ?"); params.append(detail_main_raw_text)
        if not sets and seller_sku is None:
            return False
        with _open_db(self.db_path) as conn:
            if seller_sku is not None:
                clean = (seller_sku or "").strip().upper()[:25]
                # Master SKU intentionally matches its listings' SKUs (unified
                # cross-shop scheme). Don't check tkshop_products — masters and
                # their listings SHARE one SKU by design now.
                # Uniqueness is still enforced inside local_products below.
                while conn.execute(
                    "SELECT 1 FROM local_products "
                    "WHERE UPPER(seller_sku) = ? AND id != ?",
                    (clean, local_product_id),
                ).fetchone():
                    # Append disambiguator using the same scheme as
                    # _ensure_unique_seller_sku.
                    base = clean
                    n = 2
                    while True:
                        suffix = f"-{n}"
                        candidate = base[:25 - len(suffix)] + suffix
                        if not conn.execute(
                            "SELECT 1 FROM local_products WHERE UPPER(seller_sku) = ?",
                            (candidate,),
                        ).fetchone() and not conn.execute(
                            "SELECT 1 FROM tkshop_products WHERE UPPER(seller_sku) = ?",
                            (candidate,),
                        ).fetchone():
                            clean = candidate
                            break
                        n += 1
                sets.append("seller_sku = ?"); params.append(clean)
            sets.append("updated_at = ?"); params.append(int(time.time()))
            cur = conn.execute(
                f"UPDATE local_products SET {', '.join(sets)} WHERE id = ?",
                tuple(params) + (local_product_id,),
            )
            conn.commit()
            return cur.rowcount > 0

    def batch_set_local_default_quantity(
        self, local_product_ids: list[int], quantity: int,
    ) -> int:
        """Set ``default_quantity`` (local default stock) on many masters.

        Local-only: does NOT touch any live TikTok inventory. The value is
        used as the listing quantity on the NEXT push (see
        ``_build_publish_payload``). Returns the number of rows updated.
        """
        try:
            qty = max(0, int(quantity))
        except (TypeError, ValueError):
            raise ValueError(f"invalid quantity {quantity!r}")
        ids: list[int] = []
        for v in local_product_ids:
            try:
                ids.append(int(v))
            except (TypeError, ValueError):
                continue
        if not ids:
            return 0
        now = int(time.time())
        ph = ",".join(["?"] * len(ids))
        with _open_db(self.db_path) as conn:
            cur = conn.execute(
                f"UPDATE local_products "
                f"   SET default_quantity = ?, updated_at = ? "
                f" WHERE id IN ({ph})",
                tuple([qty, now] + ids),
            )
            conn.commit()
        logger.info(
            "batch_set_local_default_quantity: qty=%d applied to %d/%d masters",
            qty, cur.rowcount, len(ids),
        )
        return cur.rowcount

    def list_local_product_images(self, local_product_id: int) -> list[dict]:
        with _open_db(self.db_path) as conn:
            rows = conn.execute(
                """
                SELECT id, role, source, local_path, sort_order, ai_prompt, created_at
                  FROM local_product_images
                 WHERE local_product_id = ?
                 ORDER BY role = 'main' DESC, sort_order ASC, id ASC
                """,
                (local_product_id,),
            ).fetchall()
        return [dict(r) for r in rows]

    def add_local_product_image(
        self,
        local_product_id: int,
        *,
        local_path: str,
        role: str = "main",
        source: str = "manual",
        ai_prompt: str = "",
        sort_order: Optional[int] = None,
    ) -> int:
        """Insert one row into local_product_images. Caller has already
        persisted bytes via PackStore (or auto_design_images)."""
        with _open_db(self.db_path) as conn:
            if sort_order is None:
                # Append at end within the role bucket.
                row = conn.execute(
                    "SELECT COALESCE(MAX(sort_order), -1) + 1 AS s "
                    "FROM local_product_images "
                    "WHERE local_product_id = ? AND role = ?",
                    (local_product_id, role),
                ).fetchone()
                sort_order = row["s"]
            cur = conn.execute(
                """
                INSERT INTO local_product_images
                    (local_product_id, role, source, local_path,
                     sort_order, ai_prompt, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (local_product_id, role, source, local_path,
                 sort_order, ai_prompt, int(time.time())),
            )
            conn.commit()
            return int(cur.lastrowid)

    def update_local_product_image(
        self,
        image_id: int,
        *,
        role: Optional[str] = None,
        sort_order: Optional[int] = None,
    ) -> dict[str, Any]:
        sets, params = [], []
        if role is not None:
            sets.append("role = ?"); params.append(role)
        if sort_order is not None:
            sets.append("sort_order = ?"); params.append(int(sort_order))
        if not sets:
            return {"ok": False, "error": "nothing to update"}
        with _open_db(self.db_path) as conn:
            cur = conn.execute(
                f"UPDATE local_product_images SET {', '.join(sets)} WHERE id = ?",
                tuple(params) + (image_id,),
            )
            conn.commit()
            if cur.rowcount == 0:
                raise ValueError(f"local_product_image #{image_id} not found")
            return {"ok": True, "image_id": image_id}

    def delete_local_product_image(self, image_id: int) -> bool:
        """Remove image row. On-disk file is left in place — it may still be
        referenced by a listing's tkshop_product_images snapshot."""
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                """
                SELECT li.local_path, li.local_product_id, lp.pack_id
                  FROM local_product_images li
                  JOIN local_products lp ON lp.id = li.local_product_id
                 WHERE li.id = ?
                """,
                (image_id,),
            ).fetchone()
            if not row:
                return False
            cur = conn.execute(
                "DELETE FROM local_product_images WHERE id = ?", (image_id,),
            )
            if row["pack_id"]:
                self._clear_pack_cover_if_matches(
                    row["pack_id"], row["local_path"], conn,
                )
            conn.execute(
                "UPDATE local_products SET updated_at = ? WHERE id = ?",
                (int(time.time()), row["local_product_id"]),
            )
            conn.commit()
            return cur.rowcount > 0

    def auto_design_images_for_local_product(
        self,
        local_product_id: int,
        *,
        shop: Optional[str] = None,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """Master-level AI image generation.

        Resolves a tkshop_products listing for this master (reuses an
        existing draft on ``shop`` if available, else mints a fresh draft
        via ``create_product_from_pack``) and delegates to
        ``auto_design_images``. The existing mirror step inside that
        method copies AI outputs back into ``local_product_images``, so
        the master's image gallery sees the new images automatically.

        The draft listing is left in 'draft' status — never auto-published.
        It represents the operator's intent to eventually list on this
        shop; once they're happy with the AI output they can publish from
        the listing page or the master detail page.
        """
        replace_existing_ai = bool(kwargs.get("replace_existing_ai", True))
        secondary_count = int(kwargs.get("secondary_count", 3))
        mode = (kwargs.get("mode") or "all").strip().lower()
        append_one = (
            not replace_existing_ai
            and secondary_count == 1
            and mode in {"all", "secondary"}
            and fifth_image_style_is_hand_hold()
        )
        if append_one:
            gen = self.generate_hand_hold_secondary_for_local(local_product_id)
            return {
                "local_product_id": local_product_id,
                "listing_id": None,
                "shop": shop or TKSHOP_DEFAULT_SHOP,
                "product_id": None,
                "generated": 1 if gen.get("ok") else 0,
                "failed": 0 if gen.get("ok") else 1,
                "results": [gen],
                "append_only": True,
            }
        with _open_db(self.db_path) as conn:
            row = conn.execute(
                "SELECT pack_id FROM local_products WHERE id = ?",
                (local_product_id,),
            ).fetchone()
            if not row:
                raise ValueError(f"local_product #{local_product_id} not found")
            pack_id = row["pack_id"]
        target_shop = (shop or TKSHOP_DEFAULT_SHOP).strip() or TKSHOP_DEFAULT_SHOP
        if target_shop not in TKSHOP_SHOPS:
            raise ValueError(
                f"unknown shop {target_shop!r}; configured: {TKSHOP_SHOPS}"
            )
        # create_product_from_pack is idempotent: returns existing draft if
        # any, else creates one. Keeps draft status untouched.
        listing_id = self.create_product_from_pack(pack_id, target_shop)
        result = self.auto_design_images(listing_id, **kwargs)
        result["local_product_id"] = local_product_id
        result["listing_id"] = listing_id
        result["shop"] = target_shop
        return result

    def delete_local_product(self, local_product_id: int) -> dict[str, Any]:
        """Drop master + its images. Refuses if any listing still exists
        (operator must delete listings first — preserves audit trail)."""
        with _open_db(self.db_path) as conn:
            listings = conn.execute(
                "SELECT COUNT(*) FROM tkshop_products WHERE local_product_id = ?",
                (local_product_id,),
            ).fetchone()[0]
            if listings:
                raise ValueError(
                    f"local product #{local_product_id} still has {listings} "
                    "listing(s); delete those first"
                )
            conn.execute(
                "DELETE FROM local_product_images WHERE local_product_id = ?",
                (local_product_id,),
            )
            cur = conn.execute(
                "DELETE FROM local_products WHERE id = ?", (local_product_id,),
            )
            conn.commit()
            return {"ok": True, "deleted_rows": cur.rowcount}


# ---------------------------------------------------------------------------
# Module singleton
# ---------------------------------------------------------------------------

_svc: Optional[TKShopService] = None


def get_tkshop_service() -> TKShopService:
    global _svc
    if _svc is None:
        _svc = TKShopService()
    return _svc
